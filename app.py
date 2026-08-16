from __future__ import annotations
from dotenv import load_dotenv
load_dotenv()
import hashlib
import hmac as _hmac
import importlib, json, uuid, threading
import os
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Optional
from fastapi import BackgroundTasks, Body, FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from pydantic import BaseModel
from database import (
    init_db, listar_regras as db_listar_regras,
    inserir_regra, atualizar_regra, excluir_regra,
    substituir_todas_regras,
    listar_fila as db_listar_fila,
    buscar_item_fila, inserir_item_fila, atualizar_status_fila,
    stats_fila, limpar_invalidos_fila, reset_fila,
    ja_existe_pendente, get_config as db_get_config,
    migrar_json_legado,
)
from scheduler import iniciar_scheduler_background, parar_scheduler
from scbot import iniciar_scbot, parar_scbot, executar_ciclo as scbot_executar, carregar_status as scbot_status
from logging_config import configurar_logging
from auth import verificar_api_key

configurar_logging()

# Cache de links por SKU (evita chamadas repetidas à API)
_links_cache: dict = {}
_links_cache_ttl: dict = {}
_LINKS_CACHE_SECONDS = 3600  # 1 hora

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
PAGES_DIR = BASE_DIR / "pages"
REGRAS_PATH = DATA_DIR / "regras.json"
FILA_PATH = DATA_DIR / "fila_aprovacao.json"
CFG_PATH = DATA_DIR / "config.json"
LOG_PATH = DATA_DIR / "historico_precificacao.jsonl"
DATA_DIR.mkdir(exist_ok=True)
PAGES_DIR.mkdir(exist_ok=True)

app = FastAPI(title="Shinsei Pricing")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# ── Flags de controle de background ──────────────────────────────────
_scheduler_pausado: bool = False

# ── Estado da conferência ML em background ────────────────────────────
_conf_ml: dict = {
    "rodando": False, "concluido": False, "erro": None,
    "pagina": 0, "max_paginas": 20,
    "verificados": 0, "divergencias": 0, "sem_sku": 0, "erros": 0,
    "iniciado_em": None, "concluido_em": None, "resultado": None,
}

def _conf_ml_callback(pagina, max_paginas, verificados, divergencias, sem_sku, erros):
    _conf_ml.update({"pagina": pagina, "max_paginas": max_paginas,
                     "verificados": verificados, "divergencias": divergencias,
                     "sem_sku": sem_sku, "erros": erros})

def _rodar_conf_ml_bg():
    try:
        from ml_estoque_conferencia import conferir_estoques_ml
        from bling_client import BlingClient as _BC
        client = _BC()
        resultado = conferir_estoques_ml(
            client,
            max_paginas=_conf_ml["max_paginas"],
            progresso_cb=_conf_ml_callback,
        )
        _conf_ml.update({"rodando": False, "concluido": True,
                         "resultado": resultado, "concluido_em": datetime.utcnow().isoformat()})
    except Exception as e:
        logger.exception("Erro na conferência ML em background: %s", e)
        _conf_ml.update({"rodando": False, "concluido": True, "erro": str(e),
                         "concluido_em": datetime.utcnow().isoformat()})
from routes.batch import router as batch_router
from routes.ml_unificado import router as ml_router
from routes.bling import router as bling_page_router
from monitoring import router as monitoring_router
from routes.gmc import router as gmc_router
from routes.ads import router as ads_router
from routes.shopee import router as shopee_router, aplicar_preco_shopee_por_sku
from routes.amazon import router as amazon_router
from routes.estoque_sync import router as estoque_sync_router, _rebuild_caches_bg, sync_estoque_bling
from routes.shopify_webhooks import router as shopify_webhooks_router
from routes.ml_ads import router as ml_ads_router
from routes.bling_upload import router as bling_upload_router
app.include_router(batch_router)
from routes.mercado_livre import router as ml_page_router
app.include_router(ml_page_router)
app.include_router(ml_router)
app.include_router(bling_page_router)
app.include_router(bling_upload_router)
app.include_router(monitoring_router)
app.include_router(gmc_router)
app.include_router(ads_router)
app.include_router(shopee_router)
app.include_router(amazon_router)
app.include_router(estoque_sync_router)
app.include_router(shopify_webhooks_router)
app.include_router(ml_ads_router)
from routes.credentials import router as credentials_router
app.include_router(credentials_router)
try:
    from routes.frete import router as frete_router
    app.include_router(frete_router)
    logger.info("Motor de frete Shinsei registrado em /frete")
except Exception as _frete_exc:
    logger.warning("Motor de frete não carregado: %s", _frete_exc)

try:
    from dashboard_blueprint import router as dashboard_router
    app.include_router(dashboard_router)
    logger.info("Dashboard blueprint registrado em /dashboard")
except Exception as _dash_exc:
    logger.warning("Dashboard blueprint não carregado: %s", _dash_exc)

try:
    from routes.vinculo_ml import router as vinculo_ml_router
    app.include_router(vinculo_ml_router)
    logger.info("Ferramenta de vínculo ML registrada em /conferencia/ml/vincular")
except Exception as _vml_exc:
    logger.warning("Ferramenta de vínculo ML não carregada: %s", _vml_exc)

@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    return await verificar_api_key(request, call_next)

@app.on_event("startup")
def startup():
    init_db()
    migrar_json_legado()
    iniciar_scheduler_background()
    iniciar_scbot()
    # Inicia thread de auto-refresh de tokens (Bling, Shopee, ML)
    try:
        from token_autorefresh import iniciar as iniciar_autorefresh
        iniciar_autorefresh()
        logger.info("Token auto-refresh iniciado")
    except Exception as _e:
        logger.warning("Token auto-refresh não iniciado: %s", _e)
    # Pré-aquece caches de SKU (Shopify + ML) para sync de estoque em tempo real
    try:
        _rebuild_caches_bg()
        logger.info("Cache de SKU (estoque sync) sendo construído em background")
    except Exception as _e:
        logger.warning("Cache de SKU não iniciado: %s", _e)
    logger.info("Shinsei Pricing iniciado")

@app.on_event("shutdown")
def shutdown():
    parar_scheduler()
    parar_scbot()

def _optional_import(module_name: str):
    try: return importlib.import_module(module_name)
    except Exception: return None

pricing_module = _optional_import("pricing_engine_real") or _optional_import("pricing_engine")
if pricing_module is None:
    raise RuntimeError("pricing_engine_real.py ou pricing_engine.py não encontrado.")
montar_precificacao_bling: Optional[Callable[..., dict]] = getattr(pricing_module, "montar_precificacao_bling", None)
if montar_precificacao_bling is None:
    raise RuntimeError("Seu motor atual não expõe montar_precificacao_bling().")
_motor_anti_colapso_fn = getattr(pricing_module, "motor_anti_colapso", None)
_calcular_canais_fn = getattr(pricing_module, "calcular_canais", None)

product_intelligence_mod = _optional_import("product_intelligence")
_calculate_sie_fn = getattr(product_intelligence_mod, "calculate_sie", None) if product_intelligence_mod else None
_classify_sie_fn = getattr(product_intelligence_mod, "classify_sie", None) if product_intelligence_mod else None

bling_mod = _optional_import("bling_client")
BlingClient = getattr(bling_mod, "BlingClient", None) if bling_mod else None
bling_update_module = _optional_import("bling_update_engine")
aplicar_precos_multicanal = getattr(bling_update_module, "aplicar_precos_multicanal", None) if bling_update_module else None

DEFAULT_CFG = {"modo_aprovacao":"manual","fila_auto_ao_calcular":True,"peso_forca":0.4,"peso_equilibrio":0.4,"peso_lucro":0.2,"forcas_canais":{"Mercado Livre Classico":0.8,"Mercado Livre Premium":0.75,"Shopee":0.6,"Amazon":0.7,"Shein":0.55,"Shopify":0.65},"regra_estoque":{"ativo":False,"limite":2,"tipo":"percentual","valor":0}}
CANAL_ALIAS = {"Mercado Livre Classico":"mercado_livre_classico","Mercado Livre Premium":"mercado_livre_premium","Shopee":"shopee","Amazon":"amazon","Shein":"shein","Shopify":"shopify","Shopfy":"shopify"}

def _load_json(path: Path, default: Any):
    if not path.exists(): return default
    try: return json.loads(path.read_text(encoding="utf-8"))
    except Exception: return default

def _save_json(path: Path, data: Any) -> None:
    path.parent.mkdir(exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def _append_jsonl(path: Path, payload: dict) -> None:
    path.parent.mkdir(exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(payload, ensure_ascii=False) + "\n")

def _first_existing(path_options: list[Path]):
    for path in path_options:
        if path.exists(): return path
    return None

def carregar_regras(apenas_ativas: bool = False) -> list[dict]:
    try:
        db_regras = db_listar_regras(apenas_ativas=apenas_ativas)
        if db_regras:
            return db_regras
    except Exception:
        pass
    # Fallback: JSON (usado no Cloud Run onde o SQLite começa vazio)
    regras = _load_json(REGRAS_PATH, [])
    if not isinstance(regras, list): return []
    for r in regras:
        if isinstance(r, dict): r.setdefault("ativo", True)
    return [r for r in regras if isinstance(r, dict) and (r.get("ativo", True) or not apenas_ativas)]

def carregar_cfg() -> dict:
    data = _load_json(CFG_PATH, {})
    cfg = json.loads(json.dumps(DEFAULT_CFG))
    if isinstance(data, dict):
        cfg.update(data)
        cfg["forcas_canais"] = {**DEFAULT_CFG["forcas_canais"], **data.get("forcas_canais", {})}
        cfg["regra_estoque"] = {**DEFAULT_CFG["regra_estoque"], **data.get("regra_estoque", {})}
    return cfg

def carregar_fila() -> list[dict]:
    try:
        return db_listar_fila()
    except Exception:
        itens = _load_json(FILA_PATH, [])
        return itens if isinstance(itens, list) else []

def salvar_fila(itens: list[dict]) -> None:
    try:
        reset_fila()
        for item in itens:
            inserir_item_fila(item)
    except Exception:
        _save_json(FILA_PATH, itens)

def _fila_stats(itens: list[dict]) -> dict:
    stats = {"pendente":0,"aprovado":0,"rejeitado":0}
    for item in itens:
        status = str(item.get("status","pendente")).lower()
        if status in stats: stats[status] += 1
    stats["total"] = len(itens)
    return stats

def _normalizar_marketplaces(itens: list[dict]) -> dict[str, dict]:
    marketplaces = {}
    for item in itens or []:
        canal = item.get("canal") or "Canal"
        key = CANAL_ALIAS.get(canal, canal.lower().replace(" ","_"))
        marketplaces[key] = {
            "label": canal,
            "preco": item.get("preco_virtual") or item.get("preco_cheio") or item.get("preco_sugerido") or item.get("preco_promocional") or item.get("preco_final") or 0,
            "preco_promocional": item.get("preco_promocional") or item.get("preco_final") or 0,
            "lucro": item.get("lucro_liquido") or item.get("lucro") or 0,
            "margem": item.get("margem") or item.get("margem_liquida_percentual") or 0,
            "comissao": item.get("comissao") or 0,
            "frete": item.get("frete") or 0,
            "taxa_fixa": item.get("taxa_fixa") or 0,
            "imposto": item.get("imposto") or 0,
            "custo_total": item.get("custo_total") or item.get("custo_produto") or 0,
            "faixa_aplicada": item.get("faixa_aplicada") or "",
            "indice_final": item.get("indice_final") or 0,
            "raw": item,
        }
    return marketplaces

def _marketplaces_validos(marketplaces: dict) -> bool:
    return isinstance(marketplaces, dict) and len(marketplaces) > 0

def _diagnostico_preview(preview: dict) -> dict:
    aud = preview.get("auditoria") or {}
    produto = preview.get("produto") or {}
    marketplaces = preview.get("marketplaces") or {}
    if aud.get("erro"):
        codigo = str(aud.get("erro_codigo") or "").strip()
        if codigo not in {"custo_ausente", "peso_ausente", "composicao_sem_custo"}:
            erro_txt = str(aud.get("erro") or "").lower()
            if "sem peso" in erro_txt:
                codigo = "peso_ausente"
            elif "composição" in erro_txt or "composicao" in erro_txt:
                codigo = "composicao_sem_custo"
            elif "sem custo" in erro_txt:
                codigo = "custo_ausente"
            else:
                codigo = "erro_motor"
        return {"ok":False,"codigo":codigo,"mensagem":str(aud.get("erro")),"detalhe":aud.get("acao") or ""}
    if not (produto.get("codigo") or aud.get("sku")):
        return {"ok":False,"codigo":"sku_ausente","mensagem":"SKU ausente no retorno do Bling.","detalhe":"Verifique se o produto encontrado possui Código (SKU) cadastrado."}
    custo = float(aud.get("custo_usado") or 0)
    peso = float(aud.get("peso_usado") or 0)
    tipo_custo = str(aud.get("tipo_custo") or "").lower()
    componentes = aud.get("componentes_custo") or []
    if peso <= 0:
        return {"ok":False,"codigo":"peso_ausente","mensagem":"Produto sem peso no Bling.","detalhe":"Preencha o peso no produto ou use peso override no simulador."}
    if custo <= 0 and tipo_custo == "composicao":
        faltando = [c.get("sku") or str(c.get("id") or "-") for c in componentes if float(c.get("custo_unitario") or 0) <= 0]
        return {"ok":False,"codigo":"composicao_sem_custo","mensagem":"Composição sem custo válido nos componentes.","detalhe":("Componentes sem custo: " + ", ".join(faltando)) if faltando else "Nenhum componente retornou custo válido."}
    if custo <= 0:
        return {"ok":False,"codigo":"custo_ausente","mensagem":"Produto sem custo no estoque do Bling.","detalhe":"Preencha o preço de compra/custo do produto no estoque."}
    if not _marketplaces_validos(marketplaces):
        return {"ok":False,"codigo":"sem_canais","mensagem":"Nenhum canal válido foi calculado.","detalhe":"Verifique peso, custo e faixas da Aba2 para este produto."}
    return {"ok":True,"codigo":"preview_valido","mensagem":"Preview válido.","detalhe":""}

def _preview_valido(preview: dict):
    diag = _diagnostico_preview(preview)
    return bool(diag.get("ok")), str(diag.get("mensagem") or "Preview inválido.")

def _montar_item_fila(preview: dict, payload_original: dict | None = None) -> dict:
    aud = preview.get("auditoria") or {}
    produto = preview.get("produto") or {}
    agora = datetime.utcnow().isoformat()
    return {"id":str(uuid.uuid4()),"status":"pendente","criado_em":agora,"atualizado_em":agora,"sku":aud.get("sku") or produto.get("codigo") or "","nome":produto.get("nome") or "","produto_bling":produto,"marketplaces":preview.get("marketplaces") or {},"auditoria":aud,"payload_original":payload_original or preview.get("raw") or {},"historico_decisao":[],"resultado_aplicacao":None}

def _ja_existe_pendente_semelhante(itens: list[dict], sku: str, auditoria: dict) -> bool:
    custo = round(float(auditoria.get("custo_usado") or 0), 2)
    peso = round(float(auditoria.get("peso_usado") or 0), 3)
    for item in itens:
        if item.get("status") != "pendente": continue
        if str(item.get("sku") or "").strip() != str(sku).strip(): continue
        aud = item.get("auditoria") or {}
        if round(float(aud.get("custo_usado") or 0), 2) == custo and round(float(aud.get("peso_usado") or 0), 3) == peso:
            return True
    return False

class IntegracaoPayload(BaseModel):
    criterio: str = "sku"
    valor_busca: str = ""
    embalagem: Optional[float] = None  # None = usa embalagem_padrao da config
    imposto: float = 4
    quantidade: int = 1
    objetivo: str = "lucro_liquido"
    tipo_alvo: str = "percentual"
    valor_alvo: float = 30
    peso_override: float = 0
    score_config: Optional[dict] = None
    modo_aprovacao: str = "manual"
    modo_preco_virtual: str = "percentual_acima"
    acrescimo_percentual: float = 20
    acrescimo_nominal: float = 0
    preco_manual: float = 0
    arredondamento: str = "90"
    preco_compra_anterior_bling: float = 0

class DebugSkuPayload(BaseModel):
    sku: str

class AtualizacaoCampoBlingPayload(BaseModel):
    produto_id: int
    valor: float

class ImagemVariacaoPayload(BaseModel):
    produto_id: int
    variacao_id: int
    image_url: str

class BuscaProdutoPayload(BaseModel):
    nome: str
    limite: int = 10

FALLBACK_HTML = "<!DOCTYPE html><html lang='pt-BR'><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width, initial-scale=1.0'><title>Shinsei Pricing</title></head><body><h1>Shinsei Pricing</h1></body></html>"

def _prepare_product_patch(existing: dict) -> dict:
    patch = dict(existing) if isinstance(existing, dict) else {}
    if "data" in patch and isinstance(patch["data"], dict):
        patch = dict(patch["data"])
    return patch



@app.get("/auditoria/ml-sem-sku")
def auditoria_ml_sem_sku():
    """Lista anúncios ML ativos sem seller_custom_field (SKU) vinculado."""
    try:
        import json as _json, requests as _req
        from pathlib import Path
        _tokens_path = BASE_DIR / "data" / "ml_tokens.json"
        if not _tokens_path.exists():
            return {"ok": False, "erro": "Token ML não configurado.", "itens": [], "stats": {"sem_sku": 0}}
        _tokens = _json.loads(_tokens_path.read_text(encoding="utf-8"))
        _token = _tokens.get("access_token", "")
        _h = {"Authorization": f"Bearer {_token}"}
        
        # Busca anúncios ativos
        _sem_sku = []
        _offset = 0
        _limit = 100
        while _offset < 500:  # máximo 500 anúncios por vez
            _r = _req.get(
                f"https://api.mercadolibre.com/users/733168645/items/search",
                params={"status": "active", "limit": _limit, "offset": _offset},
                headers=_h, timeout=15
            )
            if _r.status_code != 200:
                break
            _data = _r.json()
            _items = _data.get("results", [])
            if not _items:
                break
            
            # Busca detalhes em lote (até 20 por vez)
            for i in range(0, len(_items), 20):
                _batch = _items[i:i+20]
                _ids = ",".join(_batch)
                _r2 = _req.get(
                    f"https://api.mercadolibre.com/items",
                    params={"ids": _ids, "attributes": "id,title,available_quantity,seller_custom_field,status,catalog_product_id"},
                    headers=_h, timeout=15
                )
                if _r2.status_code == 200:
                    for entry in _r2.json():
                        item = entry.get("body") or entry
                        if not item.get("seller_custom_field"):
                            _sem_sku.append({
                                "id": item.get("id"),
                                "titulo": item.get("title", "")[:60],
                                "estoque": item.get("available_quantity", 0),
                                "status": item.get("status"),
                                "catalogo": bool(item.get("catalog_product_id")),
                                "link": f"https://www.mercadolivre.com.br/anuncios/{item.get('id')}/editar",
                            })
            
            _offset += _limit
            if len(_items) < _limit:
                break
        
        return {
            "ok": True,
            "itens": _sem_sku,
            "stats": {"sem_sku": len(_sem_sku)},
        }
    except Exception as e:
        logger.error("Erro ao verificar ML sem SKU: %s", e)
        return {"ok": False, "erro": str(e), "itens": [], "stats": {"sem_sku": 0}}

@app.get("/", response_class=HTMLResponse)
def home():
    html_file = _first_existing([BASE_DIR / "index.html", PAGES_DIR / "simulador.html"])
    return HTMLResponse(html_file.read_text(encoding="utf-8")) if html_file else HTMLResponse(FALLBACK_HTML)

@app.get("/simulador", response_class=HTMLResponse)
def simulador_page():
    html_file = _first_existing([PAGES_DIR / "simulador.html", BASE_DIR / "index.html"])
    return HTMLResponse(html_file.read_text(encoding="utf-8")) if html_file else HTMLResponse(FALLBACK_HTML)

@app.get("/taxas/status")
def taxas_status():
    """Retorna quais fontes de taxa estão ativas (api real vs tabela fallback)."""
    status = {}
    # ML
    try:
        from ml_pricing_engine import _load_token as _ml_token
        status["ml"] = {"disponivel": bool(_ml_token()), "fonte": "api" if bool(_ml_token()) else "fallback"}
    except Exception:
        status["ml"] = {"disponivel": False, "fonte": "fallback"}
    # Amazon
    try:
        from amazon_client import AmazonClient
        ac = AmazonClient()
        ac._get_access_token()
        status["amazon"] = {"disponivel": True, "fonte": "sp_api"}
    except Exception:
        status["amazon"] = {"disponivel": False, "fonte": "tabela_br"}
    # Shopee
    try:
        from shopee_pricing_engine import _load_tokens as _sh_tokens
        toks = _sh_tokens()
        import os
        ok = bool(toks and toks.get("access_token") and os.getenv("SHOPEE_PARTNER_KEY"))
        status["shopee"] = {"disponivel": ok, "fonte": "pedidos_reais" if ok else "tabela_br"}
    except Exception:
        status["shopee"] = {"disponivel": False, "fonte": "tabela_br"}
    return status

@app.get("/fila", response_class=HTMLResponse)
def fila_page():
    html_file = PAGES_DIR / "fila.html"
    if html_file.exists():
        return HTMLResponse(html_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="pages/fila.html não encontrado.")

@app.get("/hub", response_class=HTMLResponse)
def hub_page():
    html_file = PAGES_DIR / "hub.html"
    return HTMLResponse(html_file.read_text(encoding="utf-8")) if html_file.exists() else HTMLResponse(FALLBACK_HTML)

# ── Produtos ──────────────────────────────────────────────────────────────────
@app.get("/produtos", response_class=HTMLResponse)
def produtos_page():
    return HTMLResponse((PAGES_DIR / "produtos.html").read_text(encoding="utf-8"))

import subprocess, threading, queue as _queue

_bot_processes: dict = {}
_bot_queues: dict = {}

def _run_bot_background(bot_id: str, script_path: str, q: "_queue.Queue"):
    try:
        proc = subprocess.Popen(
            ["python", script_path],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, encoding="utf-8", errors="replace",
            cwd=str(BASE_DIR)
        )
        _bot_processes[bot_id] = proc
        for line in proc.stdout:
            q.put({"line": line.rstrip()})
        proc.wait()
        ok = proc.returncode == 0
        q.put({"done": True, "ok": ok, "summary": {}})
    except Exception as e:
        q.put({"line": f"ERRO: {e}"})
        q.put({"done": True, "ok": False, "summary": {}})
    finally:
        _bot_processes.pop(bot_id, None)

from fastapi.responses import StreamingResponse
import json as _json

@app.get("/produtos/bots/{bot_id}/run")
def bot_run(bot_id: str):
    scripts = {
        "kit":              str(BASE_DIR / "bling_anexar_imagens_kit.py"),
        "import-shopify":   str(BASE_DIR / "bots" / "importar_shopify_bling.py"),
        "import-shopee":    str(BASE_DIR / "bots" / "importar_shopee_bling.py"),
        "import-ml":        str(BASE_DIR / "bots" / "importar_ml_bling.py"),
        "export-ml":        str(BASE_DIR / "bots" / "exportar_bling_ml.py"),
        "export-shopify":   str(BASE_DIR / "bots" / "exportar_bling_shopify.py"),
        "export-shopee":    str(BASE_DIR / "bots" / "exportar_bling_shopee.py"),
        "sync-estoque":     str(BASE_DIR / "bots" / "sync_estoque_canais.py"),
        "sync-preco-shopee":str(BASE_DIR / "bots" / "sync_preco_shopee.py"),
    }
    if bot_id not in scripts:
        return HTMLResponse("Bot não encontrado", status_code=404)

    q: _queue.Queue = _queue.Queue()
    _bot_queues[bot_id] = q
    t = threading.Thread(target=_run_bot_background, args=(bot_id, scripts[bot_id], q), daemon=True)
    t.start()

    def event_stream():
        while True:
            try:
                msg = q.get(timeout=120)
                yield f"data: {_json.dumps(msg, ensure_ascii=False)}\n\n"
                if msg.get("done"):
                    break
            except _queue.Empty:
                yield "data: {\"line\": \"[timeout]\"}\n\n"
                break

    return StreamingResponse(event_stream(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

@app.post("/produtos/bots/{bot_id}/stop")
def bot_stop(bot_id: str):
    proc = _bot_processes.get(bot_id)
    if proc:
        proc.terminate()
    return {"ok": True}

@app.get("/sistema/bling", response_class=HTMLResponse)
def sistema_bling():
    return HTMLResponse((PAGES_DIR / "sistema_bling.html").read_text(encoding="utf-8"))

@app.get("/sistema/ml", response_class=HTMLResponse)
def sistema_ml():
    return HTMLResponse((PAGES_DIR / "sistema_ml.html").read_text(encoding="utf-8"))

@app.get("/sistema/shopify", response_class=HTMLResponse)
def sistema_shopify_page():
    return HTMLResponse((PAGES_DIR / "sistema_shopify.html").read_text(encoding="utf-8"))

@app.get("/sistema/amazon", response_class=HTMLResponse)
def sistema_amazon():
    return HTMLResponse((PAGES_DIR / "sistema_amazon.html").read_text(encoding="utf-8"))

@app.get("/sistema/shopee", response_class=HTMLResponse)
def sistema_shopee():
    return HTMLResponse((PAGES_DIR / "sistema_shopee.html").read_text(encoding="utf-8"))

@app.get("/sistema/google", response_class=HTMLResponse)
def sistema_google():
    return HTMLResponse((PAGES_DIR / "sistema_google.html").read_text(encoding="utf-8"))

@app.get("/cost-engine", response_class=HTMLResponse)
def cost_engine_page():
    return HTMLResponse((PAGES_DIR / "cost_engine.html").read_text(encoding="utf-8"))

@app.get("/cost-allocation", response_class=HTMLResponse)
def cost_allocation_page():
    return HTMLResponse((PAGES_DIR / "cost_allocation.html").read_text(encoding="utf-8"))

@app.get("/oee", response_class=HTMLResponse)
def oee_page():
    return HTMLResponse((PAGES_DIR / "oee.html").read_text(encoding="utf-8"))

@app.get("/perfis", response_class=HTMLResponse)
def perfis_page():
    return HTMLResponse((PAGES_DIR / "perfis.html").read_text(encoding="utf-8"))

@app.get("/regras-calculo", response_class=HTMLResponse)
def regras_calculo_page():
    return HTMLResponse((PAGES_DIR / "regras_calculo.html").read_text(encoding="utf-8"))

@app.get("/sie", response_class=HTMLResponse)
def sie_page():
    return HTMLResponse((PAGES_DIR / "sie.html").read_text(encoding="utf-8"))

@app.get("/conferencia-sku", response_class=HTMLResponse)
def conferencia_sku_page():
    return HTMLResponse((PAGES_DIR / "conferencia_sku.html").read_text(encoding="utf-8"))

@app.get("/dashboard", response_class=HTMLResponse)
def dashboard_page():
    return HTMLResponse((PAGES_DIR / "dashboard.html").read_text(encoding="utf-8"))

@app.get("/conferencia/ml", response_class=HTMLResponse)
def conferencia_ml_page():
    return HTMLResponse((PAGES_DIR / "conferencia_ml.html").read_text(encoding="utf-8"))

@app.get("/conferencia/shopify", response_class=HTMLResponse)
def conferencia_shopify_page():
    return HTMLResponse((PAGES_DIR / "conferencia_shopify.html").read_text(encoding="utf-8"))

@app.get("/conferencia/amazon", response_class=HTMLResponse)
def conferencia_amazon_page():
    return HTMLResponse((PAGES_DIR / "conferencia_amazon.html").read_text(encoding="utf-8"))

@app.get("/conferencia/shopee", response_class=HTMLResponse)
def conferencia_shopee_page():
    return HTMLResponse((PAGES_DIR / "conferencia_shopee.html").read_text(encoding="utf-8"))


@app.get("/shopee/item/preview")
def shopee_item_preview(item_id: int = 0):
    """
    Busca detalhes de um item Shopee (nome, SKU, preço) + todos os modelos/variações.
    Usado pelo modal de importação para Bling.
    """
    if not item_id:
        return {"error": "Informe o item_id da Shopee."}
    try:
        from services.shopee import ShopeeService
        svc = ShopeeService()
        return svc.obter_item_completo(item_id)
    except RuntimeError as e:
        return {"error": str(e)}
    except Exception as e:
        logger.error("shopee_item_preview erro: %s", e)
        return {"error": f"Erro inesperado: {e}"}


@app.post("/shopee/item/importar-bling")
async def shopee_importar_bling(request: Request):
    """
    Importa anúncio Shopee para o Bling em 2 etapas:
    1. Cria o produto pai (tipo 'P') sem variações
    2. Para cada variação com SKU preenchido, cria via /produtos/{id}/variacoes

    Body: {
      nome, codigo, preco, situacao?,
      variacoes: [{nome, codigo, preco, estoque}]
    }
    """
    try:
        body = await request.json()
    except Exception:
        return {"ok": False, "error": "Body JSON inválido."}

    nome   = str(body.get("nome", "")).strip()
    codigo = str(body.get("codigo", "")).strip()
    preco  = float(body.get("preco", 0) or 0)

    if not nome:
        return {"ok": False, "error": "Nome do produto é obrigatório."}
    if not codigo:
        return {"ok": False, "error": "Código (SKU pai) é obrigatório."}

    variacoes_raw = body.get("variacoes") or []
    variacoes = [
        {
            "nome":    str(v.get("nome", "")).strip(),
            "codigo":  str(v.get("codigo", "")).strip(),
            "preco":   round(float(v.get("preco", preco) or preco), 2),
            "estoque": int(v.get("estoque", 0) or 0),
        }
        for v in variacoes_raw
        if str(v.get("codigo", "")).strip()
    ]
    if variacoes_raw and not variacoes:
        return {"ok": False, "error": "Nenhuma variação com SKU preenchido."}

    # ── Etapa 1: Cria produto pai ─────────────────────────────────────────────
    payload_pai: dict = {
        "nome":     nome,
        "codigo":   codigo,
        "preco":    round(preco, 2),
        "situacao": body.get("situacao", "A"),
        "tipo":     "P",  # Bling v3: P=Produto, S=Serviço, N=Serviço fiscal
    }

    try:
        client = BlingClient()
        res_pai = client.criar_produto(payload_pai)
    except Exception as e:
        logger.error("shopee_importar_bling — criar pai: %s", e)
        return {"ok": False, "step": "criar_produto", "error": str(e)}

    prod_id = (res_pai.get("data") or {}).get("id")
    if not prod_id:
        return {"ok": False, "step": "criar_produto", "error": f"Bling não retornou ID do produto: {res_pai}"}

    # ── Etapa 2: Cria variações (se houver) ──────────────────────────────────
    if not variacoes:
        return {"ok": True, "id": prod_id, "variacoes_criadas": 0}

    erros_var = []
    criadas = 0
    for v in variacoes:
        payload_var = {
            "nome":    v["nome"],
            "codigo":  v["codigo"],
            "preco":   v["preco"],
            "estoque": {"saldoInicial": v["estoque"]},
        }
        try:
            client.criar_variacao(prod_id, payload_var)
            criadas += 1
        except Exception as e:
            logger.warning("shopee_importar_bling — criar variação '%s': %s", v["codigo"], e)
            erros_var.append({"codigo": v["codigo"], "nome": v["nome"], "error": str(e)})

    return {
        "ok": True,
        "id": prod_id,
        "variacoes_criadas": criadas,
        "variacoes_erro": erros_var,
    }


@app.get("/auditoria/canais", response_class=HTMLResponse)
def auditoria_canais_page():
    return HTMLResponse((PAGES_DIR / "auditoria_canais.html").read_text(encoding="utf-8"))

@app.get("/auditoria/canais/dados")
def auditoria_canais_dados(
    canal: str = "",
    tipo: str = "",
    prioridade: str = "",
    limit: int = 200,
    offset: int = 0,
):
    """Retorna dados de auditoria filtrados por canal/tipo/prioridade."""
    from conferencia_sku import get_resultado
    dados = get_resultado()
    if not dados:
        return {"ok": False, "erro": "Nenhuma conferência executada ainda.", "auditoria": [], "resumo": {}}

    auditoria = dados.get("auditoria", [])

    if canal:
        auditoria = [a for a in auditoria if any(p["canal"] == canal for p in a["problemas"])]
    if tipo:
        auditoria = [a for a in auditoria if any(p["tipo"] == tipo for p in a["problemas"])]
    if prioridade:
        auditoria = [a for a in auditoria if any(p["prioridade"] == prioridade for p in a["problemas"])]

    total = len(auditoria)
    return {
        "ok": True,
        "executado_em": dados.get("executado_em"),
        "resumo": dados.get("resumo_auditoria", {}),
        "total_filtrado": total,
        "auditoria": auditoria[offset: offset + limit],
    }

@app.get("/bling/debug/sku-get")
def bling_debug_sku_get(sku: str = ""):
    """GET simplificado para buscar produto Bling por SKU — usado no debug modal ML."""
    if not BlingClient:
        return {"erro": "bling_client.py não encontrado"}
    sku = sku.strip()
    if not sku:
        return {"erro": "SKU não informado"}
    try:
        client = BlingClient()
        # Busca produto por código (SKU)
        result = client._get("/produtos", params={"codigo": sku, "limite": 5})
        produtos = result.get("data", [])
        if not produtos:
            return {"encontrado": False, "sku": sku, "mensagem": "Produto não encontrado no Bling com esse código/SKU"}
        # Pega o mais relevante
        prod = produtos[0]
        estoque = prod.get("estoque") or {}
        situacao = prod.get("situacao") or ""
        return {
            "encontrado": True,
            "id": prod.get("id"),
            "codigo": prod.get("codigo"),
            "nome": prod.get("nome"),
            "situacao": situacao,          # "A" = ativo, "I" = inativo
            "preco": prod.get("preco"),
            "estoque_atual": estoque.get("saldoVirtualTotal") or estoque.get("saldoFisicoTotal") or prod.get("estoqueAtual") or 0,
            "estoque_fisico": estoque.get("saldoFisicoTotal") or 0,
            "estoque_virtual": estoque.get("saldoVirtualTotal") or 0,
            "estrutura": prod.get("estrutura"),  # se é kit/componente
            "tipo": prod.get("tipo"),
            "raw_estoque": estoque,
        }
    except Exception as e:
        return {"erro": f"Erro Bling: {e}"}


@app.get("/ml/debug/item")
def ml_debug_item(mlb: str = ""):
    """
    Busca detalhes de um anúncio ML pelo código MLB.
    Retorna: id, titulo, status, seller_custom_field, available_quantity, variações.
    Útil para diagnosticar por que um anúncio não está sincronizando estoque.
    """
    import requests as _req
    mlb = mlb.strip().upper()
    if not mlb:
        return {"erro": "Informe o código MLB (ex: MLB3798077573)"}

    tp = DATA_DIR / "ml_tokens.json"
    if not tp.exists():
        return {"erro": "ML não conectado — faça login em /ml/login"}
    try:
        tokens = json.loads(tp.read_text(encoding="utf-8"))
    except Exception:
        return {"erro": "Erro ao ler tokens ML"}
    token = tokens.get("access_token", "")
    if not token:
        return {"erro": "access_token ML vazio"}

    h = {"Authorization": f"Bearer {token}"}
    try:
        r = _req.get(
            "https://api.mercadolibre.com/items",
            params={
                "ids": mlb,
                "attributes": "id,title,available_quantity,seller_custom_field,status,variations,listing_type_id,catalog_product_id",
            },
            headers=h,
            timeout=12,
        )
        if r.status_code == 401:
            return {"erro": "Token ML expirado — acesse /ml/login para reconectar"}
        if r.status_code != 200:
            return {"erro": f"ML API retornou HTTP {r.status_code}", "body": r.text[:300]}
        data = r.json()
        if not data:
            return {"erro": "Nenhum resultado para este MLB"}
        item = (data[0].get("body") or data[0]) if isinstance(data, list) else data

        # Estrutura de resposta amigável
        variations_out = []
        for v in (item.get("variations") or []):
            variations_out.append({
                "id":                  v.get("id"),
                "seller_custom_field": v.get("seller_custom_field"),
                "available_quantity":  v.get("available_quantity"),
                "attribute_combinations": [
                    f"{a.get('name')}: {a.get('value_name')}"
                    for a in (v.get("attribute_combinations") or [])
                ],
            })

        return {
            "mlb":                 item.get("id"),
            "titulo":              item.get("title"),
            "status":              item.get("status"),
            "seller_custom_field": item.get("seller_custom_field"),  # SKU configurado no anúncio pai
            "available_quantity":  item.get("available_quantity"),   # 0 para pais com variações
            "listing_type_id":     item.get("listing_type_id"),
            "catalog_product_id":  item.get("catalog_product_id"),
            "tem_variacoes":       len(variations_out) > 0,
            "total_variacoes":     len(variations_out),
            "variacoes":           variations_out,
            "diagnostico": _ml_diagnostico(item, variations_out),
        }
    except Exception as e:
        return {"erro": f"Erro ao consultar ML: {e}"}


def _ml_diagnostico(item: dict, variations: list) -> list[str]:
    """Gera lista de avisos/diagnósticos para um item ML."""
    diag = []
    scf = (item.get("seller_custom_field") or "").strip()
    qty = item.get("available_quantity", 0)
    status = (item.get("status") or "").lower()

    if status != "active":
        diag.append(f"⚠️ Anúncio não está ativo (status: {status}) — Bling não sincroniza estoque de anúncios inativos/pausados")

    if variations:
        diag.append(f"📦 Produto com {len(variations)} variação(ões) — o pai sempre tem qty=0; estoque é por variação")
        sem_sku = [v for v in variations if not (v.get("seller_custom_field") or "").strip()]
        com_sku = [v for v in variations if (v.get("seller_custom_field") or "").strip()]
        if sem_sku:
            diag.append(f"❌ {len(sem_sku)} variação(ões) sem seller_custom_field (SKU) — Bling não consegue identificar e sincronizar")
        if com_sku:
            diag.append(f"✅ {len(com_sku)} variação(ões) com SKU configurado — podem ser sincronizadas")
    else:
        # Produto simples
        if not scf:
            diag.append("❌ seller_custom_field vazio — Bling não consegue identificar este anúncio pelo SKU; sincronização não funcionará via integração padrão")
        else:
            diag.append(f"✅ seller_custom_field = '{scf}' — SKU configurado no anúncio")
        if qty == 0 and status == "active":
            diag.append("⚠️ Estoque zerado no ML com anúncio ativo — verifique se o Bling está enviando atualizações")

    return diag


# ─── Conferência de SKUs ────────────────────────────────────────────────────

@app.post("/conferencia-sku/executar")
def conferencia_sku_executar():
    """Inicia a conferência cruzada de SKUs em background."""
    if not BlingClient:
        raise HTTPException(status_code=500, detail="Bling não disponível.")
    from conferencia_sku import iniciar_conferencia_em_background, get_estado
    estado = get_estado()
    if estado.get("status") == "rodando":
        return {"ok": False, "msg": "Conferência já em andamento.", "estado": estado}
    bling = BlingClient()
    iniciou = iniciar_conferencia_em_background(bling)
    return {"ok": iniciou, "msg": "Conferência iniciada." if iniciou else "Já rodando.", "estado": get_estado()}

@app.post("/conferencia-sku/cancelar")
def conferencia_sku_cancelar():
    """Cancela a conferência em andamento (marca como cancelado)."""
    from conferencia_sku import _set_estado, get_estado
    estado = get_estado()
    if estado.get("status") == "rodando":
        _set_estado("erro", 0, "cancelado", "Conferência cancelada pelo usuário.")
        return {"ok": True, "msg": "Conferência cancelada."}
    return {"ok": False, "msg": "Nenhuma conferência em andamento."}

@app.get("/conferencia-sku/status")
def conferencia_sku_status():
    """Retorna o progresso da conferência em andamento."""
    from conferencia_sku import get_estado
    return get_estado()

@app.get("/conferencia-sku/resultado")
def conferencia_sku_resultado(
    canal: str = "",
    ausente_em: str = "",
    busca: str = "",
    so_bling: bool = False,
    limit: int = 500,
    offset: int = 0,
):
    """
    Retorna o resultado da última conferência com filtros.
    canal: filtra por canal específico (ml, shopify, amazon, shopee)
    ausente_em: mostra só SKUs AUSENTES naquele canal
    busca: filtra por SKU ou nome
    so_bling: mostra só produtos que estão APENAS no Bling (sem presença em nenhum canal)
    """
    from conferencia_sku import get_resultado
    dados = get_resultado()
    if not dados:
        return {"ok": False, "erro": "Nenhuma conferência executada ainda.", "matrix": [], "stats": {}}

    matrix = dados.get("matrix", [])

    # Filtros
    if canal:
        # Retorna apenas itens que têm dados do canal (presente_X não é None)
        campo = f"presente_{canal}"
        matrix = [r for r in matrix if r.get(campo) is not None]

    if busca:
        b = busca.lower()
        matrix = [r for r in matrix if b in r["sku"].lower() or b in (r.get("nome") or "").lower()]

    if ausente_em:
        campo = f"presente_{ausente_em}"
        matrix = [r for r in matrix if r.get(campo) is False]

    if so_bling:
        matrix = [
            r for r in matrix
            if r.get("presente_bling")
            and not r.get("presente_ml")
            and not r.get("presente_shopify")
            and not r.get("presente_amazon")
            and not r.get("presente_shopee")
        ]

    total = len(matrix)
    matrix_paginada = matrix[offset : offset + limit]

    return {
        "ok": True,
        "executado_em": dados.get("executado_em"),
        "duracao_segundos": dados.get("duracao_segundos"),
        "stats": dados.get("stats", {}),
        "sem_sku_ml": dados.get("sem_sku_ml", []),
        "ml_anuncios_remapeados": dados.get("ml_anuncios_remapeados", 0),
        "ml_gtin_remapeados": dados.get("ml_gtin_remapeados", 0),
        "resumo_auditoria": dados.get("resumo_auditoria", {}),
        "ml_sem_bling": dados.get("ml_sem_bling", []),
        "shopify_sem_bling": dados.get("shopify_sem_bling", []),
        "amazon_sem_bling": dados.get("amazon_sem_bling", []),
        "shopee_sem_bling": dados.get("shopee_sem_bling", []),
        "total_filtrado": total,
        "matrix": matrix_paginada,
    }

@app.get("/health")
def health():
    itens = [i for i in carregar_fila() if i.get("status") in ("pendente","incompleto")]
    return {"status":"Shinsei Pricing rodando","engine":pricing_module.__name__,"bling_client":bool(BlingClient),"bling_update_engine":bool(aplicar_precos_multicanal),"modo_busca":"sku_only","fila":_fila_stats(itens)}

@app.get("/bling/status")
def bling_status():
    if not BlingClient: return {"ok":False,"erro":"bling_client.py não encontrado."}
    try:
        client = BlingClient()
        return {"ok":True,"configurado":bool(getattr(client,"client_id","") and getattr(client,"client_secret","") and getattr(client,"redirect_uri","")),"token_local":bool(client.has_local_tokens())}
    except Exception as exc:
        return {"ok":False,"erro":str(exc)}

@app.get("/bling/raw-token")
def bling_raw_token():
    """Retorna o access token atual do Bling para debug/uso local."""
    if not BlingClient: return {"ok": False, "erro": "bling_client.py não encontrado."}
    try:
        client = BlingClient()
        headers = client._get_headers()
        return {"ok": True, "access_token": headers["Authorization"].replace("Bearer ", "")}
    except Exception as exc:
        return {"ok": False, "erro": str(exc)}

@app.get("/shopee/raw-token")
def shopee_raw_token():
    """Retorna os tokens atuais da Shopee para sync local."""
    try:
        from services.shopee import _carregar_tokens, ShopeeOAuthService, token_expirado
        if token_expirado():
            svc = ShopeeOAuthService()
            res = svc.renovar_token()
            if not res.get("success"):
                return {"ok": False, "erro": res.get("error", "Refresh falhou")}
        t = _carregar_tokens()
        if not t or not t.get("access_token"):
            return {"ok": False, "erro": "Shopee não autenticada"}
        return {"ok": True, "access_token": t["access_token"], "refresh_token": t.get("refresh_token",""),
                "shop_id": t.get("shop_id", 0), "expires_at": t.get("expires_at", 0)}
    except Exception as exc:
        return {"ok": False, "erro": str(exc)}


@app.get("/bling/auth")
def bling_auth():
    if not BlingClient: raise HTTPException(status_code=500, detail="bling_client.py não encontrado.")
    try:
        client = BlingClient()
        return RedirectResponse(client.build_authorize_url())
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))

@app.get("/bling/callback")
def bling_callback(code: str | None = Query(None), state: str | None = Query(None), error: str | None = Query(None), error_description: str | None = Query(None)):
    if not BlingClient: raise HTTPException(status_code=500, detail="bling_client.py não encontrado.")
    if error: raise HTTPException(status_code=400, detail=f"Bling OAuth retornou erro: {error}. {error_description or ''}".strip())
    if not code: raise HTTPException(status_code=400, detail="Callback do Bling sem code de autorização.")
    try:
        client = BlingClient()
        token = client.exchange_code_for_token(code, state=state)
        return {"ok":True,"message":"Conexão com Bling realizada.","expires_in":token.get("expires_in")}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


# ── Bling AKG (segunda conta) ─────────────────────────────────────────────────
_BLING_AKG_REDIRECT_URI  = os.getenv("BLING_AKG_REDIRECT_URI",  "https://shinsei-pricing.onrender.com/bling/callback2")
_BLING_AKG_TOKEN_PATH    = Path("data/bling_tokens_akg.json")
_BLING_AKG_STATE_PATH    = Path("data/bling_oauth_state_akg.json")
_CREDENTIALS_PATH        = Path("data/credentials.json")

def _bling_akg_creds() -> tuple[str, str]:
    """Lê client_id e client_secret do credentials.json (salvo pela página de integrações)."""
    try:
        creds = json.loads(_CREDENTIALS_PATH.read_text(encoding="utf-8"))
        akg = creds.get("bling_akg", {})
        cid = akg.get("client_id", "") or os.getenv("BLING_AKG_CLIENT_ID", "")
        sec = akg.get("client_secret", "") or os.getenv("BLING_AKG_CLIENT_SECRET", "")
        return cid, sec
    except Exception:
        return os.getenv("BLING_AKG_CLIENT_ID", ""), os.getenv("BLING_AKG_CLIENT_SECRET", "")
def _bling_akg_save(data: dict):
    _BLING_AKG_TOKEN_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

def _bling_akg_load() -> dict:
    if not _BLING_AKG_TOKEN_PATH.exists():
        return {}
    try:
        return json.loads(_BLING_AKG_TOKEN_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}

def _bling_akg_headers() -> dict:
    import time as _time
    import requests as _req
    tokens = _bling_akg_load()
    access = tokens.get("access_token", "")
    expires_at = float(tokens.get("expires_at", 0) or 0)
    if not access or _time.time() >= expires_at:
        _cid, _csec = _bling_akg_creds()
        r = _req.post("https://www.bling.com.br/Api/v3/oauth/token",
            data={"grant_type": "refresh_token",
                  "refresh_token": tokens.get("refresh_token", "")},
            auth=(_cid, _csec),
            timeout=20)
        if r.status_code == 200:
            new = r.json()
            new["expires_at"] = _time.time() + new.get("expires_in", 3600) - 60
            _bling_akg_save(new)
            access = new["access_token"]
    return {"Authorization": f"Bearer {access}", "Content-Type": "application/json"}

@app.get("/bling/auth2")
def bling_auth2():
    """Inicia OAuth Bling para conta AKG (segunda empresa)."""
    import secrets as _sec
    import time as _time
    from urllib.parse import urlencode
    state = _sec.token_urlsafe(32)
    _BLING_AKG_STATE_PATH.write_text(json.dumps({"state": state, "created_at": int(_time.time())}), encoding="utf-8")
    _cid, _csec = _bling_akg_creds()
    params = urlencode({
        "response_type": "code",
        "client_id": _cid,
        "redirect_uri": _BLING_AKG_REDIRECT_URI,
        "state": state,
    })
    return RedirectResponse(f"https://www.bling.com.br/Api/v3/oauth/authorize?{params}")

@app.get("/bling/callback2")
def bling_callback2(code: str | None = Query(None), state: str | None = Query(None),
                    error: str | None = Query(None), error_description: str | None = Query(None)):
    """Callback OAuth Bling conta AKG."""
    if error:
        raise HTTPException(status_code=400, detail=f"Bling OAuth erro: {error}. {error_description or ''}")
    if not code:
        raise HTTPException(status_code=400, detail="Callback sem code de autorização.")
    import time as _time
    import requests as _req
    _cid, _csec = _bling_akg_creds()
    r = _req.post("https://www.bling.com.br/Api/v3/oauth/token",
        data={"grant_type": "authorization_code",
              "code": code,
              "redirect_uri": _BLING_AKG_REDIRECT_URI},
        auth=(_cid, _csec),
        timeout=20)
    if r.status_code != 200:
        raise HTTPException(status_code=400, detail=f"Erro ao trocar code: {r.text[:300]}")
    tokens = r.json()
    tokens["expires_at"] = _time.time() + tokens.get("expires_in", 3600) - 60
    _bling_akg_save(tokens)
    return {"ok": True, "message": "Conta AKG conectada ao Bling com sucesso.", "expires_in": tokens.get("expires_in")}

@app.get("/bling/status2")
def bling_status2():
    """Status da conexão Bling conta AKG."""
    import time as _time
    tokens = _bling_akg_load()
    access = tokens.get("access_token", "")
    if not access:
        return {"ok": False, "conectado": False, "message": "Conta AKG não conectada. Acesse /bling/auth2"}
    expires_at = float(tokens.get("expires_at", 0) or 0)
    expirado = _time.time() >= expires_at
    return {"ok": True, "conectado": True, "expirado": expirado,
            "expires_at": expires_at, "message": "Conta AKG conectada."}

@app.get("/bling/tokens2")
def bling_tokens2():
    """Exporta token Bling AKG para scripts locais (igual /ml/tokens2)."""
    tokens = _bling_akg_load()
    if not tokens.get("access_token"):
        return {"success": False, "error": "Bling AKG não conectado."}
    return {"success": True, "data": tokens}

@app.get("/ml/akg/copiar-shinsei/status")
def ml_akg_copy_status():
    """Status do job de cópia Shinsei → AKG."""
    from services.ml_copy_service import _load_progress, _summary
    progress = _load_progress()
    return _summary(progress)

@app.post("/ml/akg/copiar-shinsei/dry-run")
def ml_akg_copy_dry_run(limit: int = Query(default=10)):
    """Simula a cópia sem publicar nada. Mostra os primeiros `limit` itens que seriam criados."""
    try:
        from services.ml_copy_service import run_copy
        return run_copy(limit=limit, dry_run=True, reset=True)
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/ml/akg/copiar-shinsei/iniciar")
def ml_akg_copy_iniciar(
    background_tasks: BackgroundTasks,
    limit: int = Query(default=0, description="0 = sem limite"),
    reset: bool = Query(default=False, description="Reinicia do zero ignorando progresso"),
):
    """Inicia a cópia em background dos anúncios da Shinsei para o ML da AKG."""
    _log_msgs: list[str] = []

    def _callback(msg: str):
        _log_msgs.append(msg)
        if len(_log_msgs) % 50 == 0:
            import logging
            logging.getLogger("shinsei.ml_copy").info("progresso: %d mensagens", len(_log_msgs))

    def _run():
        from services.ml_copy_service import run_copy
        run_copy(limit=limit, dry_run=False, reset=reset, status_callback=_callback)

    background_tasks.add_task(_run)
    return {"ok": True, "message": "Job iniciado em background. Acompanhe em /ml/akg/copiar-shinsei/status"}

@app.post("/ml/injetar-tokens")
async def ml_injetar_tokens(request: Request):
    """Injeta tokens ML diretamente (para renovar sem OAuth quando o servidor reinicia)."""
    import time as _t
    body = await request.json()
    shinsei = body.get("shinsei")
    akg = body.get("akg")
    resultado = {}
    if shinsei:
        p = DATA_DIR / "ml_tokens.json"
        shinsei["expires_at"] = _t.time() + shinsei.get("expires_in", 21600) - 300
        p.write_text(json.dumps(shinsei, indent=2, ensure_ascii=False), encoding="utf-8")
        resultado["shinsei"] = "ok"
    if akg:
        p = DATA_DIR / "ml_tokens_akg.json"
        akg["expires_at"] = _t.time() + akg.get("expires_in", 21600) - 300
        p.write_text(json.dumps(akg, indent=2, ensure_ascii=False), encoding="utf-8")
        resultado["akg"] = "ok"
    return {"ok": True, "resultado": resultado}

@app.post("/ml/akg/copiar-shinsei/resetar")
def ml_akg_copy_reset():
    """Limpa o progresso salvo para recomeçar do zero."""
    from services.ml_copy_service import PROGRESS_PATH
    if PROGRESS_PATH.exists():
        PROGRESS_PATH.unlink()
    return {"ok": True, "message": "Progresso resetado."}

@app.post("/ml/akg/copiar-shinsei/teste-real")
def ml_akg_copy_teste_real():
    """Publica 1 item real no ML AKG e retorna o ID criado para verificação."""
    try:
        from services.ml_copy_service import run_copy
        result = run_copy(limit=1, dry_run=False, reset=True)
        return result
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/ml/shinsei/item-raw/{item_id}")
def ml_shinsei_item_raw(item_id: str):
    """Retorna o JSON completo de um item da Shinsei no ML (para diagnóstico)."""
    import requests as _req
    from services.ml_copy_service import _shinsei_token, _headers
    try:
        token = _shinsei_token()
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))
    r = _req.get(f"https://api.mercadolibre.com/items/{item_id}",
                 headers=_headers(token), timeout=15)
    if r.status_code != 200:
        raise HTTPException(status_code=r.status_code, detail=r.text[:500])
    return r.json()

@app.post("/ml/akg/debug-post/{item_id}")
def ml_akg_debug_post(item_id: str, listing_type: str = Query(default="free")):
    """Tenta criar o item no ML AKG e retorna o erro completo (sem truncar)."""
    import requests as _req
    from services.ml_copy_service import _shinsei_token, _akg_token, _headers, _get_items_details, _build_payload
    try:
        shin_token = _shinsei_token()
        akg_token = _akg_token()
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))
    items = _get_items_details([item_id], shin_token)
    if not items:
        raise HTTPException(status_code=404, detail=f"Item {item_id} não encontrado na Shinsei")
    item = items[0]
    payload = _build_payload(item)
    if not payload:
        raise HTTPException(status_code=400, detail="Payload inválido (sem título/categoria/preço)")
    payload["listing_type_id"] = listing_type  # override para teste
    r = _req.post("https://api.mercadolibre.com/items",
                  json=payload, headers=_headers(akg_token), timeout=30)
    return {
        "status_code": r.status_code,
        "payload_enviado": payload,
        "resposta_ml": r.json() if r.headers.get("content-type","").startswith("application/json") else r.text,
    }

@app.get("/ml/akg/debug-payload/{item_id}")
def ml_akg_debug_payload(item_id: str):
    """Mostra o item raw da Shinsei e o payload que seria enviado ao ML AKG."""
    import requests as _req
    from services.ml_copy_service import _shinsei_token, _headers, _build_payload
    try:
        token = _shinsei_token()
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))
    # Busca individual sem filtro para receber family_name e todos os campos privados
    r = _req.get(
        f"https://api.mercadolibre.com/items/{item_id}",
        headers=_headers(token), timeout=15
    )
    if r.status_code != 200:
        raise HTTPException(status_code=r.status_code, detail=r.text[:500])
    item = r.json()
    payload = _build_payload(item)
    return {
        "item_raw_fields": {
            "id": item.get("id"),
            "family_name": item.get("family_name"),
            "domain_id": item.get("domain_id"),
            "category_id": item.get("category_id"),
            "seller_custom_field": item.get("seller_custom_field"),
            "attributes_ids": [a.get("id") for a in item.get("attributes", [])],
        },
        "payload_enviado": payload,
    }

@app.get("/ml/akg/verificar-item/{item_id}")
def ml_akg_verificar_item(item_id: str):
    """Verifica se um item existe no ML AKG e retorna seus dados principais."""
    import requests as _req
    from services.ml_copy_service import _akg_token, _headers
    try:
        token = _akg_token()
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))
    r = _req.get(f"https://api.mercadolibre.com/items/{item_id}",
                 headers=_headers(token), timeout=15)
    if r.status_code != 200:
        raise HTTPException(status_code=r.status_code, detail=r.text[:300])
    d = r.json()
    return {
        "id": d.get("id"),
        "title": d.get("title"),
        "status": d.get("status"),
        "listing_type_id": d.get("listing_type_id"),
        "price": d.get("price"),
        "seller_custom_field": d.get("seller_custom_field"),
        "permalink": d.get("permalink"),
        "thumbnail": d.get("thumbnail"),
    }

@app.get("/bling/akg/lojas")
def bling_akg_lojas():
    """Descobre idLoja ML no Bling AKG via anúncios existentes."""
    import requests as _req
    hdrs = _bling_akg_headers()
    # Tenta pegar um anúncio existente para extrair o idLoja
    endpoints_tentados = []
    for ep in ["/lojas", "/integracoes", "/canais-de-venda", "/canaisdevenda"]:
        r = _req.get(f"https://api.bling.com.br/Api/v3{ep}", headers=hdrs, timeout=20)
        endpoints_tentados.append({"endpoint": ep, "status": r.status_code, "body": r.text[:200]})
    # Tenta /anuncios sem filtro para ver o formato de retorno e extrair idLoja
    r2 = _req.get("https://api.bling.com.br/Api/v3/anuncios", params={"pagina": 1, "limite": 5},
                  headers=hdrs, timeout=20)
    anuncios_sample = r2.json() if r2.status_code == 200 else {"erro": r2.text[:300]}
    return {"endpoints_testados": endpoints_tentados, "anuncios_sample": anuncios_sample}


@app.get("/bling/akg/anuncios-sem-ml")
def bling_akg_anuncios_sem_ml(id_loja: int = Query(..., description="ID da loja ML no Bling AKG")):
    """Lista produtos do Bling AKG que ainda não têm anúncio no ML."""
    import requests as _req
    hdrs = _bling_akg_headers()

    # Busca anúncios já publicados
    publicados_skus: set[str] = set()
    pagina = 1
    while True:
        r = _req.get("https://api.bling.com.br/Api/v3/anuncios",
                     params={"idLoja": id_loja, "pagina": pagina, "limite": 100},
                     headers=hdrs, timeout=20)
        if r.status_code != 200:
            break
        items = r.json().get("data", [])
        if not items:
            break
        for a in items:
            sku = a.get("produto", {}).get("codigo", "") or a.get("idVendedor", "")
            if sku:
                publicados_skus.add(sku)
        pagina += 1

    # Busca todos os produtos
    todos: list[dict] = []
    pagina = 1
    while True:
        r = _req.get("https://api.bling.com.br/Api/v3/produtos",
                     params={"pagina": pagina, "limite": 100, "situacao": "A"},
                     headers=hdrs, timeout=20)
        if r.status_code != 200:
            break
        items = r.json().get("data", [])
        if not items:
            break
        todos.extend(items)
        pagina += 1

    sem_ml = [p for p in todos if p.get("codigo", "") not in publicados_skus]

    return {
        "total_produtos": len(todos),
        "publicados_no_ml": len(publicados_skus),
        "sem_anuncio_ml": len(sem_ml),
        "skus_sem_ml": [{"id": p.get("id"), "codigo": p.get("codigo"), "nome": p.get("nome")} for p in sem_ml[:200]],
    }


@app.get("/bling/exportar-tokens")
def bling_exportar_tokens(api_key: str = Query(...)):
    """Exporta tokens Bling para configurar como env vars no Cloud Run. Protegido por api_key (middleware)."""
    if not BlingClient:
        raise HTTPException(status_code=500, detail="bling_client.py não encontrado.")
    try:
        from bling_client import TOKEN_PATH, _decrypt_tokens
        if not TOKEN_PATH.exists():
            return {"ok": False, "erro": "Arquivo bling_tokens.json não encontrado. Faça a autorização em /bling/auth primeiro."}
        raw = json.loads(TOKEN_PATH.read_text(encoding="utf-8"))
        if "encrypted" in raw:
            tokens = _decrypt_tokens(raw["encrypted"])
        else:
            tokens = raw
        access_token = tokens.get("access_token", "")
        refresh_token = tokens.get("refresh_token", "")
        if not access_token or not refresh_token:
            return {"ok": False, "erro": "Tokens inválidos ou ausentes. Reautorize em /bling/auth."}
        cmd = (
            f"gcloud run services update shinsei-pricing "
            f"--region southamerica-east1 "
            f"--update-env-vars BLING_ACCESS_TOKEN={access_token},BLING_REFRESH_TOKEN={refresh_token}"
        )
        return {
            "ok": True,
            "access_token": access_token,
            "refresh_token": refresh_token,
            "expires_at": tokens.get("expires_at"),
            "gcloud_cmd": cmd,
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))

@app.post("/bling/debug/sku")
def bling_debug_sku(payload: DebugSkuPayload):
    if not BlingClient: raise HTTPException(status_code=500, detail="bling_client.py não encontrado.")
    try:
        client = BlingClient()
        return client.debug_product_by_sku(payload.sku)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))

@app.post("/bling/produto/buscar")
def bling_produto_buscar(payload: DebugSkuPayload):
    if not BlingClient: raise HTTPException(status_code=500, detail="bling_client.py não encontrado.")
    try:
        client = BlingClient()
        return client.get_product_by_sku(payload.sku)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))

@app.post("/bling/produto/atualizar-peso")
def bling_produto_atualizar_peso(payload: AtualizacaoCampoBlingPayload):
    if not BlingClient:
        raise HTTPException(status_code=500, detail="bling_client.py não encontrado.")
    if float(payload.valor or 0) <= 0:
        raise HTTPException(status_code=400, detail="Informe um peso válido.")
    try:
        client = BlingClient()
        existing = client.get_product(int(payload.produto_id))
        patch = _prepare_product_patch(existing)
        patch["id"] = int(payload.produto_id)
        patch["pesoLiquido"] = float(payload.valor)
        patch["peso"] = float(payload.valor)
        if not patch.get("pesoBruto"):
            patch["pesoBruto"] = float(payload.valor)
        result = client.update_product(int(payload.produto_id), patch)
        atualizado = client.get_product(int(payload.produto_id))
        return {"ok": True, "message": "Peso atualizado no Bling.", "produto": atualizado, "raw": result}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Falha ao atualizar peso no Bling: {exc}")

@app.post("/bling/produto/atualizar-preco")
def bling_produto_atualizar_preco(payload: AtualizacaoCampoBlingPayload):
    if not BlingClient:
        raise HTTPException(status_code=500, detail="bling_client.py não encontrado.")
    if float(payload.valor or 0) <= 0:
        raise HTTPException(status_code=400, detail="Informe um preço válido.")
    try:
        client = BlingClient()
        existing = client.get_product(int(payload.produto_id))
        valor = round(float(payload.valor), 2)

        # Tenta via fornecedor primeiro
        fornecedor = existing.get("fornecedor") if isinstance(existing, dict) else None
        if isinstance(fornecedor, dict) and fornecedor.get("id"):
            patch = _prepare_product_patch(existing) if "_prepare_product_patch" in dir() else dict(existing)
            patch["id"] = int(payload.produto_id)
            fornecedor_patch = dict(fornecedor)
            fornecedor_patch["precoCusto"] = valor
            fornecedor_patch["precoCompra"] = valor
            patch["fornecedor"] = fornecedor_patch
            client.update_product(int(payload.produto_id), patch)
            logger.info("Custo atualizado via fornecedor: produto_id=%s valor=%s", payload.produto_id, valor)
            return {"ok": True, "message": f"Custo R${valor:.2f} atualizado no Bling."}
        else:
            # Fallback: salva custo localmente como override
            custo_override_path = DATA_DIR / "custo_override.json"
            overrides = _load_json(custo_override_path, {})
            sku = existing.get("codigo") or str(payload.produto_id) if isinstance(existing, dict) else str(payload.produto_id)
            prod_id_str = str(payload.produto_id)
            entry = {"custo": valor, "produto_id": payload.produto_id, "atualizado_em": datetime.utcnow().isoformat()}
            overrides[sku] = entry
            overrides[prod_id_str] = entry
            _save_json(custo_override_path, overrides)
            logger.info("Custo salvo localmente (sem fornecedor): sku=%s valor=%s", sku, valor)
            return {"ok": True, "message": f"Custo R${valor:.2f} salvo localmente. Recalcule para ver o resultado.", "local_only": True}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Falha ao atualizar custo: {exc}")


@app.get("/bling/debug/produto-raw/{produto_id}")
def bling_debug_produto_raw(produto_id: int):
    """Retorna a resposta RAW do Bling para um produto. Sem auth."""
    if not BlingClient:
        raise HTTPException(status_code=500, detail="bling_client.py não encontrado.")
    client = BlingClient()
    raw = client._get(f"/produtos/{int(produto_id)}")
    return {"produto_id": produto_id, "raw": raw}


@app.get("/bling/produto/{produto_id}/variacoes")
def bling_produto_variacoes(produto_id: int):
    """Retorna o produto com todas as variações (IDs e opções). Sem auth."""
    if not BlingClient:
        raise HTTPException(status_code=500, detail="bling_client.py não encontrado.")
    client = BlingClient()
    # Tenta o endpoint /variacoes do Bling v3 primeiro
    try:
        raw_var = client._get(f"/produtos/{int(produto_id)}/variacoes")
        variacoes_data = raw_var.get("data", [])
        if isinstance(variacoes_data, list) and variacoes_data:
            result = []
            for v in variacoes_data:
                result.append({
                    "id": v.get("id"),
                    "nome": v.get("nome"),
                    "codigo": v.get("codigo"),
                    "preco": v.get("preco"),
                    "imagens": v.get("imagens", []),
                    "variacao": v.get("variacao", {}),
                })
            # Get product name
            produto_name = None
            try:
                produto = client.get_product(produto_id)
                produto_name = produto.get("nome")
            except Exception:
                pass
            return {"produto_id": produto_id, "nome": produto_name, "total": len(result), "variacoes": result, "source": "variacoes_endpoint"}
    except Exception:
        pass
    # Fallback: GET /produtos/{id} inline variacoes
    produto = client.get_product(produto_id)
    if isinstance(produto, dict) and "data" in produto:
        produto = produto["data"]
    variacoes = produto.get("variacoes", [])
    result = []
    for v in variacoes:
        result.append({
            "id": v.get("id"),
            "nome": v.get("nome"),
            "codigo": v.get("codigo"),
            "preco": v.get("preco"),
            "imagens": v.get("imagens", []),
            "variacao": v.get("variacao", {}),
        })
    return {"produto_id": produto_id, "nome": produto.get("nome"), "total": len(result), "variacoes": result, "source": "inline"}


@app.get("/bling/produto/variacoes/{produto_id}")
def bling_listar_variacoes(produto_id: int):
    """Lista variações de um produto Bling com id e nome. Sem auth."""
    client = BlingClient()
    data = client.get_product(produto_id)
    p = data.get("data", data) if isinstance(data, dict) else {}
    variacoes = p.get("variacoes", [])
    return {
        "produto_id": produto_id,
        "nome": p.get("nome", ""),
        "total": len(variacoes),
        "variacoes": [{"id": v["id"], "nome": v.get("nome", ""), "imagemURL": v.get("imagemURL", "")} for v in variacoes],
    }


class ImagensBatchPayload(BaseModel):
    produto_id: int
    imagens: dict  # {variacao_id (str): image_url}


@app.post("/bling/produto/atualizar-imagens-batch")
def bling_atualizar_imagens_batch(payload: ImagensBatchPayload):
    """Atualiza imagemURL de múltiplas variações num único PATCH. Sem auth."""
    import requests as _req
    client = BlingClient()
    data = client.get_product(payload.produto_id)
    p = data.get("data", data) if isinstance(data, dict) else {}
    variacoes = p.get("variacoes", [])
    if not variacoes:
        raise HTTPException(status_code=404, detail="Produto não encontrado ou sem variações.")

    id_to_url = {int(k): v for k, v in payload.imagens.items()}
    vars_payload = []
    for v in variacoes:
        vp = {"id": v["id"]}
        if v.get("estrutura"):
            vp["estrutura"] = v["estrutura"]
        if v.get("gtin"):
            vp["gtin"] = v["gtin"]
        url = id_to_url.get(int(v["id"]))
        if url:
            vp["imagemURL"] = url
        vars_payload.append(vp)

    patch = {
        "nome": p.get("nome", ""), "tipo": p.get("tipo", "P"),
        "situacao": p.get("situacao", "A"), "formato": p.get("formato", "V"),
        "variacoes": vars_payload,
    }
    hdrs = client._get_headers()
    hdrs["Content-Type"] = "application/json"
    resp = _req.patch(
        f"{client.base_url}/produtos/{payload.produto_id}",
        headers=hdrs, json=patch, timeout=90,
    )
    if resp.status_code == 200:
        aplicadas = sum(1 for v in vars_payload if v.get("imagemURL"))
        return {"ok": True, "produto_id": payload.produto_id, "imagens_aplicadas": aplicadas}
    raise HTTPException(status_code=resp.status_code, detail=resp.text[:500])


@app.post("/bling/produto/atualizar-imagem-variacao")
def bling_atualizar_imagem_variacao(payload: ImagemVariacaoPayload):
    """Atualiza a imagem principal de uma variação no Bling. Sem auth."""
    if not BlingClient:
        raise HTTPException(status_code=500, detail="bling_client.py não encontrado.")
    client = BlingClient()
    existing = client.get_product(payload.produto_id)
    existing_data = existing.get("data", existing) if isinstance(existing, dict) else {}
    # Verifica se a variação existe
    all_vars = existing_data.get("variacoes") or []
    variacao_encontrada = any(int(v.get("id", 0)) == payload.variacao_id for v in all_vars)
    if not variacao_encontrada:
        raise HTTPException(status_code=404, detail=f"Variação {payload.variacao_id} não encontrada.")
    # Payload mínimo com campos obrigatórios do Bling v3 + variação alvo
    patch = {
        "nome": existing_data.get("nome", ""),
        "tipo": existing_data.get("tipo", "V"),
        "situacao": existing_data.get("situacao", "A"),
        "formato": existing_data.get("formato", "E"),
        "variacoes": [{"id": payload.variacao_id, "imagens": [{"link": payload.image_url}]}],
    }
    result = client.update_product(payload.produto_id, patch)
    return {"ok": True, "produto_id": payload.produto_id, "variacao_id": payload.variacao_id, "image_url": payload.image_url, "raw": result}


@app.post("/bling/produto/buscar-por-nome")
def bling_buscar_por_nome(payload: BuscaProdutoPayload):
    """Busca produtos no Bling pelo nome. Sem auth."""
    if not BlingClient:
        raise HTTPException(status_code=500, detail="bling_client.py não encontrado.")
    client = BlingClient()
    results = client.search_by_name(payload.nome, limit=payload.limite)
    return {"total": len(results), "produtos": results}


class ImagemSimplesBlingPayload(BaseModel):
    produto_id: int
    image_url: str

class BuscaSkuPayload(BaseModel):
    sku: str


@app.post("/bling/produto/atualizar-imagem-simples")
def bling_atualizar_imagem_simples(payload: ImagemSimplesBlingPayload):
    """Atualiza a imagem de um produto simples/composto (tipo=P) no Bling.
    Sem auth. Exclui camposCustomizados para evitar erro de permissão."""
    if not BlingClient:
        raise HTTPException(status_code=500, detail="bling_client.py não encontrado.")
    client = BlingClient()
    # Busca produto existente (raw para ter todos os campos)
    raw_resp = client._get(f"/produtos/{int(payload.produto_id)}")
    existing = raw_resp.get("data", {})
    if not isinstance(existing, dict):
        existing = {}
    # Campos a EXCLUIR (causam erros de permissão ou não devem ser enviados)
    EXCLUIR = {
        "camposCustomizados",  # permissão negada
        "variacoes",           # não aplicável para tipo=P
        "variacao",            # não aplicável para tipo=P
        "estoque",             # gerenciado separadamente
        "actionEstoque",       # apenas para consulta
    }
    patch = {k: v for k, v in existing.items() if k not in EXCLUIR}
    patch["id"] = payload.produto_id
    # NOTA: Bling API v3 NÃO salva imagens via externas (testado exaustivamente).
    # Preserva internas existentes para não apagar imagens já hospedadas na Bling S3.
    existing_midia = existing.get("midia", {}) or {}
    existing_imagens = existing_midia.get("imagens", {}) or {}
    existing_internas = existing_imagens.get("internas", []) or []
    patch["midia"] = {
        "video": existing_midia.get("video", {"url": ""}),
        "imagens": {
            "externas": [{"link": payload.image_url}],
            "internas": existing_internas,  # Preserva imagens já hospedadas na Bling
            "imagensURL": [],
        }
    }
    result = client.update_product(payload.produto_id, patch)
    return {"ok": True, "produto_id": payload.produto_id, "image_url": payload.image_url,
            "nota": "Bling API v3 não salva URLs externas — imagem deve ser inserida pela interface web do Bling",
            "raw": result}


@app.post("/bling/debug/testar-imagem")
def bling_debug_testar_imagem(payload: ImagemSimplesBlingPayload):
    """Debug: testa diferentes abordagens para salvar imagem no Bling.
    Retorna estado antes, payload enviado, resposta e estado depois."""
    if not BlingClient:
        raise HTTPException(status_code=500, detail="bling_client.py não encontrado.")
    import requests as _req
    client = BlingClient()
    produto_id = int(payload.produto_id)
    image_url = payload.image_url

    # Estado ANTES
    raw_antes = client._get(f"/produtos/{produto_id}")
    data_antes = raw_antes.get("data", {})
    externas_antes = data_antes.get("midia", {}).get("imagens", {}).get("externas", [])

    headers = client._get_headers()
    results = []

    # Abordagem A: PUT com todos os campos exceto camposCustomizados/variacoes/estoque
    EXCLUIR_A = {"camposCustomizados", "variacoes", "variacao", "estoque", "actionEstoque"}
    patch_a = {k: v for k, v in data_antes.items() if k not in EXCLUIR_A}
    patch_a["id"] = produto_id
    patch_a["midia"] = {
        "video": {"url": ""},
        "imagens": {"externas": [{"link": image_url}], "internas": [], "imagensURL": []}
    }
    resp_a = _req.put(f"{client.base_url}/produtos/{produto_id}", headers=headers, json=patch_a, timeout=30)

    import time as _time; _time.sleep(2)
    after_a = client._get(f"/produtos/{produto_id}").get("data", {}).get("midia", {}).get("imagens", {}).get("externas", [])
    results.append({"abordagem": "A_full_minus_custom", "status": resp_a.status_code, "response": resp_a.text[:300], "externas_depois": after_a})

    # Abordagem B: PUT sem estrutura (só campos básicos + midia)
    EXCLUIR_B = EXCLUIR_A | {"estrutura", "artigoPerigoso", "descricaoEmbalagemDiscreta", "linhaProduto",
                              "tributacao", "dimensoes", "fornecedor", "categoria", "descricaoComplementar",
                              "linkExterno", "observacoes", "gtinEmbalagem", "itensPorCaixa", "volumes", "dataValidade"}
    patch_b = {k: v for k, v in data_antes.items() if k not in EXCLUIR_B}
    patch_b["id"] = produto_id
    patch_b["midia"] = {
        "video": {"url": ""},
        "imagens": {"externas": [{"link": image_url}], "internas": [], "imagensURL": []}
    }
    resp_b = _req.put(f"{client.base_url}/produtos/{produto_id}", headers=headers, json=patch_b, timeout=30)
    _time.sleep(2)
    after_b = client._get(f"/produtos/{produto_id}").get("data", {}).get("midia", {}).get("imagens", {}).get("externas", [])
    results.append({"abordagem": "B_sem_estrutura", "status": resp_b.status_code, "response": resp_b.text[:300], "externas_depois": after_b})

    # Abordagem C: PATCH (se suportado) com só midia
    resp_c = _req.patch(
        f"{client.base_url}/produtos/{produto_id}",
        headers=headers,
        json={"midia": {"video": {"url": ""}, "imagens": {"externas": [{"link": image_url}], "internas": [], "imagensURL": []}}},
        timeout=30
    )
    _time.sleep(2)
    after_c = client._get(f"/produtos/{produto_id}").get("data", {}).get("midia", {}).get("imagens", {}).get("externas", [])
    results.append({"abordagem": "C_patch_midia_only", "status": resp_c.status_code, "response": resp_c.text[:300], "externas_depois": after_c})

    # Abordagem D: PUT com imagensURL (campo diferente)
    EXCLUIR_A2 = {"camposCustomizados", "variacoes", "variacao", "estoque", "actionEstoque"}
    patch_d = {k: v for k, v in data_antes.items() if k not in EXCLUIR_A2}
    patch_d["id"] = produto_id
    patch_d["midia"] = {
        "video": {"url": ""},
        "imagens": {"externas": [], "internas": [], "imagensURL": [image_url]}
    }
    resp_d = _req.put(f"{client.base_url}/produtos/{produto_id}", headers=headers, json=patch_d, timeout=30)
    _time.sleep(2)
    after_d = client._get(f"/produtos/{produto_id}").get("data", {}).get("midia", {})
    results.append({"abordagem": "D_imagensURL", "status": resp_d.status_code, "response": resp_d.text[:300], "midia_depois": after_d})

    # Abordagem E: PUT com internas = [{"link": url}] (mesma estrutura que externas mas em internas)
    patch_e = {k: v for k, v in data_antes.items() if k not in EXCLUIR_A2}
    patch_e["id"] = produto_id
    patch_e["midia"] = {
        "video": {"url": ""},
        "imagens": {"externas": [], "internas": [{"link": image_url}], "imagensURL": []}
    }
    resp_e = _req.put(f"{client.base_url}/produtos/{produto_id}", headers=headers, json=patch_e, timeout=30)
    _time.sleep(2)
    after_e = client._get(f"/produtos/{produto_id}").get("data", {}).get("midia", {})
    results.append({"abordagem": "E_internas_url", "status": resp_e.status_code, "response": resp_e.text[:300], "midia_depois": after_e})

    # Abordagem F: download image + multipart upload
    import io as _io
    img_resp = _req.get(image_url, timeout=20)
    if img_resp.status_code == 200:
        files = {"imagem": ("kit_6.4.jpg", _io.BytesIO(img_resp.content), "image/jpeg")}
        headers_f = {k: v for k, v in headers.items() if k.lower() != "content-type"}
        headers_f = {k: v for k, v in headers.items() if k.lower() != "content-type"}
        # F1: POST /produtos/{id}/imagens multipart
        resp_f = _req.post(f"{client.base_url}/produtos/{produto_id}/imagens",
                           headers=headers_f, files=files, timeout=30)
        results.append({"abordagem": "F1_multipart_post_imagens", "status": resp_f.status_code, "response": resp_f.text[:500]})
        # F2: POST /produtos/{id}/fotos multipart
        img_resp2 = _req.get(image_url, timeout=20)
        files2 = {"imagem": ("kit_6.4.jpg", _io.BytesIO(img_resp2.content), "image/jpeg")}
        resp_f2 = _req.post(f"{client.base_url}/produtos/{produto_id}/fotos",
                            headers=headers_f, files=files2, timeout=30)
        results.append({"abordagem": "F2_multipart_post_fotos", "status": resp_f2.status_code, "response": resp_f2.text[:500]})
        # F3: POST /imagens (root-level)
        img_resp3 = _req.get(image_url, timeout=20)
        files3 = {"imagem": ("kit_6.4.jpg", _io.BytesIO(img_resp3.content), "image/jpeg")}
        resp_f3 = _req.post(f"{client.base_url}/imagens", headers=headers_f, files=files3, timeout=30)
        results.append({"abordagem": "F3_root_imagens_multipart", "status": resp_f3.status_code, "response": resp_f3.text[:500]})
        # F4: POST /produtos/{id}/imagens with JSON {url: ...}
        resp_f4 = _req.post(f"{client.base_url}/produtos/{produto_id}/imagens",
                            headers=headers, json={"url": image_url}, timeout=30)
        results.append({"abordagem": "F4_json_url_post_imagens", "status": resp_f4.status_code, "response": resp_f4.text[:500]})
        # F5: PATCH /produtos/{id} with json = {midia...}
        resp_f5 = _req.patch(f"{client.base_url}/produtos/{produto_id}",
                             headers=headers,
                             json={"midia": {"imagens": {"externas": [{"link": image_url}]}}},
                             timeout=30)
        results.append({"abordagem": "F5_patch_midia_nested", "status": resp_f5.status_code, "response": resp_f5.text[:500]})
    else:
        results.append({"abordagem": "F_multipart_upload", "status": "download_failed", "response": str(img_resp.status_code)})

    # Abordagem G: POST /uploads com multipart (padrão de upload antes de associar)
    if img_resp.status_code == 200:
        import io as _io2
        headers_g = {k: v for k, v in headers.items() if k.lower() != "content-type"}
        files_g = {"file": ("image.jpg", _io2.BytesIO(img_resp.content), "image/jpeg")}
        resp_g = _req.post(f"{client.base_url}/uploads", headers=headers_g, files=files_g, timeout=30)
        results.append({"abordagem": "G1_post_uploads_file", "status": resp_g.status_code, "response": resp_g.text[:500]})

        # G2: base64 no internas
        import base64 as _b64
        img_data = _b64.b64encode(img_resp.content).decode("utf-8")
        EXCLUIR_G = {"camposCustomizados", "variacoes", "variacao", "estoque", "actionEstoque"}
        patch_g = {k: v for k, v in data_antes.items() if k not in EXCLUIR_G}
        patch_g["id"] = produto_id
        patch_g["midia"] = {
            "video": {"url": ""},
            "imagens": {
                "externas": [],
                "internas": [{"imagem": img_data, "tipo": "jpg"}],
                "imagensURL": []
            }
        }
        resp_g2 = _req.put(f"{client.base_url}/produtos/{produto_id}", headers=headers, json=patch_g, timeout=60)
        _time.sleep(2)
        after_g2 = client._get(f"/produtos/{produto_id}").get("data", {}).get("midia", {})
        results.append({"abordagem": "G2_base64_internas", "status": resp_g2.status_code,
                        "response": resp_g2.text[:500], "midia_depois": after_g2})

        # G3: POST /imagens com JSON {link: url}
        resp_g3 = _req.post(f"{client.base_url}/imagens", headers=headers,
                            json={"link": image_url}, timeout=30)
        results.append({"abordagem": "G3_post_imagens_json_link", "status": resp_g3.status_code, "response": resp_g3.text[:500]})

    return {
        "produto_id": produto_id,
        "image_url": image_url,
        "externas_antes": externas_antes,
        "resultados": results,
    }


@app.post("/bling/produto/buscar-por-sku")
def bling_buscar_por_sku(payload: BuscaSkuPayload):
    """Busca um produto no Bling pelo SKU/código. Sem auth."""
    if not BlingClient:
        raise HTTPException(status_code=500, detail="bling_client.py não encontrado.")
    client = BlingClient()
    try:
        raw_resp = client._get("/produtos", params={"codigo": payload.sku.strip(), "limite": 5})
        items = raw_resp.get("data", [])
        results = []
        for item in items:
            p = item if isinstance(item, dict) else {}
            results.append({
                "id": p.get("id"),
                "codigo": p.get("codigo"),
                "nome": p.get("nome"),
                "tipo": p.get("tipo"),
                "imagens": p.get("imagens", []),
            })
        return {"sku": payload.sku, "total": len(results), "produtos": results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/integracao/preview")
def integracao_preview(payload: IntegracaoPayload):
    regras = carregar_regras(apenas_ativas=True)
    if not regras:
        raise HTTPException(status_code=400, detail="Nenhuma regra cadastrada. Importe a Aba2 primeiro.")
    if (payload.criterio or "sku").strip().lower() != "sku":
        raise HTTPException(status_code=400, detail="A precificação integrada aceita apenas busca por SKU. Use criterio='sku'.")
    try:
        # Embalagem: usa valor do payload ou, se não enviado, lê embalagem_padrao da config
        _cfg_int = carregar_integracao_cfg()
        embalagem_efetiva = payload.embalagem if payload.embalagem is not None else float(_cfg_int.get("embalagem_padrao") or 0)

        # Flags de API em tempo real — lê do config quando não especificado no payload
        _intel = dict(payload.score_config or {})
        for _flag in ("ml_api_real", "amazon_api_real", "shopee_api_real"):
            if _flag not in _intel:
                _intel[_flag] = bool(_cfg_int.get(_flag, True))

        # Carrega score_config SIE salvo (ajuste automático de margem por score)
        _sie_score_cfg = _load_json(_MOD_DIR / "sie_score_config.json", {})
        _score_config_efetivo = payload.score_config or (_sie_score_cfg if _sie_score_cfg.get("ajuste_ativo") else None)
        resultado = montar_precificacao_bling(
            regras=regras, criterio="sku", valor_busca=payload.valor_busca, embalagem=embalagem_efetiva, imposto=payload.imposto,
            quantidade=payload.quantidade, objetivo=payload.objetivo, tipo_alvo=payload.tipo_alvo, valor_alvo=payload.valor_alvo,
            peso_override=payload.peso_override, intelligence_config=_intel, score_config=_score_config_efetivo,
            modo_aprovacao=payload.modo_aprovacao,
            preco_compra_anterior_bling=payload.preco_compra_anterior_bling, modo_preco_virtual=payload.modo_preco_virtual,
            acrescimo_percentual=payload.acrescimo_percentual, acrescimo_nominal=payload.acrescimo_nominal, preco_manual=payload.preco_manual,
            arredondamento=payload.arredondamento, regra_estoque=carregar_cfg().get("regra_estoque"),
        )
        if resultado.get("erro"):
            preview = {"ok":False,"criterio_usado":"sku","produto":resultado.get("produto_bling") or {},"melhor_canal":"","modo_aprovacao":payload.modo_aprovacao,"marketplaces":{},"auditoria":resultado,"raw":resultado}
            preview["diagnostico"] = _diagnostico_preview(preview)
            preview["fila_auto"] = {"adicionado":False,"motivo":preview["diagnostico"]["mensagem"]}
            return preview
        itens = (resultado.get("integracao") or {}).get("itens") or resultado.get("itens_precificacao") or resultado.get("itens") or []
        preview = {"ok":True,"criterio_usado":"sku","produto":resultado.get("produto_bling") or {},"melhor_canal":resultado.get("melhor_canal") or "","modo_aprovacao":payload.modo_aprovacao,"marketplaces":_normalizar_marketplaces(itens or resultado.get("canais", [])),"auditoria":resultado.get("auditoria") or resultado,"raw":resultado}
        diagnostico = _diagnostico_preview(preview)
        preview["diagnostico"] = diagnostico
        valido, motivo = _preview_valido(preview)
        fila_auto = {"adicionado":False,"motivo":""}
        if carregar_cfg().get("fila_auto_ao_calcular", True) and valido:
            sku = preview["auditoria"].get("sku") or preview["produto"].get("codigo") or ""
            if ja_existe_pendente(sku):
                fila_auto = {"adicionado":False,"motivo":"Já existe item pendente equivalente na fila."}
            else:
                item = _montar_item_fila(preview, payload.dict())
                inserir_item_fila(item)
                _append_jsonl(LOG_PATH, {"evento":"fila_auto_preview","item_id":item["id"],"sku":item["sku"],"quando":item["criado_em"]})
                fila_auto = {"adicionado":True,"item_id":item["id"]}
        else:
            fila_auto = {"adicionado":False,"motivo":motivo or diagnostico.get("mensagem") or "Fila automática desativada."}
        preview["fila_auto"] = fila_auto
        logger.info("Precificação: SKU=%s melhor_canal=%s fila=%s", payload.valor_busca, preview.get("melhor_canal"), fila_auto.get("adicionado"))
        return preview
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Falha no preview: {exc}")

@app.get("/fila/lista")
def fila_lista():
    itens = [i for i in carregar_fila() if i.get("status") in ("pendente","incompleto")]
    return {"itens":itens,"stats":stats_fila()}

@app.post("/fila/adicionar")
def fila_adicionar(payload: dict = Body(...)):
    preview = {"ok":payload.get("ok", True),"produto":payload.get("produto_bling") or payload.get("produto") or {},"marketplaces":payload.get("marketplaces") or {},"auditoria":payload.get("auditoria") or {},"raw":payload.get("raw") or {}}
    diag = _diagnostico_preview(preview)
    if not diag.get("ok"): raise HTTPException(status_code=400, detail=f"Preview inválido para fila: {diag.get('mensagem')}")
    sku = (preview.get("auditoria") or {}).get("sku") or (preview.get("produto") or {}).get("codigo") or ""
    if ja_existe_pendente(sku):
        return {"ok":True,"duplicado":True,"message":"Já existe item pendente equivalente na fila.","stats":stats_fila()}
    item = _montar_item_fila(preview, payload.get("payload_original") or payload.get("raw") or {})
    inserir_item_fila(item)
    _append_jsonl(LOG_PATH, {"evento":"fila_adicionar_manual","item_id":item["id"],"sku":item["sku"],"quando":item["criado_em"]})
    return {"ok":True,"item":item,"stats":stats_fila()}

@app.post("/fila/limpar-invalidos")
def fila_limpar_invalidos():
    removidos_n = limpar_invalidos_fila()
    return {"ok":True,"removidos":removidos_n,"stats":stats_fila()}

@app.post("/fila/reset-total")
def fila_reset_total():
    reset_fila()
    return {"ok":True,"message":"Fila completamente limpa","stats":stats_fila()}


_POPULAR_STATE_FILE = DATA_DIR / "popular_estoque_state.json"

def _popular_salvar_estado(state: dict):
    try:
        _POPULAR_STATE_FILE.write_text(json.dumps(state, ensure_ascii=False), encoding="utf-8")
    except Exception as e:
        logger.warning("[POPULAR-FILA] erro ao salvar estado: %s", e)

def _popular_carregar_estado() -> dict:
    try:
        if _POPULAR_STATE_FILE.exists():
            return json.loads(_POPULAR_STATE_FILE.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {}


@app.post("/fila/popular-estoque")
def fila_popular_estoque(lote: int = 50):
    """
    Processa um lote de produtos simples do Bling e insere na fila de aprovação.
    Chame repetidamente até status.concluido=true.
    Persiste progresso em arquivo — sobrevive a reinicializações do Cloud Run.
    """
    import time as _time
    estado = _popular_carregar_estado()

    bc = BlingClient()
    regras = carregar_regras(apenas_ativas=True)
    _sie_score_cfg = _load_json(_MOD_DIR / "sie_score_config.json", {})
    score_cfg = _sie_score_cfg if _sie_score_cfg.get("ajuste_ativo") else None

    # ── Fase 1: coletar SKUs (só na primeira chamada) ──────────────────────────
    if not estado.get("skus"):
        estado = {"skus": [], "cursor": 0, "adicionados": 0, "erros": 0, "concluido": False}
        pagina = 1
        while True:
            try:
                resp = bc.list_products(page=pagina, limit=100)
                itens = resp.get("data") or []
                if not itens:
                    break
                for item in itens:
                    if item.get("situacao") not in ("A", None, ""):
                        continue
                    if item.get("tipo") not in ("P", None):
                        continue
                    if item.get("formato") in ("K",):
                        continue
                    estrutura = item.get("estrutura") or {}
                    if estrutura.get("tipo") in ("M", "K"):
                        continue
                    sku = item.get("codigo") or ""
                    estoque = item.get("estoque") or {}
                    saldo = float(
                        estoque.get("fisico") or estoque.get("saldoFisico")
                        or estoque.get("saldoVirtualTotal") or 0
                    )
                    if sku and saldo > 0:
                        estado["skus"].append({
                            "sku": sku,
                            "nome": item.get("nome", "")[:80],
                            "preco_custo": float(item.get("precoCusto") or 0),
                        })
                pagina += 1
                _time.sleep(0.25)
            except Exception as e:
                logger.warning("[POPULAR-FILA] erro coleta pg %s: %s", pagina, e)
                break
        _popular_salvar_estado(estado)
        logger.info("[POPULAR-FILA] Coletados %d produtos simples.", len(estado["skus"]))

    skus      = estado["skus"]
    cursor    = estado.get("cursor", 0)
    total     = len(skus)
    adicionados = estado.get("adicionados", 0)
    erros     = estado.get("erros", 0)

    if cursor >= total:
        estado["concluido"] = True
        _popular_salvar_estado(estado)
        return {"ok": True, "concluido": True, "total": total,
                "adicionados": adicionados, "erros": erros,
                "msg": f"Concluído: {adicionados} adicionados, {erros} erros."}

    # ── Fase 2: processar lote ─────────────────────────────────────────────────
    fim = min(cursor + lote, total)
    for item in skus[cursor:fim]:
        sku = item["sku"]
        try:
            if ja_existe_pendente(sku):
                continue
            preco_custo = item.get("preco_custo") or None
            resultado = montar_precificacao_bling(
                regras=regras, criterio="sku", valor_busca=sku,
                embalagem="", imposto=0, quantidade=1,
                objetivo="margem", tipo_alvo="percentual", valor_alvo=0,
                score_config=score_cfg, regra_estoque=carregar_cfg().get("regra_estoque"),
                preco_compra_anterior_bling=preco_custo,
                peso_override=0.3,
            )
            if resultado.get("erro"):
                erros += 1
                continue
            itens_prec = (resultado.get("integracao") or {}).get("itens") or resultado.get("itens_precificacao") or resultado.get("itens") or []
            preview = {
                "ok": True, "criterio_usado": "sku",
                "produto": resultado.get("produto_bling") or {},
                "melhor_canal": resultado.get("melhor_canal") or "",
                "modo_aprovacao": "manual",
                "marketplaces": _normalizar_marketplaces(itens_prec),
                "auditoria": resultado.get("auditoria") or resultado,
                "raw": resultado,
            }
            preview["diagnostico"] = _diagnostico_preview(preview)
            item_fila = _montar_item_fila(preview, {"valor_busca": sku, "criterio": "sku"})
            inserir_item_fila(item_fila)
            adicionados += 1
        except Exception as exc:
            logger.warning("[POPULAR-FILA] SKU=%s erro: %s", sku, exc)
            erros += 1

    estado["cursor"]     = fim
    estado["adicionados"] = adicionados
    estado["erros"]      = erros
    estado["concluido"]  = fim >= total
    _popular_salvar_estado(estado)

    return {
        "ok":         True,
        "concluido":  estado["concluido"],
        "total":      total,
        "processados": fim,
        "adicionados": adicionados,
        "erros":      erros,
        "msg":        f"{fim}/{total} processados — {adicionados} na fila",
    }


@app.get("/fila/popular-estoque/status")
def fila_popular_estoque_status():
    estado = _popular_carregar_estado()
    if not estado:
        return {"concluido": False, "total": 0, "processados": 0, "adicionados": 0, "erros": 0, "msg": "Nenhum processo iniciado."}
    return {
        "concluido":   estado.get("concluido", False),
        "total":       len(estado.get("skus", [])),
        "processados": estado.get("cursor", 0),
        "adicionados": estado.get("adicionados", 0),
        "erros":       estado.get("erros", 0),
        "msg":         f"{estado.get('cursor',0)}/{len(estado.get('skus',[]))} processados — {estado.get('adicionados',0)} na fila",
    }


@app.post("/fila/popular-estoque/reset")
def fila_popular_estoque_reset():
    """Limpa o estado do processo para iniciar do zero."""
    try:
        _POPULAR_STATE_FILE.unlink(missing_ok=True)
    except Exception:
        pass
    return {"ok": True, "msg": "Estado resetado."}


# ══════════════════════════════════════════════════════════════════════════════
# FILA DE CONFERÊNCIA DE PREÇO DE CUSTO
# ══════════════════════════════════════════════════════════════════════════════

_FILA_CUSTO_FILE = DATA_DIR / "fila_custo.json"


def _fila_custo_carregar() -> list:
    try:
        if _FILA_CUSTO_FILE.exists():
            return json.loads(_FILA_CUSTO_FILE.read_text(encoding="utf-8"))
    except Exception:
        pass
    return []


def _fila_custo_salvar(itens: list):
    _FILA_CUSTO_FILE.write_text(json.dumps(itens, ensure_ascii=False, indent=2), encoding="utf-8")


@app.get("/fila-custo", response_class=HTMLResponse)
def fila_custo_page():
    html = BASE_DIR / "pages" / "fila_custo.html"
    if not html.exists():
        raise HTTPException(status_code=404, detail="fila_custo.html não encontrado.")
    return HTMLResponse(html.read_text(encoding="utf-8"))


@app.post("/fila-custo/popular")
def fila_custo_popular(lote: int = 100):
    """
    Varre o Bling e cria a fila de conferência de preço de custo.
    Lotes de 100 produtos por chamada — chame até popular=true.
    """
    import time as _time

    estado = _popular_carregar_estado()
    skus_com_estoque = estado.get("skus", [])

    # Se não tem estado ainda, precisa rodar /fila/popular-estoque primeiro
    if not skus_com_estoque:
        # Coleta direto do Bling
        bc = BlingClient()
        pagina = 1
        while True:
            try:
                resp = bc.list_products(page=pagina, limit=100)
                itens = resp.get("data") or []
                if not itens:
                    break
                for item in itens:
                    if item.get("situacao") not in ("A", None, ""):
                        continue
                    if item.get("tipo") not in ("P", None):
                        continue
                    if item.get("formato") in ("K",):
                        continue
                    estrutura = item.get("estrutura") or {}
                    if estrutura.get("tipo") in ("M", "K"):
                        continue
                    estoque_raw = item.get("estoque") or {}
                    saldo = float(
                        estoque_raw.get("fisico") or estoque_raw.get("saldoFisico")
                        or estoque_raw.get("saldoVirtualTotal") or 0
                    )
                    sku = item.get("codigo") or ""
                    if sku and saldo > 0:
                        skus_com_estoque.append({
                            "id": item.get("id"),
                            "sku": sku,
                            "nome": item.get("nome", "")[:100],
                            "preco_custo": float(item.get("precoCusto") or 0),
                            "saldo": saldo,
                        })
                pagina += 1
                _time.sleep(0.25)
            except Exception as e:
                logger.warning("[FILA-CUSTO] erro coleta pg %s: %s", pagina, e)
                break
    else:
        skus_com_estoque = [
            {"id": None, "sku": s["sku"], "nome": s["nome"],
             "preco_custo": s.get("preco_custo", 0), "saldo": 0}
            for s in skus_com_estoque
        ]

    # Monta itens da fila (sem duplicar)
    existentes = {i["sku"] for i in _fila_custo_carregar()}
    novos = [s for s in skus_com_estoque if s["sku"] not in existentes]
    fila = _fila_custo_carregar() + [
        {
            "id":          s.get("id"),
            "sku":         s["sku"],
            "nome":        s.get("nome", ""),
            "preco_custo": s.get("preco_custo", 0),
            "saldo":       s.get("saldo", 0),
            "status":      "pendente",
            "novo_custo":  None,
        }
        for s in novos
    ]
    _fila_custo_salvar(fila)
    return {"ok": True, "total": len(fila), "novos": len(novos),
            "msg": f"{len(fila)} produtos na fila de custo."}


@app.get("/fila-custo/lista")
def fila_custo_lista(status: str = "pendente", busca: str = ""):
    itens = _fila_custo_carregar()
    if status:
        itens = [i for i in itens if i.get("status") == status]
    if busca:
        b = busca.lower()
        itens = [i for i in itens if b in i.get("sku", "").lower() or b in i.get("nome", "").lower()]
    return {"itens": itens, "total": len(itens)}


@app.post("/fila-custo/atualizar")
def fila_custo_atualizar(payload: dict = Body(...)):
    """Atualiza o preço de custo de um produto no Bling e marca como revisado na fila."""
    sku        = payload.get("sku", "")
    novo_custo = float(payload.get("novo_custo") or 0)
    if not sku:
        raise HTTPException(status_code=400, detail="SKU obrigatório.")
    if novo_custo <= 0:
        raise HTTPException(status_code=400, detail="Novo custo deve ser maior que zero.")

    fila = _fila_custo_carregar()
    item = next((i for i in fila if i["sku"] == sku), None)
    if not item:
        raise HTTPException(status_code=404, detail="SKU não encontrado na fila de custo.")

    # Atualiza no Bling
    try:
        bc = BlingClient()
        # Busca produto pelo SKU para pegar o ID e estrutura completa
        prod = bc.get_product_by_sku(sku) if hasattr(bc, "get_product_by_sku") else None
        if not prod:
            res_busca = bc._get("/produtos", params={"codigo": sku, "limite": 1})
            prods = (res_busca.get("data") or [])
            if prods:
                prod_id = prods[0].get("id")
                prod = bc.get_product(prod_id) if prod_id else None

        if not prod:
            raise HTTPException(status_code=404, detail=f"Produto {sku} não encontrado no Bling.")

        prod_id = prod.get("id") or item.get("id")
        if not prod_id:
            raise HTTPException(status_code=400, detail="ID do produto não disponível.")

        # Monta patch mínimo com precoCusto
        patch = {"precoCusto": round(novo_custo, 2)}
        fornecedor = prod.get("fornecedor")
        if isinstance(fornecedor, dict) and fornecedor.get("id"):
            patch["fornecedor"] = {**fornecedor, "precoCusto": round(novo_custo, 2), "precoCompra": round(novo_custo, 2)}

        bc.update_product(int(prod_id), patch)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Erro ao atualizar Bling: {exc}")

    # Atualiza fila
    item["preco_custo"] = novo_custo
    item["novo_custo"]  = novo_custo
    item["status"]      = "revisado"
    _fila_custo_salvar(fila)

    return {"ok": True, "sku": sku, "novo_custo": novo_custo}


@app.post("/fila-custo/pular")
def fila_custo_pular(payload: dict = Body(...)):
    sku = payload.get("sku", "")
    fila = _fila_custo_carregar()
    item = next((i for i in fila if i["sku"] == sku), None)
    if item:
        item["status"] = "pulado"
        _fila_custo_salvar(fila)
    return {"ok": True}


@app.post("/fila-custo/reset")
def fila_custo_reset():
    _FILA_CUSTO_FILE.unlink(missing_ok=True)
    return {"ok": True, "msg": "Fila de custo limpa."}


@app.get("/fila/links/{sku}")
async def fila_links(sku: str):
    """Retorna links diretos para edição do produto em cada marketplace."""
    import time
    # Verifica cache
    now = time.time()
    if sku in _links_cache and now - _links_cache_ttl.get(sku, 0) < _LINKS_CACHE_SECONDS:
        return {"ok": True, "sku": sku, "links": _links_cache[sku], "cached": True}
    links = {}
    try:
        client = BlingClient()
        busca = client.get_product_by_sku(sku)
        if busca.get("encontrado"):
            produto_id = busca.get("produto", {}).get("id")
            if produto_id:
                links["bling"] = f"https://www.bling.com.br/produtos.php#edit/{produto_id}"
    except Exception:
        pass
    # ML - busca MLB IDs pelo SKU
    try:
        import json as _j, requests as _rq
        _ml_tokens = _j.loads((BASE_DIR / "data" / "ml_tokens.json").read_text(encoding="utf-8"))
        _ml_token = _ml_tokens.get("access_token", "")
        _ml_r = _rq.get(
            f"https://api.mercadolibre.com/users/733168645/items/search?seller_custom_field={sku}",
            headers={"Authorization": f"Bearer {_ml_token}"},
            timeout=8
        )
        if _ml_r.status_code == 200:
            items = _ml_r.json().get("results", [])
            if items:
                links["ml"] = f"https://www.mercadolivre.com.br/anuncios/{items[0]}/editar"
    except Exception:
        pass
    # Amazon - link com SKU no inventário
    links["amazon"] = f"https://sellercentral.amazon.com.br/myinventory/inventory?searchField=all&searchTerm={sku}"
    # Shopify - busca product_id pelo SKU via API
    try:
        import requests as _req, json as _json
        _cfg = _json.loads((BASE_DIR / "data" / "shopify_config.json").read_text(encoding="utf-8"))
        _shop = _cfg.get("shop_url", "pknw4n-eg.myshopify.com")
        _token = _cfg.get("access_token", "")
        _r = _req.get(
            f"https://{_shop}/admin/api/2024-01/variants.json?sku={sku}&limit=1",
            headers={"X-Shopify-Access-Token": _token},
            timeout=8
        )
        if _r.status_code == 200:
            variants = _r.json().get("variants", [])
            if variants:
                product_id = variants[0].get("product_id")
                if product_id:
                    links["shopify"] = f"https://admin.shopify.com/store/pknw4n-eg/products/{product_id}"
    except Exception:
        pass
    # Salva no cache
    _links_cache[sku] = links
    _links_cache_ttl[sku] = time.time()
    return {"ok": True, "sku": sku, "links": links}

@app.post("/fila/aprovar/{item_id}")
def fila_aprovar(item_id: str):
    item = buscar_item_fila(item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item não encontrado na fila.")
    if item.get("status") != "pendente":
        raise HTTPException(status_code=400, detail=f"Item já com status '{item.get("status")}'.")
    if not BlingClient or not aplicar_precos_multicanal:
        raise HTTPException(status_code=500, detail="Integração de aplicação no Bling indisponível.")
    item_com_gordura = _aplicar_gordura_no_item(item)
    try:
        client = BlingClient()
        resultado = aplicar_precos_multicanal(client, item_com_gordura)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Falha ao aplicar preços no Bling: {exc}")

    # ── Tenta aplicar na Shopee se SKU estiver mapeado ──
    try:
        sku = item.get("sku") or (item.get("produto_bling") or {}).get("codigo") or ""
        marketplaces = item.get("marketplaces") or {}
        shopee_resultado = aplicar_preco_shopee_por_sku(sku, marketplaces)
        if shopee_resultado:
            resultado["shopee"] = shopee_resultado
            if shopee_resultado.get("success"):
                logger.info("Shopee: preço aplicado sku=%s item_id=%s preco=%.2f",
                            sku, shopee_resultado.get("item_id"), shopee_resultado.get("preco_aplicado", 0))
            else:
                logger.warning("Shopee: falha ao aplicar sku=%s motivo=%s", sku, shopee_resultado.get("motivo"))
    except Exception as exc_shopee:
        logger.warning("Shopee: exceção ao aplicar preço sku=%s: %s", item.get("sku"), exc_shopee)

    agora = datetime.utcnow().isoformat()
    atualizar_status_fila(item_id, "aprovado", resultado=resultado)
    _append_jsonl(LOG_PATH, {"evento": "fila_aprovado", "item_id": item_id, "quando": agora})
    logger.info("Item aprovado: id=%s sku=%s estrategia=%s", item_id, item.get("sku"), resultado.get("estrategia"))
    return {"ok": True, "message": "Preços aplicados no Bling.", "resultado": resultado, "stats": stats_fila()}



@app.post("/fila/completar/{item_id}")
async def fila_completar(item_id: str, request: Request):
    """Salva peso ou custo localmente para produto incompleto e agenda recalculo."""
    import json as _json
    data = await request.json()
    tipo = data.get("tipo")  # "peso" ou "custo"
    valor = float(data.get("valor", 0))
    if tipo not in ("peso", "custo") or valor <= 0:
        raise HTTPException(status_code=400, detail="Tipo ou valor inv\u00e1lido.")
    item = buscar_item_fila(item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item n\u00e3o encontrado na fila.")
    sku = item.get("sku")
    if not sku:
        raise HTTPException(status_code=400, detail="SKU n\u00e3o encontrado no item.")
    try:
        if tipo == "peso":
            _override_path = BASE_DIR / "data" / "peso_override.json"
            _overrides = {}
            if _override_path.exists():
                try:
                    _overrides = _json.loads(_override_path.read_text(encoding="utf-8"))
                except Exception:
                    pass
            _overrides[str(sku)] = valor
            _override_path.write_text(_json.dumps(_overrides, ensure_ascii=False, indent=2), encoding="utf-8")
            logger.info("Peso override salvo: SKU %s = %s kg", sku, valor)
        else:
            _override_path = BASE_DIR / "data" / "custo_override.json"
            _overrides = {}
            if _override_path.exists():
                try:
                    _overrides = _json.loads(_override_path.read_text(encoding="utf-8"))
                except Exception:
                    pass
            # Salva override para o SKU principal
            _overrides[str(sku)] = {"custo": valor, "origem": "manual_fila"}
            # Se for composição, salva também para o componente e tenta atualizar no Bling
            dados_inc = item.get("dados_incompletos") or {}
            comps_sem_custo = dados_inc.get("componentes_sem_custo") or []
            if comps_sem_custo:
                comp = comps_sem_custo[0]
                comp_sku = comp.get("sku")
                comp_id = comp.get("id")
                if comp_sku:
                    _overrides[str(comp_sku)] = {"custo": valor, "origem": "manual_fila"}
                    logger.info("Custo override componente: SKU %s = R$%s", comp_sku, valor)
                # Tenta atualizar custo do componente no Bling via fornecedor
                if comp_id and BlingClient:
                    try:
                        _bling = BlingClient()
                        _prod_comp = _bling.get_product(int(comp_id))
                        _forn = _prod_comp.get("fornecedor") or {}
                        if _forn.get("id"):
                            _forn_patch = {**_forn, "precoCusto": valor, "precoCompra": valor}
                            _patch = {k: v for k, v in _prod_comp.items() if k not in ("estoque", "variacoes", "estrutura", "midia")}
                            _patch["fornecedor"] = _forn_patch
                            _bling.update_product(int(comp_id), _patch)
                            logger.info("Custo componente atualizado no Bling: id=%s valor=%s", comp_id, valor)
                    except Exception as _e:
                        logger.warning("Erro ao atualizar componente %s no Bling: %s", comp_id, _e)
            _override_path.write_text(_json.dumps(_overrides, ensure_ascii=False, indent=2), encoding="utf-8")
            logger.info("Custo override salvo: SKU %s = R$%s", sku, valor)
        atualizar_status_fila(item_id, "rejeitado", resultado={"motivo": f"{tipo} preenchido: {valor}"})
        return {"ok": True, "mensagem": f"{tipo.capitalize()} salvo. O produto ser\u00e1 recalculado no pr\u00f3ximo ciclo do scheduler."}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Erro ao completar produto %s: %s", sku, e)
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/fila/exportar-sem-dados")
def fila_exportar_sem_dados():
    """Varre o Bling e retorna CSV com todos os produtos sem peso ou custo."""
    from fastapi.responses import StreamingResponse
    import io as _io, csv as _csv
    if not BlingClient:
        raise HTTPException(status_code=500, detail="BlingClient não disponível.")
    client = BlingClient()
    linhas = [["id_bling","sku","nome","tipo_produto","formato","peso_atual","peso_novo","custo_atual","custo_novo"]]
    pagina = 1
    while True:
        try:
            resp = client._get("/produtos", params={"situacao": "A", "pagina": pagina, "limite": 100})
        except Exception:
            break
        items = resp.get("data") or []
        if not items:
            break
        for p in items:
            peso = float(p.get("pesoLiquido") or p.get("pesoBruto") or p.get("peso") or 0)
            custo = float(p.get("precoCusto") or p.get("precoCompra") or 0)
            fmt = (p.get("formato") or "").upper()
            eh_kit = fmt == "E"
            sem_peso = peso <= 0
            sem_custo = custo <= 0 and not eh_kit
            if sem_peso or sem_custo:
                tipo = "kit_composicao" if eh_kit else "simples"
                nome = (p.get("nome") or "").replace('"', '""')
                custo_novo = "(calculado pelos componentes)" if eh_kit else ""
                linhas.append([
                    p.get("id", ""), p.get("codigo", ""), f'"{nome}"',
                    tipo, fmt,
                    f"{peso:.3f}", "",
                    f"{custo:.2f}", custo_novo,
                ])
        if len(items) < 100:
            break
        pagina += 1
    buf = _io.StringIO()
    buf.write("sep=;\r\n")
    writer = _csv.writer(buf, delimiter=";")
    for row in linhas:
        writer.writerow(row)
    content = "﻿" + buf.getvalue()
    return StreamingResponse(
        _io.BytesIO(content.encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=produtos_sem_dados_{__import__('datetime').date.today()}.csv"},
    )


_IMPORT_JOB: dict = {"status": "idle", "total": 0, "ok": 0, "erros": 0, "ignorados": 0, "log": []}

def _run_importar_correcoes(linhas: list[dict]):
    import json as _json, time as _time
    global _IMPORT_JOB
    _IMPORT_JOB.update({"status": "running", "total": len(linhas), "ok": 0, "erros": 0, "ignorados": 0, "log": []})
    peso_path  = BASE_DIR / "data" / "peso_override.json"
    custo_path = BASE_DIR / "data" / "custo_override.json"
    peso_ov  = _json.loads(peso_path.read_text(encoding="utf-8"))  if peso_path.exists()  else {}
    custo_ov = _json.loads(custo_path.read_text(encoding="utf-8")) if custo_path.exists() else {}
    bling = BlingClient() if BlingClient else None

    for row in linhas:
        id_bling = (row.get("id_bling") or "").strip()
        sku      = (row.get("sku") or "").strip()
        tipo     = (row.get("tipo_produto") or "simples").strip().lower()
        peso_str = (row.get("peso_novo") or "").strip().replace(",", ".")
        custo_str= (row.get("custo_novo") or "").strip().replace(",", ".")
        if not sku:
            _IMPORT_JOB["ignorados"] += 1
            continue

        peso_val  = float(peso_str)  if peso_str  and _is_number(peso_str)  else None
        custo_val = float(custo_str) if custo_str and _is_number(custo_str) and tipo == "simples" else None

        if peso_val is None and custo_val is None:
            _IMPORT_JOB["ignorados"] += 1
            continue

        # Atualiza overrides locais
        if peso_val and peso_val > 0:
            peso_ov[sku] = peso_val
        if custo_val and custo_val > 0:
            custo_ov[sku] = {"custo": custo_val, "origem": "importacao_planilha"}

        # Atualiza no Bling via GET + PUT
        if bling and id_bling:
            try:
                prod = bling.get_product(int(id_bling))
                if prod:
                    patch = {}
                    if peso_val and peso_val > 0:
                        patch["pesoLiquido"] = peso_val
                    if custo_val and custo_val > 0:
                        patch["precoCusto"] = custo_val
                    if patch:
                        merged = {**prod, **patch}
                        # Remover campos que a API rejeita no PUT (padrão bling_update_engine)
                        for _k in ("camposCustomizados", "customFields", "midias", "anexos", "producao", "tipoEstoque"):
                            merged.pop(_k, None)
                        # Corrigir tipoEstoque dentro de estrutura
                        if isinstance(merged.get("estrutura"), dict):
                            est = dict(merged["estrutura"])
                            if est.get("tipoEstoque") not in ("V", "F"):
                                est["tipoEstoque"] = "V"
                            merged["estrutura"] = est
                        bling.update_product(int(id_bling), merged)
                        _IMPORT_JOB["ok"] += 1
                        _IMPORT_JOB["log"].append(f"OK {sku}: {patch}")
                        _time.sleep(0.25)
                else:
                    _IMPORT_JOB["erros"] += 1
            except Exception as _e:
                logger.warning("Bling update id=%s sku=%s: %s", id_bling, sku, _e)
                _IMPORT_JOB["erros"] += 1
                _IMPORT_JOB["log"].append(f"ERRO {sku}: {_e}")
        else:
            _IMPORT_JOB["ok"] += 1

    peso_path.write_text(_json.dumps(peso_ov,  ensure_ascii=False, indent=2), encoding="utf-8")
    custo_path.write_text(_json.dumps(custo_ov, ensure_ascii=False, indent=2), encoding="utf-8")
    _IMPORT_JOB["status"] = "done"
    logger.info("Importação concluída: ok=%d erros=%d ignorados=%d", _IMPORT_JOB["ok"], _IMPORT_JOB["erros"], _IMPORT_JOB["ignorados"])

def _is_number(s: str) -> bool:
    try: float(s); return True
    except: return False


@app.post("/fila/importar-correcoes")
async def fila_importar_correcoes(request: Request):
    """Recebe CSV com correções em lote e inicia job em background para atualizar o Bling."""
    import io as _io, csv as _csv, threading as _thr
    if _IMPORT_JOB.get("status") == "running":
        return {"ok": False, "detail": "Já existe um import em andamento. Aguarde."}
    body = await request.body()
    text = body.decode("utf-8-sig")
    # Pular linha sep= se presente
    if text.startswith("sep="):
        text = text[text.index("\n")+1:]
    reader = _csv.DictReader(_io.StringIO(text), delimiter=";")
    linhas = [row for row in reader if (row.get("peso_novo") or row.get("custo_novo") or "").strip()]
    if not linhas:
        return {"ok": False, "detail": "Nenhuma linha com peso_novo ou custo_novo preenchido."}
    _thr.Thread(target=_run_importar_correcoes, args=(linhas,), daemon=True).start()
    return {"ok": True, "iniciado": True, "total": len(linhas), "mensagem": f"{len(linhas)} produtos enfileirados. Acompanhe em /fila/importar-correcoes/status"}


@app.get("/fila/importar-correcoes/status")
def fila_importar_status():
    return _IMPORT_JOB


@app.post("/fila/rejeitar/{item_id}")
def fila_rejeitar(item_id: str, payload: dict = Body(default={})):
    item = buscar_item_fila(item_id)
    if not item: raise HTTPException(status_code=404, detail="Item não encontrado na fila.")
    agora = datetime.utcnow().isoformat()
    motivo = payload.get("motivo") or "Rejeitado manualmente."
    atualizar_status_fila(item_id, "rejeitado", resultado={"motivo": motivo})
    _append_jsonl(LOG_PATH, {"evento":"fila_rejeitado","item_id":item_id,"quando":agora,"motivo":motivo})
    logger.info("Item rejeitado: id=%s sku=%s motivo=%s", item_id, item.get("sku"), motivo)
    return {"ok":True,"message":"Item marcado como rejeitado.","stats":stats_fila()}


@app.get("/fila/stats-detalhados")
def fila_stats_detalhados():
    """
    Retorna stats analíticos da fila sem carregar todos os itens.
    Inclui distribuição por canal, faixa de margem, e piores/melhores casos.
    """
    itens = carregar_fila()
    pendentes = [i for i in itens if i.get("status") in ("pendente", "incompleto")]

    canais: dict = {}
    margens: list = []
    for i in pendentes:
        mp = i.get("marketplaces") or {}
        for canal_key, dados in mp.items():
            if not isinstance(dados, dict): continue
            label = dados.get("label") or canal_key
            canais[label] = canais.get(label, 0) + 1
            m = float(dados.get("margem") or 0)
            if m != 0:
                margens.append(m)

    dist_margem = {
        "negativa": sum(1 for m in margens if m < 0),
        "0_a_10":   sum(1 for m in margens if 0 <= m < 10),
        "10_a_20":  sum(1 for m in margens if 10 <= m < 20),
        "acima_20": sum(1 for m in margens if m >= 20),
    }
    avg_margem = round(sum(margens) / len(margens), 2) if margens else 0

    # 10 mais antigos na fila (maior risco de desatualização)
    mais_antigos = sorted(
        [i for i in pendentes if i.get("criado_em")],
        key=lambda x: str(x.get("criado_em", "")),
    )[:10]

    return {
        "total_pendentes": len(pendentes),
        "canais": dict(sorted(canais.items(), key=lambda x: -x[1])),
        "dist_margem": dist_margem,
        "avg_margem": avg_margem,
        "mais_antigos": [{"id": i.get("id"), "sku": i.get("sku"), "nome": str(i.get("nome",""))[:50], "criado_em": i.get("criado_em")} for i in mais_antigos],
    }


@app.post("/fila/aprovar-lote")
async def fila_aprovar_lote(request: Request):
    """
    Aprova um lote de itens por IDs.
    Body: {"ids": ["id1", "id2", ...]}
    Processa até 50 por chamada. Retorna {ok, aprovados, erros, [detalhes]}.
    """
    if not BlingClient or not aplicar_precos_multicanal:
        raise HTTPException(status_code=500, detail="Integração Bling indisponível.")
    body = await request.json()
    ids = body.get("ids", [])[:50]  # máx 50 por vez
    aprovados = 0
    erros = []
    for item_id in ids:
        try:
            item = buscar_item_fila(item_id)
            if not item or item.get("status") != "pendente":
                continue
            item_com_gordura = _aplicar_gordura_no_item(item)
            client = BlingClient()
            resultado = aplicar_precos_multicanal(client, item_com_gordura)
            atualizar_status_fila(item_id, "aprovado", resultado=resultado)
            _append_jsonl(LOG_PATH, {"evento": "fila_aprovado_lote", "item_id": item_id, "quando": datetime.utcnow().isoformat()})
            aprovados += 1
        except Exception as exc:
            erros.append({"id": item_id, "erro": str(exc)})
    return {"ok": True, "aprovados": aprovados, "erros": len(erros), "detalhes_erros": erros, "stats": stats_fila()}


@app.post("/fila/rejeitar-lote")
async def fila_rejeitar_lote(request: Request):
    """Rejeita um lote de itens. Body: {"ids": [...], "motivo": "..."}"""
    body = await request.json()
    ids = body.get("ids", [])
    motivo = body.get("motivo") or "Rejeitado em lote."
    count = 0
    for item_id in ids:
        item = buscar_item_fila(item_id)
        if item and item.get("status") == "pendente":
            atualizar_status_fila(item_id, "rejeitado", resultado={"motivo": motivo})
            count += 1
    return {"ok": True, "rejeitados": count, "stats": stats_fila()}

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# FASE 5 â€” Integração Comercial
# Cole este bloco no app.py logo antes de:
#   
def _aplicar_gordura_no_item(item: dict) -> dict:
    import copy
    item_mod = copy.deepcopy(item)
    cfg = carregar_integracao_cfg()
    gordura_por_canal = cfg.get("gordura_por_canal", {})
    arredondamento = str(cfg.get("arredondamento", "90"))
    marketplaces = item_mod.get("marketplaces", {})
    if isinstance(marketplaces, dict):
        for canal_key, dados in marketplaces.items():
            if not isinstance(dados, dict): continue
            preco_calculado = float(dados.get("preco_promocional") or dados.get("preco_final") or dados.get("preco") or 0)
            if preco_calculado <= 0: continue
            gordura = _buscar_gordura_canal(canal_key, gordura_por_canal)
            preco_virtual = calcular_preco_virtual(preco_calculado, gordura, arredondamento)
            dados["preco_promocional"] = round(preco_calculado, 2)
            dados["preco"] = preco_virtual
            dados["preco_virtual"] = preco_virtual
            dados["gordura_aplicada"] = gordura
    itens_lista = item_mod.get("itens", [])
    if isinstance(itens_lista, list):
        for it in itens_lista:
            if not isinstance(it, dict): continue
            canal_key = it.get("canal", "")
            preco_calculado = float(it.get("preco_promocional") or it.get("preco_final") or it.get("preco") or 0)
            if preco_calculado <= 0: continue
            gordura = _buscar_gordura_canal(canal_key, gordura_por_canal)
            preco_virtual = calcular_preco_virtual(preco_calculado, gordura, arredondamento)
            it["preco_promocional"] = round(preco_calculado, 2)
            it["preco"] = preco_virtual
            it["preco_virtual"] = preco_virtual
    return item_mod


def _buscar_gordura_canal(canal_key: str, gordura_por_canal: dict) -> dict:
    padrao = {"tipo": "percentual", "valor": 20.0}
    if canal_key in gordura_por_canal: return gordura_por_canal[canal_key]
    def _norm(s): return s.lower().replace(" ", "_").replace("-", "_")
    canal_norm = _norm(canal_key)
    for nome, gordura in gordura_por_canal.items():
        if _norm(nome) == canal_norm: return gordura
    aliases = {
        "mercado_livre_classico": ["ml_classico","mercadolivre_classico","mercado livre classico"],
        "mercado_livre_premium": ["ml_premium","mercadolivre_premium","mercado livre premium"],
        "shopee": ["shopee"], "amazon": ["amazon"], "shein": ["shein"], "shopify": ["shopify","shopfy"],
    }
    for chave_cfg, alias_list in aliases.items():
        if canal_norm in [_norm(a) for a in alias_list]:
            if chave_cfg in gordura_por_canal: return gordura_por_canal[chave_cfg]
    return padrao


def _verificar_assinatura_bling(body_bytes: bytes, header: str) -> bool:
    """Valida HMAC-SHA256 do webhook Bling (X-Bling-Signature-256: sha256=<hex>)."""
    secret = os.getenv("BLING_WEBHOOK_SECRET", "")
    if not secret or not header:
        return True  # Sem segredo configurado: aceita tudo
    try:
        expected = "sha256=" + _hmac.new(secret.encode(), body_bytes, hashlib.sha256).hexdigest()
        return _hmac.compare_digest(expected, header)
    except Exception:
        return False

class IntegracaoPayload(BaseModel):
    criterio: str = "sku"
    valor_busca: str = ""
    embalagem: Optional[float] = None  # None = usa embalagem_padrao da config
    imposto: float = 4
    quantidade: int = 1
    objetivo: str = "lucro_liquido"
    tipo_alvo: str = "percentual"
    valor_alvo: float = 30
    peso_override: float = 0
    score_config: Optional[dict] = None
    modo_aprovacao: str = "manual"
    modo_preco_virtual: str = "percentual_acima"
    acrescimo_percentual: float = 20
    acrescimo_nominal: float = 0
    preco_manual: float = 0
    arredondamento: str = "90"
    preco_compra_anterior_bling: float = 0

@app.post("/shopify/kits/limpar-barcodes")
def shopify_kits_limpar_barcodes(dry_run: bool = False):
    """
    Remove o barcode (GTIN/EAN) de todos os produtos Kit/Combo/Duo no Shopify.

    O GMC cruza GTINs com bases externas e pode reprovar kits se o barcode
    de um componente coincidir com produto de categoria restrita (ex: tabaco).
    Kits não possuem um único EAN — é seguro remover o barcode deles.

    Query params:
        dry_run=true   → apenas lista o que seria alterado, sem modificar nada
        dry_run=false  → executa a limpeza (padrão)
    """
    try:
        from limpar_barcodes_kits import limpar_barcodes_kits
        result = limpar_barcodes_kits(dry_run=dry_run)
        return result
    except Exception as e:
        logger.exception("Erro em /shopify/kits/limpar-barcodes")
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────────────────────────────────────
# Webhook Shopify — limpeza de barcode em tempo real para kits
# ─────────────────────────────────────────────────────────────────────────────

import base64 as _base64

def _verificar_assinatura_shopify(body_bytes: bytes, hmac_header: str) -> bool:
    """Valida X-Shopify-Hmac-Sha256 (base64 de HMAC-SHA256 com SHOPIFY_WEBHOOK_SECRET)."""
    secret = os.getenv("SHOPIFY_WEBHOOK_SECRET", "")
    if not secret:
        return True  # Sem segredo configurado: aceita tudo (dev/testing)
    if not hmac_header:
        return False
    try:
        computed = _base64.b64encode(
            _hmac.new(secret.encode("utf-8"), body_bytes, hashlib.sha256).digest()
        ).decode()
        return _hmac.compare_digest(computed, hmac_header)
    except Exception:
        return False


@app.post("/webhooks/shopify/produto")
async def webhook_shopify_produto(request: Request, background_tasks: BackgroundTasks):
    """
    Recebe events products/create e products/update do Shopify.

    Se o produto for kit/combo (título contém keywords), limpa automaticamente
    os barcodes das variantes em background — evitando violações de política
    de tabaco no Google Merchant Center.

    Registro do webhook via: POST /shopify/webhooks/setup
    """
    raw = await request.body()

    # ── Verificação de assinatura ──────────────────────────────────────
    hmac_header = request.headers.get("X-Shopify-Hmac-Sha256", "")
    if not _verificar_assinatura_shopify(raw, hmac_header):
        logger.warning("[WEBHOOK-SHOPIFY] Assinatura inválida — rejeitando")
        raise HTTPException(status_code=401, detail="assinatura inválida")

    try:
        product = json.loads(raw)
    except Exception:
        return {"ok": True}  # Payload inválido: responde 200 para Shopify não reenviar

    topic = request.headers.get("X-Shopify-Topic", "")
    title = product.get("title", "")

    # ── Verifica se é kit ──────────────────────────────────────────────
    try:
        from limpar_barcodes_kits import is_kit, clear_variant_barcode
    except ImportError:
        logger.error("[WEBHOOK-SHOPIFY] Não foi possível importar limpar_barcodes_kits")
        return {"ok": True}

    if not is_kit(product):
        return {"ok": True, "kit": False}

    # ── Filtra variantes que têm barcode ──────────────────────────────
    variants_com_barcode = [
        v for v in product.get("variants", [])
        if v.get("barcode")
    ]

    if not variants_com_barcode:
        return {"ok": True, "kit": True, "barcodes": 0}

    prod_id = product.get("id")
    logger.info(
        "[WEBHOOK-SHOPIFY] Kit detectado: '%s' (id=%s, topic=%s) — %d variante(s) com barcode",
        title[:60], prod_id, topic, len(variants_com_barcode),
    )

    # ── Limpa barcodes em background (Shopify exige resposta <5 s) ────
    def _limpar_bg():
        import time
        cleared = erros = 0
        for v in variants_com_barcode:
            time.sleep(0.15)
            if clear_variant_barcode(v["id"]):
                cleared += 1
                logger.info(
                    "[WEBHOOK-SHOPIFY] Barcode removido: variant %s ('%s') produto '%s'",
                    v["id"], v.get("barcode", ""), title[:50],
                )
            else:
                erros += 1
                logger.warning("[WEBHOOK-SHOPIFY] Falha ao limpar variant %s", v["id"])
        logger.info(
            "[WEBHOOK-SHOPIFY] Concluído produto '%s': %d limpos, %d erros",
            title[:50], cleared, erros,
        )

    background_tasks.add_task(_limpar_bg)

    return {
        "ok": True,
        "kit": True,
        "produto": title[:80],
        "barcodes_a_limpar": len(variants_com_barcode),
    }


@app.post("/shopify/webhooks/setup")
def shopify_webhooks_setup():
    """
    Registra (ou confirma) os webhooks necessários no Shopify:
      - products/create
      - products/update

    Requer SHOPIFY_WEBHOOK_SECRET no .env / Railway para validação HMAC.
    O webhook aponta para {APP_URL}/webhooks/shopify/produto.

    Execute uma vez após o deploy: POST /shopify/webhooks/setup
    """
    import requests as _req

    cfg = json.loads((DATA_DIR / "shopify_config.json").read_text(encoding="utf-8"))
    token = cfg["access_token"]
    base  = "https://pknw4n-eg.myshopify.com/admin/api/2024-01"
    headers = {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}

    app_url = os.getenv("APP_URL", "").rstrip("/")
    if not app_url:
        raise HTTPException(
            status_code=400,
            detail="Variável de ambiente APP_URL não definida. "
                   "Defina APP_URL=https://seu-dominio.railway.app no Railway.",
        )

    endpoint = f"{app_url}/webhooks/shopify/produto"
    topics   = ["products/create", "products/update"]
    results  = []

    # Lista webhooks existentes para evitar duplicatas
    existing = _req.get(f"{base}/webhooks.json", headers=headers, timeout=15).json()
    existing_addresses = {w["address"]: w for w in existing.get("webhooks", [])}

    for topic in topics:
        if endpoint in existing_addresses:
            wh = existing_addresses[endpoint]
            results.append({"topic": topic, "status": "já_existe", "id": wh["id"], "address": endpoint})
            logger.info("[WEBHOOK-SETUP] Já existe: %s → %s", topic, endpoint)
            continue

        payload = {"webhook": {"topic": topic, "address": endpoint, "format": "json"}}
        r = _req.post(f"{base}/webhooks.json", headers=headers, json=payload, timeout=15)
        if r.status_code == 201:
            wh = r.json().get("webhook", {})
            results.append({"topic": topic, "status": "criado", "id": wh.get("id"), "address": endpoint})
            logger.info("[WEBHOOK-SETUP] Criado: %s → %s (id=%s)", topic, endpoint, wh.get("id"))
        else:
            results.append({"topic": topic, "status": "erro", "code": r.status_code, "detail": r.text[:200]})
            logger.error("[WEBHOOK-SETUP] Erro ao criar %s: %s %s", topic, r.status_code, r.text[:200])

    return {"ok": True, "endpoint": endpoint, "webhooks": results}


@app.get("/shopify/webhooks/listar")
def shopify_webhooks_listar():
    """Lista todos os webhooks registrados no Shopify."""
    import requests as _req
    cfg = json.loads((DATA_DIR / "shopify_config.json").read_text(encoding="utf-8"))
    token = cfg["access_token"]
    r = _req.get(
        "https://pknw4n-eg.myshopify.com/admin/api/2024-01/webhooks.json",
        headers={"X-Shopify-Access-Token": token},
        timeout=15,
    )
    r.raise_for_status()
    return r.json()


@app.post("/scheduler/ciclo", tags=["Scheduler"])
def scheduler_ciclo_manual():
    """
    Dispara manualmente um ciclo completo do scheduler.
    Útil após entrada de nota fiscal ou reajuste de custos no Bling.
    """
    try:
        from scheduler import _ciclo_atualizacao
        import threading

        resultado = {}
        erro = {}

        def _rodar():
            try:
                resultado.update(_ciclo_atualizacao())
            except Exception as e:
                erro["msg"] = str(e)

        t = threading.Thread(target=_rodar, daemon=True)
        t.start()
        t.join(timeout=60)  # aguarda até 60s

        if erro:
            return {"ok": False, "erro": erro["msg"]}
        return {"ok": True, "resultado": resultado}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/scheduler/status", tags=["Scheduler"])
def scheduler_status():
    """Retorna estado atual do scheduler e fila de aprovação."""
    try:
        from scheduler import _scheduler_thread
        ativo = _scheduler_thread is not None and _scheduler_thread.is_alive()
    except Exception:
        ativo = False

    from database import stats_fila
    stats = stats_fila()
    return {
        "scheduler_ativo": ativo,
        "intervalo_segundos": int(os.getenv("SCHEDULER_INTERVALO", "300")),
        "fila": stats,
    }


@app.get("/webhooks/bling")
async def webhook_bling_verify():
    """Responde ao ping de verificacao do Bling (GET) para manter o webhook ativo."""
    return {"ok": True, "status": "ativo"}


@app.post("/webhooks/bling")
async def webhook_bling(request: Request, background_tasks: BackgroundTasks):
    raw = await request.body()
    sig = request.headers.get("X-Bling-Signature-256", "")
    if not _verificar_assinatura_bling(raw, sig):
        logger.warning("Webhook Bling: assinatura inválida — ignorando")
        return {"ok": False, "erro": "assinatura inválida"}
    try:
        body = json.loads(raw)
    except Exception:
        body = {}
    evento = body.get("evento") or body.get("event") or "desconhecido"
    logger.info("Webhook Bling recebido: evento=%s", evento)
    _append_jsonl(LOG_PATH, {"evento":"webhook_bling","tipo":evento,"quando":datetime.utcnow().isoformat(),"payload":body})
    # Propaga atualização de estoque em tempo real para Shopify e ML
    background_tasks.add_task(sync_estoque_bling, body)
    return {"ok": True, "recebido": True}

@app.post("/webhooks/ml")
async def webhook_ml(request: Request, background_tasks: BackgroundTasks):
    """
    Recebe notificacoes do Mercado Livre.
    O ML envia POST: {"resource": "...", "user_id": 123, "topic": "price_suggestion"}
    Responde 200 imediatamente; processamento pesado vai para background task.
    """
    try:
        body = await request.json()
    except Exception:
        body = {}
    topic = body.get("topic") or body.get("type") or "desconhecido"
    resource = body.get("resource", "")
    user_id = body.get("user_id", "")
    logger.info("Webhook ML recebido: topic=%s resource=%s user_id=%s", topic, resource, user_id)
    _append_jsonl(LOG_PATH, {
        "evento": "webhook_ml",
        "topic": topic,
        "resource": resource,
        "user_id": user_id,
        "quando": datetime.utcnow().isoformat(),
        "payload": body,
    })

    # Processa sugestoes de preco em background (nao bloqueia resposta ao ML)
    if topic == "price_suggestion" and resource:
        def _processar_bg():
            try:
                from services.ml_price_suggestions import processar_price_suggestion
                regras = carregar_regras(apenas_ativas=True)
                client = BlingClient() if BlingClient else None
                processar_price_suggestion(
                    resource=resource,
                    user_id=str(user_id),
                    bling_client=client,
                    regras=regras,
                )
            except Exception as exc:
                import traceback
                logger.warning("Erro bg price_suggestion %s: %s\n%s", resource, exc, traceback.format_exc())
        background_tasks.add_task(_processar_bg)

    # Captura candidatos a promocao — ML envia preco sugerido diretamente!
    elif topic == "public_candidates" and resource:
        def _processar_candidate_bg():
            try:
                from services.ml_price_suggestions import _buscar_candidate_promo, _load_sugestoes, _save_sugestoes
                data = _buscar_candidate_promo(resource)
                if data:
                    # Salva o payload bruto para analise posterior
                    _append_jsonl(LOG_PATH, {
                        "evento": "public_candidate_ml",
                        "resource": resource,
                        "user_id": str(user_id),
                        "quando": datetime.utcnow().isoformat(),
                        "payload": data,
                    })
                    logger.info("public_candidate salvo: %s", json.dumps(data, ensure_ascii=False)[:400])
            except Exception as exc:
                logger.warning("Erro bg public_candidate %s: %s", resource, exc)
        background_tasks.add_task(_processar_candidate_bg)

    return {"ok": True}

FALLBACK_HTML = "<!DOCTYPE html><html lang='pt-BR'><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width, initial-scale=1.0'><title>Shinsei Pricing</title></head><body><h1>Shinsei Pricing</h1></body></html>"

@app.get("/marketing/ml/sugestoes")
def marketing_ml_sugestoes():
    """Retorna sugestoes de preco recebidas via webhook price_suggestion."""
    from services.ml_price_suggestions import carregar_sugestoes, resumo_sugestoes
    return {
        "ok": True,
        "resumo": resumo_sugestoes(),
        "sugestoes": carregar_sugestoes(),
    }


@app.delete("/marketing/ml/sugestoes")
def marketing_ml_sugestoes_limpar():
    """Limpa todas as sugestoes salvas."""
    from services.ml_price_suggestions import limpar_sugestoes
    limpar_sugestoes()
    return {"ok": True}

@app.post("/marketing/ml/sugestoes/limpar")
def marketing_ml_sugestoes_limpar_post():
    """Alternativa POST para limpar sugestoes (compativel com ngrok/proxies)."""
    from services.ml_price_suggestions import limpar_sugestoes
    limpar_sugestoes()
    return {"ok": True}

@app.post("/marketing/ml/sugestoes/reprocessar")
async def marketing_ml_sugestoes_reprocessar(background_tasks: BackgroundTasks):
    """Reprocessa todas as sugestoes salvas para calcular margem atual."""
    from services.ml_price_suggestions import carregar_sugestoes

    sugestoes = carregar_sugestoes()
    if not sugestoes:
        return {"ok": True, "mensagem": "Nenhuma sugestao para reprocessar."}

    def _reprocessar_bg():
        try:
            from services.ml_price_suggestions import (
                processar_price_suggestion, _load_sugestoes, _save_sugestoes
            )
            regras = carregar_regras(apenas_ativas=True)
            client = BlingClient() if BlingClient else None
            lista = _load_sugestoes()
            if not lista:
                return
            # Zera cooldown (usa timestamp antigo mas válido)
            for s in lista:
                s["recebido_em"] = "2020-01-01T00:00:00"
            _save_sugestoes(lista)
            # Reprocessa cada item
            for s in lista:
                try:
                    processar_price_suggestion(
                        resource=s.get("resource", f"/marketplace/benchmarks/items/{s['item_id']}/details"),
                        user_id=s.get("user_id", ""),
                        bling_client=client,
                        regras=regras,
                    )
                    import time as _t; _t.sleep(0.2)
                except Exception as e:
                    logger.warning("Reprocessar %s: %s", s.get("item_id"), e)
            logger.info("Reprocessamento concluido: %d sugestoes", len(lista))
        except Exception as exc:
            logger.warning("Erro reprocessar sugestoes: %s", exc)

    background_tasks.add_task(_reprocessar_bg)
    return {"ok": True, "mensagem": f"Reprocessando {len(sugestoes)} sugestão(ões) em background..."}



@app.get("/conferencia-estoque", response_class=HTMLResponse)
def conferencia_estoque_page():
    html_file = PAGES_DIR / "conferencia_estoque.html"
    if html_file.exists():
        return HTMLResponse(html_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="conferencia_estoque.html não encontrado.")

@app.get("/auditoria/bling-sem-shopify", response_class=HTMLResponse)
def bling_sem_shopify_page():
    html_file = PAGES_DIR / "bling_sem_shopify.html"
    if html_file.exists():
        return HTMLResponse(html_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="bling_sem_shopify.html não encontrado.")

@app.get("/amazon", response_class=HTMLResponse)
def amazon_page():
    html_file = PAGES_DIR / "amazon.html"
    if html_file.exists():
        return HTMLResponse(html_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="amazon.html não encontrado.")

@app.get("/amazon/kits", response_class=HTMLResponse)
def amazon_kits_page():
    html_file = PAGES_DIR / "amazon_kits.html"
    if html_file.exists():
        return HTMLResponse(html_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="amazon_kits.html não encontrado.")

@app.get("/shopify", response_class=HTMLResponse)
def shopify_page():
    html_file = PAGES_DIR / "shopify.html"
    if html_file.exists():
        return HTMLResponse(html_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="shopify.html não encontrado.")

@app.get("/estoque/fila")
def estoque_fila_lista(status: str = ""):
    from estoque_conferencia import carregar_fila_estoque, stats_fila_estoque
    itens = carregar_fila_estoque()
    if status:
        itens = [i for i in itens if i.get("status") == status]
    return {"itens": itens, "stats": stats_fila_estoque()}

@app.post("/estoque/conferir")
def estoque_conferir():
    from estoque_conferencia import conferir_estoques
    if not BlingClient:
        raise HTTPException(status_code=500, detail="Bling não disponível.")
    try:
        client = BlingClient()
        ml_svc = None
        try:
            from services.mercado_livre import MercadoLivreService
            ml_svc = MercadoLivreService(BASE_DIR)
        except Exception:
            pass
        return conferir_estoques(client, ml_svc)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/estoque/corrigir/{item_id}")
def estoque_corrigir(item_id: str):
    from estoque_conferencia import corrigir_item_estoque, stats_fila_estoque
    ml_svc = None
    try:
        from services.mercado_livre import MercadoLivreService
        ml_svc = MercadoLivreService(BASE_DIR)
    except Exception:
        pass
    resultado = corrigir_item_estoque(item_id, ml_svc)
    if not resultado.get("ok"):
        raise HTTPException(status_code=400, detail=resultado.get("erro","Falha ao corrigir."))
    return {"ok": True, "resultado": resultado, "stats": stats_fila_estoque()}

@app.post("/estoque/ignorar/{item_id}")
def estoque_ignorar(item_id: str):
    from estoque_conferencia import ignorar_item_estoque, stats_fila_estoque
    resultado = ignorar_item_estoque(item_id)
    return {"ok": resultado.get("ok"), "stats": stats_fila_estoque()}

@app.post("/estoque/limpar-resolvidos")
def estoque_limpar_resolvidos():
    from estoque_conferencia import carregar_fila_estoque, salvar_fila_estoque, stats_fila_estoque
    itens = carregar_fila_estoque()
    itens = [i for i in itens if i.get("status") == "pendente"]
    salvar_fila_estoque(itens)
    return {"ok": True, "stats": stats_fila_estoque()}


@app.post("/shopify-flow/pricing-suggestion")
async def shopify_flow_pricing(request: Request):
    """
    Endpoint para Shopify Flow.
    Recebe payload com SKU do produto, busca no Bling,
    calcula preços pelo motor e retorna sugestão por canal.
    Não aplica preços automaticamente.
    """
    try:
        body = await request.json()
    except Exception:
        body = {}

    # Extrai SKU do payload Shopify (vários formatos possíveis)
    sku = (
        body.get("sku")
        or body.get("variant_sku")
        or body.get("product", {}).get("variants", [{}])[0].get("sku", "") if isinstance(body.get("product"), dict) else ""
        or body.get("codigo")
        or ""
    )
    sku = str(sku).strip()

    if not sku:
        return {
            "ok": False,
            "erro": "SKU não encontrado no payload.",
            "payload_recebido": body,
        }

    if not BlingClient or not montar_precificacao_bling:
        return {"ok": False, "erro": "Motor de precificação indisponível."}

    # Carrega configuração comercial
    cfg = carregar_integracao_cfg()
    objetivo = cfg.get("objetivo", "lucro_liquido")
    tipo_alvo = cfg.get("tipo_alvo", "percentual")
    valor_alvo = float(cfg.get("valor_alvo", 30))
    arredondamento = str(cfg.get("arredondamento", "90"))

    try:
        regras = carregar_regras(apenas_ativas=True)

        _sie_cfg_sf = _load_json(_MOD_DIR / "sie_score_config.json", {})
        resultado = montar_precificacao_bling(
            regras=regras,
            criterio="sku",
            valor_busca=sku,
            embalagem=float(body.get("embalagem", 1)),
            imposto=float(body.get("imposto", 4)),
            quantidade=int(body.get("quantidade", 1)),
            objetivo=objetivo,
            tipo_alvo=tipo_alvo,
            valor_alvo=valor_alvo,
            peso_override=float(body.get("peso_override", 0)),
            arredondamento=arredondamento,
            regra_estoque=cfg.get("regra_estoque"),
            score_config=_sie_cfg_sf if _sie_cfg_sf.get("ajuste_ativo") else None,
        )
    except Exception as exc:
        logger.warning("Shopify Flow: erro no motor para SKU=%s: %s", sku, exc)
        return {"ok": False, "sku": sku, "erro": str(exc)}

    if resultado.get("erro"):
        return {"ok": False, "sku": sku, "erro": resultado["erro"]}

    # Monta sugestão por canal com gordura
    gordura_por_canal = cfg.get("gordura_por_canal", {})
    itens = (resultado.get("integracao") or {}).get("itens") or resultado.get("itens") or []
    sugestoes = []

    for item in itens:
        if not isinstance(item, dict):
            continue
        canal = item.get("canal", "")
        preco_calculado = float(
            item.get("preco_promocional") or item.get("preco_final") or item.get("preco") or 0
        )
        if preco_calculado <= 0:
            continue

        gordura = gordura_por_canal.get(canal, {"tipo": "percentual", "valor": 20})
        preco_virtual = calcular_preco_virtual(preco_calculado, gordura, arredondamento)

        sugestoes.append({
            "canal": canal,
            "preco_calculado": round(preco_calculado, 2),
            "preco_virtual": preco_virtual,
            "lucro_liquido": float(item.get("lucro_liquido") or item.get("lucro") or 0),
            "margem": float(item.get("margem") or item.get("margem_liquida_percentual") or 0),
        })

    # Sugestão do canal Shopify especificamente
    shopify_sugestao = next((s for s in sugestoes if "shopify" in s["canal"].lower()), None)

    produto = resultado.get("produto_bling") or {}
    logger.info("Shopify Flow: SKU=%s canais=%d", sku, len(sugestoes))
    _append_jsonl(LOG_PATH, {
        "evento": "shopify_flow_suggestion",
        "sku": sku,
        "quando": datetime.utcnow().isoformat(),
        "canais": len(sugestoes),
    })

    return {
        "ok": True,
        "sku": sku,
        "produto": {
            "nome": produto.get("nome") or produto.get("descricao") or "",
            "codigo": produto.get("codigo") or sku,
        },
        "sugestao_shopify": shopify_sugestao,
        "todos_canais": sugestoes,
        "objetivo_usado": objetivo,
        "valor_alvo_usado": valor_alvo,
    }


@app.get("/auditoria-automatica", response_class=HTMLResponse)
def auditoria_automatica_page():
    html_file = PAGES_DIR / "auditoria_automatica.html"
    if html_file.exists():
        return HTMLResponse(html_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="auditoria_automatica.html não encontrado.")

@app.get("/auditoria/fila")
def auditoria_fila_lista(status: str = "", tipo: str = ""):
    from auditoria_automatica import carregar_fila_auditoria, stats_fila_auditoria
    itens = carregar_fila_auditoria()
    if status:
        itens = [i for i in itens if i.get("status") == status]
    if tipo:
        itens = [i for i in itens if i.get("tipo") == tipo]
    return {"itens": itens, "stats": stats_fila_auditoria()}

@app.post("/auditoria/rodar")
def auditoria_rodar():
    from auditoria_automatica import rodar_auditoria
    if not BlingClient:
        raise HTTPException(status_code=500, detail="Bling não disponível.")
    try:
        client = BlingClient()
        ml_svc = None
        try:
            from services.mercado_livre import MercadoLivreService
            ml_svc = MercadoLivreService(BASE_DIR)
        except Exception:
            pass
        return rodar_auditoria(client, ml_svc)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/auditoria/corrigir/{item_id}")
def auditoria_corrigir(item_id: str):
    from auditoria_automatica import carregar_fila_auditoria, corrigir_estoque, corrigir_preco, stats_fila_auditoria
    fila = carregar_fila_auditoria()
    item = next((i for i in fila if i.get("id") == item_id), None)
    if not item:
        raise HTTPException(status_code=404, detail="Item não encontrado.")
    ml_svc = None
    try:
        from services.mercado_livre import MercadoLivreService
        ml_svc = MercadoLivreService(BASE_DIR)
    except Exception:
        pass
    if item.get("tipo") == "estoque":
        resultado = corrigir_estoque(item_id, ml_svc)
    else:
        resultado = corrigir_preco(item_id, ml_svc)
    if not resultado.get("ok"):
        raise HTTPException(status_code=400, detail=resultado.get("erro", "Falha ao corrigir."))
    return {"ok": True, "resultado": resultado, "stats": stats_fila_auditoria()}

@app.post("/auditoria/ignorar/{item_id}")
def auditoria_ignorar(item_id: str):
    from auditoria_automatica import ignorar_item, stats_fila_auditoria
    resultado = ignorar_item(item_id)
    return {"ok": resultado.get("ok"), "stats": stats_fila_auditoria()}

@app.post("/auditoria/limpar-resolvidos")
def auditoria_limpar():
    from auditoria_automatica import limpar_resolvidos, stats_fila_auditoria
    removidos = limpar_resolvidos()
    return {"ok": True, "removidos": removidos, "stats": stats_fila_auditoria()}


@app.get("/auditoria/ml-estoque")
def auditoria_ml_estoque_lista(status: str = "", tipo: str = ""):
    from ml_estoque_conferencia import carregar_fila_estoque_ml, stats_fila_estoque_ml
    itens = carregar_fila_estoque_ml()
    if status:
        itens = [i for i in itens if i.get("status") == status]
    if tipo:
        itens = [i for i in itens if i.get("tipo") == tipo]
    return {"itens": itens, "stats": stats_fila_estoque_ml()}

@app.post("/auditoria/ml-estoque/conferir")
def auditoria_ml_estoque_conferir():
    if not BlingClient:
        raise HTTPException(status_code=500, detail="Bling não disponível.")
    if _conf_ml["rodando"]:
        return {"ok": True, "em_andamento": True, "message": "Conferência já em andamento.", "estado": _conf_ml}
    _conf_ml.update({
        "rodando": True, "concluido": False, "erro": None,
        "pagina": 0, "verificados": 0, "divergencias": 0,
        "sem_sku": 0, "erros": 0, "resultado": None,
        "iniciado_em": datetime.utcnow().isoformat(), "concluido_em": None,
    })
    threading.Thread(target=_rodar_conf_ml_bg, daemon=True).start()
    return {"ok": True, "em_andamento": True, "message": "Conferência iniciada em background."}

@app.get("/auditoria/ml-estoque/conferir/status")
def auditoria_ml_estoque_conferir_status():
    return {"ok": True, **_conf_ml}

@app.post("/auditoria/ml-estoque/corrigir/{item_id}")
def auditoria_ml_estoque_corrigir(item_id: str):
    from ml_estoque_conferencia import corrigir_estoque_ml
    resultado = corrigir_estoque_ml(item_id)
    if not resultado.get("ok"):
        raise HTTPException(status_code=400, detail=resultado.get("erro", "Falha."))
    return {"ok": True}

@app.post("/auditoria/ml-estoque/cadastrar-sku/{item_id}")
async def auditoria_ml_cadastrar_sku(item_id: str, request: Request):
    from ml_estoque_conferencia import cadastrar_sku_ml
    data = await request.json()
    sku = data.get("sku", "").strip()
    if not sku:
        raise HTTPException(status_code=400, detail="SKU obrigatório.")
    client = BlingClient() if BlingClient else None
    resultado = cadastrar_sku_ml(item_id, sku, client)
    if not resultado.get("ok"):
        raise HTTPException(status_code=400, detail=resultado.get("erro", "Falha."))
    return resultado

@app.post("/auditoria/ml-estoque/ignorar/{item_id}")
def auditoria_ml_estoque_ignorar(item_id: str):
    from ml_estoque_conferencia import ignorar_item_ml
    return ignorar_item_ml(item_id)

@app.post("/auditoria/ml-estoque/limpar-resolvidos")
def auditoria_ml_limpar_resolvidos():
    from ml_estoque_conferencia import carregar_fila_estoque_ml, salvar_fila_estoque_ml, stats_fila_estoque_ml
    itens = carregar_fila_estoque_ml()
    itens = [i for i in itens if i.get("status") == "pendente"]
    salvar_fila_estoque_ml(itens)
    return {"ok": True, "stats": stats_fila_estoque_ml()}

@app.post("/auditoria/ml-estoque/limpar-tudo")
def auditoria_ml_limpar_tudo():
    from ml_estoque_conferencia import salvar_fila_estoque_ml, stats_fila_estoque_ml
    salvar_fila_estoque_ml([])
    return {"ok": True, "stats": stats_fila_estoque_ml()}


@app.get("/integracoes", response_class=HTMLResponse)
def integracoes_page():
    html_file = PAGES_DIR / "integracoes.html"
    if html_file.exists():
        return HTMLResponse(html_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404)



@app.get("/ml/status")
def ml_status_endpoint():
    import json
    tp = DATA_DIR / "ml_tokens.json"
    if not tp.exists(): return {"connected": False}
    tokens = json.loads(tp.read_text(encoding="utf-8"))
    at = tokens.get("access_token", "")
    return {"connected": bool(at) and at != ".", "seller_id": tokens.get("user_id"), "expires_at": tokens.get("expires_at")}

@app.get("/auditoria/shopify")
def auditoria_shopify_lista(status: str = "", tipo: str = ""):
    from shopify_conferencia import carregar_fila_shopify, stats_fila_shopify
    itens = carregar_fila_shopify()
    if status:
        itens = [i for i in itens if i.get("status") == status]
    if tipo:
        itens = [i for i in itens if i.get("tipo") == tipo]
    return {"itens": itens, "stats": stats_fila_shopify()}

@app.post("/auditoria/shopify/conferir")
def auditoria_shopify_conferir(tipo: str = ""):
    from shopify_conferencia import conferir_shopify
    if not BlingClient:
        raise HTTPException(status_code=500, detail="Bling nao disponivel.")
    global _scheduler_pausado
    _scheduler_pausado = True
    try:
        client = BlingClient()
        resultado = conferir_shopify(client, tipo=tipo)
        return resultado
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        _scheduler_pausado = False

@app.post("/auditoria/shopify/corrigir/{item_id}")
def auditoria_shopify_corrigir(item_id: str):
    from shopify_conferencia import corrigir_shopify
    resultado = corrigir_shopify(item_id)
    if not resultado.get("ok"):
        raise HTTPException(status_code=400, detail=resultado.get("erro", "Falha."))
    return {"ok": True}

@app.post("/auditoria/shopify/ignorar/{item_id}")
def auditoria_shopify_ignorar(item_id: str):
    from shopify_conferencia import ignorar_shopify
    return ignorar_shopify(item_id)

@app.post("/auditoria/shopify/limpar-resolvidos")
def auditoria_shopify_limpar():
    from shopify_conferencia import carregar_fila_shopify, salvar_fila_shopify, stats_fila_shopify
    itens = carregar_fila_shopify()
    itens = [i for i in itens if i.get("status") == "pendente"]
    salvar_fila_shopify(itens)
    return {"ok": True, "stats": stats_fila_shopify()}

@app.post("/auditoria/shopify/token")
async def auditoria_shopify_token(request: Request):
    from shopify_conferencia import salvar_shopify_token
    data = await request.json()
    token = data.get("access_token", "").strip()
    if not token:
        raise HTTPException(status_code=400, detail="Token invalido.")
    salvar_shopify_token(token)
    return {"ok": True}


def _shopify_redirect_uri(request: Request) -> str:
    """Constrói redirect_uri correto mesmo atrás de proxy HTTPS (Railway/ngrok)."""
    env_url = os.getenv("SHOPIFY_CALLBACK_URL", "")
    if env_url:
        return env_url
    scheme = request.headers.get("x-forwarded-proto", request.url.scheme)
    host = request.headers.get("x-forwarded-host", request.url.netloc)
    return f"{scheme}://{host}/shopify/callback"

@app.get("/shopify/auth")
def shopify_auth(request: Request):
    from shopify_oauth import gerar_url_auth
    from fastapi.responses import RedirectResponse
    redirect_uri = _shopify_redirect_uri(request)
    url = gerar_url_auth(redirect_uri)
    return RedirectResponse(url)

@app.get("/shopify/callback")
def shopify_callback(code: str = "", state: str = "", request: Request = None):
    from shopify_oauth import processar_callback
    from fastapi.responses import HTMLResponse
    redirect_uri = _shopify_redirect_uri(request)
    resultado = processar_callback(code, state, redirect_uri)
    if resultado.get("ok"):
        return HTMLResponse("<h2>✅ Shopify conectado! Token salvo com sucesso.</h2><p><a href='/integracoes'>Voltar para Integrações</a></p>")
    return HTMLResponse(f"<h2>❌ Erro: {resultado.get('erro')}</h2>")

@app.get("/shopify/status")
def shopify_status():
    import json
    cfg = DATA_DIR / "shopify_config.json"
    if not cfg.exists(): return {"connected": False}
    data = json.loads(cfg.read_text(encoding="utf-8"))
    token = data.get("access_token", "")
    return {"connected": bool(token) and token != ".", "scope": data.get("scope", ""), "salvo_em": data.get("salvo_em")}

@app.post("/shopify/install-gtag")
def shopify_install_gtag():
    """Instala o pixel de conversão do Google Ads na página de confirmação de pedido."""
    import json
    cfg = DATA_DIR / "shopify_config.json"
    if not cfg.exists():
        raise HTTPException(status_code=400, detail="Shopify não conectado.")
    data = json.loads(cfg.read_text(encoding="utf-8"))
    token = data.get("access_token", "")
    scope = data.get("scope", "")
    if not token:
        raise HTTPException(status_code=400, detail="Token Shopify ausente.")
    if "write_script_tags" not in scope:
        raise HTTPException(status_code=403, detail=f"Token sem write_script_tags. Escopos: {scope}")
    try:
        from shopify_oauth import _instalar_gtag_conversion
        info = _instalar_gtag_conversion(token)
        return {"ok": True, "mensagem": "Script tag de conversão instalada com sucesso.", **(info or {})}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/shopify/gtag-status")
def shopify_gtag_status():
    """Lista as script tags instaladas na loja Shopify para verificar a conversão."""
    import json, requests as req
    cfg = DATA_DIR / "shopify_config.json"
    if not cfg.exists():
        raise HTTPException(status_code=400, detail="Shopify não conectado.")
    data = json.loads(cfg.read_text(encoding="utf-8"))
    token = data.get("access_token", "")
    if not token:
        raise HTTPException(status_code=400, detail="Token Shopify ausente.")
    from shopify_oauth import SHOPIFY_STORE
    TEMA_ID = 185169445169
    base    = f"https://{SHOPIFY_STORE}.myshopify.com/admin/api/2024-01"
    headers = {"X-Shopify-Access-Token": token}
    # Script tags
    r = req.get(f"{base}/script_tags.json?limit=50", headers=headers, timeout=15)
    if r.status_code != 200:
        raise HTTPException(status_code=r.status_code, detail=r.text[:300])
    tags = r.json().get("script_tags", [])
    # Asset no tema
    asset_r = req.get(
        f"https://{SHOPIFY_STORE}.myshopify.com/admin/api/2024-01/themes/{TEMA_ID}/assets.json"
        "?asset[key]=assets/shinsei-ads-conversion.js",
        headers=headers, timeout=15
    )
    asset_info = asset_r.json().get("asset", {}) if asset_r.status_code == 200 else {"erro": asset_r.status_code}
    return {
        "script_tags": tags,
        "total": len(tags),
        "shinsei_tags": [t for t in tags if "shinsei" in t.get("src","").lower() or "AW-" in t.get("src","")],
        "asset": {
            "key": asset_info.get("key"),
            "public_url": asset_info.get("public_url"),
            "updated_at": asset_info.get("updated_at"),
            "size": asset_info.get("size"),
        }
    }


@app.post("/shopify/deploy-secao-oferta")
def shopify_deploy_secao_oferta():
    """
    Faz o deploy do arquivo shinsei_oferta_section.liquid para o tema Shopify ativo.
    Atualiza sections/shinsei-oferta-section.liquid via Admin API.
    """
    import json, requests as req
    cfg_path = BASE_DIR / "data" / "shopify_config.json"
    liquid_path = BASE_DIR / "shinsei_oferta_section.liquid"

    if not cfg_path.exists():
        raise HTTPException(status_code=400, detail="Shopify não conectado.")
    if not liquid_path.exists():
        raise HTTPException(status_code=404, detail="shinsei_oferta_section.liquid não encontrado.")

    data  = json.loads(cfg_path.read_text(encoding="utf-8"))
    token = data.get("access_token", "")
    if not token:
        raise HTTPException(status_code=400, detail="Token Shopify ausente.")

    from shopify_oauth import SHOPIFY_STORE
    TEMA_ID  = 185169445169
    base     = f"https://{SHOPIFY_STORE}.myshopify.com/admin/api/2024-01"
    headers  = {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}

    content  = liquid_path.read_text(encoding="utf-8")
    payload  = {"asset": {"key": "sections/shinsei-oferta-section.liquid", "value": content}}

    r = req.put(f"{base}/themes/{TEMA_ID}/assets.json", headers=headers, json=payload, timeout=30)
    if r.status_code not in (200, 201):
        raise HTTPException(status_code=r.status_code, detail=f"Shopify API: {r.text[:400]}")

    asset = r.json().get("asset", {})
    return {
        "ok": True,
        "key": asset.get("key"),
        "updated_at": asset.get("updated_at"),
        "size": asset.get("size"),
        "public_url": asset.get("public_url"),
        "msg": "sections/shinsei-oferta-section.liquid atualizado com sucesso no tema.",
    }


@app.get("/shopify/template-homepage")
def shopify_template_homepage():
    """Retorna o template index.json da homepage para inspecionar as seções."""
    import json, requests as req
    cfg_path = BASE_DIR / "data" / "shopify_config.json"
    if not cfg_path.exists():
        raise HTTPException(status_code=400, detail="Shopify não conectado.")
    data  = json.loads(cfg_path.read_text(encoding="utf-8"))
    token = data.get("access_token", "")
    if not token:
        raise HTTPException(status_code=400, detail="Token Shopify ausente.")
    from shopify_oauth import SHOPIFY_STORE
    TEMA_ID = 185169445169
    base = f"https://{SHOPIFY_STORE}.myshopify.com/admin/api/2024-01"
    headers = {"X-Shopify-Access-Token": token}
    r = req.get(f"{base}/themes/{TEMA_ID}/assets.json?asset[key]=templates/index.json", headers=headers, timeout=15)
    if r.status_code != 200:
        raise HTTPException(status_code=r.status_code, detail=r.text[:300])
    asset = r.json().get("asset", {})
    value = asset.get("value", "{}")
    return {"template": json.loads(value), "secoes": {k: v.get("type") for k, v in json.loads(value).get("sections", {}).items()}}


@app.post("/shopify/substituir-secao")
def shopify_substituir_secao(body: dict):
    """
    Substitui uma seção no template index.json da homepage.
    body: { "substituir": "id_ou_tipo_da_secao_antiga", "nova_secao": "shinsei-oferta-section" }
    """
    import json, requests as req
    cfg_path = BASE_DIR / "data" / "shopify_config.json"
    if not cfg_path.exists():
        raise HTTPException(status_code=400, detail="Shopify não conectado.")
    data  = json.loads(cfg_path.read_text(encoding="utf-8"))
    token = data.get("access_token", "")
    if not token:
        raise HTTPException(status_code=400, detail="Token Shopify ausente.")
    from shopify_oauth import SHOPIFY_STORE
    TEMA_ID  = 185169445169
    base     = f"https://{SHOPIFY_STORE}.myshopify.com/admin/api/2024-01"
    headers  = {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}

    # 1. Busca o template atual
    r = req.get(f"{base}/themes/{TEMA_ID}/assets.json?asset[key]=templates/index.json", headers=headers, timeout=15)
    if r.status_code != 200:
        raise HTTPException(status_code=r.status_code, detail=f"GET template: {r.text[:300]}")
    template = json.loads(r.json()["asset"]["value"])

    substituir = body.get("substituir", "")  # id ou tipo a substituir
    nova_secao = body.get("nova_secao", "shinsei-oferta-section")

    sections = template.get("sections", {})
    order    = template.get("order", [])

    # Encontra a seção alvo (por id ou por tipo)
    alvo_id = None
    for sid, sdata in sections.items():
        if substituir.lower() in sid.lower() or substituir.lower() in sdata.get("type", "").lower():
            alvo_id = sid
            break

    if not alvo_id:
        return {
            "ok": False,
            "msg": f"Seção '{substituir}' não encontrada.",
            "secoes_disponiveis": {k: v.get("type") for k, v in sections.items()},
        }

    # Substitui
    old_type = sections[alvo_id].get("type")
    sections[alvo_id] = {"type": nova_secao, "disabled": False, "settings": {}, "blocks": {}, "block_order": []}
    template["sections"] = sections
    template["order"]    = order

    # 2. PUT do template atualizado
    new_value = json.dumps(template, ensure_ascii=False)
    payload   = {"asset": {"key": "templates/index.json", "value": new_value}}
    rp = req.put(f"{base}/themes/{TEMA_ID}/assets.json", headers=headers, json=payload, timeout=30)
    if rp.status_code not in (200, 201):
        raise HTTPException(status_code=rp.status_code, detail=f"PUT template: {rp.text[:400]}")

    return {
        "ok": True,
        "substituiu": alvo_id,
        "tipo_antigo": old_type,
        "tipo_novo": nova_secao,
        "msg": f"Seção '{alvo_id}' ({old_type}) substituída por '{nova_secao}' com sucesso.",
    }


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

INTEGRACAO_CFG_PATH = DATA_DIR / "integracao_comercial.json"

DEFAULT_INTEGRACAO_CFG = {
    "objetivo": "lucro_liquido",
    "tipo_alvo": "percentual",
    "valor_alvo": 30.0,
    "arredondamento": "90",
    "modo_aprovacao": "manual",
    "fila_auto_ao_calcular": True,
    "gordura_por_canal": {
        "Mercado Livre Classico": {"tipo": "percentual", "valor": 20.0},
        "Mercado Livre Premium": {"tipo": "percentual", "valor": 20.0},
        "Shopee": {"tipo": "percentual", "valor": 20.0},
        "Amazon": {"tipo": "percentual", "valor": 20.0},
        "Shein": {"tipo": "percentual", "valor": 20.0},
        "Shopify": {"tipo": "percentual", "valor": 20.0},
    },
    "forcas_canais": {
        "Mercado Livre Classico": 0.8,
        "Mercado Livre Premium": 0.75,
        "Shopee": 0.6,
        "Amazon": 0.7,
        "Shein": 0.55,
        "Shopify": 0.65,
    },
    "peso_forca": 0.4,
    "peso_equilibrio": 0.4,
    "peso_lucro": 0.2,
    "regra_estoque": {"ativo": False, "limite": 2, "tipo": "percentual", "valor": 0},
    "modo_auto": False,
    "auto_margem_ok": 25.0,
    "auto_margem_fila": 15.0,
}


def carregar_integracao_cfg() -> dict:
    data = _load_json(INTEGRACAO_CFG_PATH, {})
    cfg = json.loads(json.dumps(DEFAULT_INTEGRACAO_CFG))
    if isinstance(data, dict):
        cfg.update(data)
        # merge profundo para sub-dicts
        if "gordura_por_canal" in data:
            cfg["gordura_por_canal"] = {
                **DEFAULT_INTEGRACAO_CFG["gordura_por_canal"],
                **data["gordura_por_canal"],
            }
        if "forcas_canais" in data:
            cfg["forcas_canais"] = {
                **DEFAULT_INTEGRACAO_CFG["forcas_canais"],
                **data["forcas_canais"],
            }
        if "regra_estoque" in data:
            cfg["regra_estoque"] = {
                **DEFAULT_INTEGRACAO_CFG["regra_estoque"],
                **data["regra_estoque"],
            }
    return cfg


def calcular_preco_virtual(preco_calculado: float, gordura: dict, arredondamento: str = "90") -> float:
    """Aplica a gordura sobre o preço calculado e arredonda."""
    tipo = gordura.get("tipo", "percentual")
    valor = float(gordura.get("valor", 20))

    if tipo == "percentual":
        virtual = preco_calculado * (1 + valor / 100)
    else:
        virtual = preco_calculado + valor

    return _arredondar_preco(virtual, arredondamento)


def _arredondar_preco(v: float, modo: str) -> float:
    if modo == "sem":
        return round(v, 2)
    sufixo = int(modo) / 100  # "90" â†’ 0.90
    base = int(v)
    proposto = base + sufixo
    if proposto >= v:
        return round(proposto, 2)
    return round(base + 1 + sufixo, 2)


# â”€â”€â”€ ROTA: página de integração comercial â”€â”€â”€
@app.get("/integracao-comercial", response_class=HTMLResponse)
def integracao_comercial_page():
    html_file = PAGES_DIR / "integracao_comercial.html"
    if html_file.exists():
        return HTMLResponse(html_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="integracao_comercial.html não encontrado.")


# â”€â”€â”€ ROTA: GET config â”€â”€â”€
@app.get("/config/integracao-comercial")
def get_integracao_config():
    return carregar_integracao_cfg()


# â”€â”€â”€ ROTA: POST config â”€â”€â”€
@app.post("/config/integracao-comercial")
async def set_integracao_config(request: Request):
    data = await request.json()
    if not isinstance(data, dict):
        raise HTTPException(status_code=400, detail="Payload inválido.")

    cfg_atual = carregar_integracao_cfg()

    # Campos simples
    for campo in ["objetivo", "tipo_alvo", "arredondamento", "modo_aprovacao"]:
        if campo in data:
            cfg_atual[campo] = str(data[campo])

    for campo in ["valor_alvo", "peso_forca", "peso_equilibrio", "peso_lucro", "auto_margem_ok", "auto_margem_fila"]:
        if campo in data:
            try:
                cfg_atual[campo] = float(data[campo])
            except (TypeError, ValueError):
                pass

    for campo in ["fila_auto_ao_calcular", "modo_auto", "ml_api_real", "amazon_api_real", "shopee_api_real"]:
        if campo in data:
            cfg_atual[campo] = bool(data[campo])
    for campo in ["embalagem_padrao", "imposto_padrao"]:
        if campo in data:
            try:
                cfg_atual[campo] = float(data[campo])
            except (TypeError, ValueError):
                pass

    # Sub-dicts
    if "gordura_por_canal" in data and isinstance(data["gordura_por_canal"], dict):
        cfg_atual["gordura_por_canal"] = {
            **cfg_atual.get("gordura_por_canal", {}),
            **data["gordura_por_canal"],
        }

    if "forcas_canais" in data and isinstance(data["forcas_canais"], dict):
        cfg_atual["forcas_canais"] = {
            **cfg_atual.get("forcas_canais", {}),
            **data["forcas_canais"],
        }

    if "regra_estoque" in data and isinstance(data["regra_estoque"], dict):
        cfg_atual["regra_estoque"] = {
            **cfg_atual.get("regra_estoque", {}),
            **data["regra_estoque"],
        }

    _save_json(INTEGRACAO_CFG_PATH, cfg_atual)
    logger.info("Configuração de integração comercial atualizada: objetivo=%s", cfg_atual.get("objetivo"))

    return {"ok": True, "config": cfg_atual}


# â”€â”€â”€ ROTA: calcular preço virtual para um canal â”€â”€â”€
@app.post("/config/calcular-preco-virtual")
async def calcular_preco_virtual_endpoint(request: Request):
    """Dado um preço calculado, retorna o preço virtual por canal com a gordura configurada."""
    data = await request.json()
    preco = float(data.get("preco", 0))
    if preco <= 0:
        raise HTTPException(status_code=400, detail="Preço inválido.")

    cfg = carregar_integracao_cfg()
    gordura_por_canal = cfg.get("gordura_por_canal", {})
    arredondamento = str(data.get("arredondamento") or cfg.get("arredondamento", "90"))

    resultado = {}
    for canal, gordura in gordura_por_canal.items():
        virtual = calcular_preco_virtual(preco, gordura, arredondamento)
        dif_nominal = round(virtual - preco, 2)
        dif_pct = round((dif_nominal / preco) * 100, 2) if preco > 0 else 0
        resultado[canal] = {
            "preco_calculado": round(preco, 2),
            "preco_virtual": virtual,
            "diferenca_nominal": dif_nominal,
            "diferenca_percentual": dif_pct,
            "gordura": gordura,
        }

    return {"ok": True, "canais": resultado}


@app.get("/auditoria/mp-status")
def auditoria_mp_status():
    import json as _json
    mp = DATA_DIR / "mp_token.json"
    data = _json.loads(mp.read_text(encoding="utf-8")) if mp.exists() else {}
    token = data.get("access_token", "")
    return {"configurado": bool(token) and token != ".", "salvo_em": data.get("salvo_em")}
if not FILA_PATH.exists(): _save_json(FILA_PATH, [])
if not CFG_PATH.exists(): _save_json(CFG_PATH, DEFAULT_CFG)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# MÓDULO REGRAS
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

MAPA_CANAIS_EXCEL = {
    "Classico": "Mercado Livre Classico",
    "Classic": "Mercado Livre Classico",
    "Premium": "Mercado Livre Premium",
    "Shopfy": "Shopify",
}

def _normalizar_canal_excel(canal: str) -> str:
    s = str(canal or "").strip()
    return MAPA_CANAIS_EXCEL.get(s, s)

def _para_float(v, default=0.0):
    if v is None or v == "": return default
    try: return float(v)
    except Exception:
        try: return float(str(v).strip().replace(".", "").replace(",", "."))
        except Exception: return default

@app.get("/regras", response_class=HTMLResponse)
def regras_page():
    html_file = PAGES_DIR / "regras.html"
    if html_file.exists():
        return HTMLResponse(html_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="pages/regras.html não encontrado.")

@app.get("/regras/listar")
def regras_listar():
    regras = carregar_regras()
    for i, r in enumerate(regras):
        if isinstance(r, dict): r["_idx"] = i
    return {"regras": regras, "total": len(regras)}

@app.post("/regras/adicionar")
def regras_adicionar(payload: dict = Body(...)):
    nova = {
        "canal": str(payload.get("canal", "")).strip(),
        "peso_min": _para_float(payload.get("peso_min"), 0),
        "peso_max": _para_float(payload.get("peso_max"), 999999),
        "preco_min": _para_float(payload.get("preco_min"), 0),
        "preco_max": _para_float(payload.get("preco_max"), 999999999),
        "taxa_fixa": _para_float(payload.get("taxa_fixa"), 0),
        "taxa_frete": _para_float(payload.get("taxa_frete"), 0),
        "comissao": _para_float(payload.get("comissao"), 0),
        "ativo": bool(payload.get("ativo", True)),
    }
    if not nova["canal"]:
        raise HTTPException(status_code=400, detail="Canal obrigatório.")
    novo_id = inserir_regra(nova)
    return {"ok": True, "id": novo_id, "total": len(carregar_regras())}

@app.post("/regras/editar/{idx}")
def regras_editar(idx: int, payload: dict = Body(...)):
    nova = {
        "canal": str(payload.get("canal", "")).strip(),
        "peso_min": _para_float(payload.get("peso_min"), 0),
        "peso_max": _para_float(payload.get("peso_max"), 999999),
        "preco_min": _para_float(payload.get("preco_min"), 0),
        "preco_max": _para_float(payload.get("preco_max"), 999999999),
        "taxa_fixa": _para_float(payload.get("taxa_fixa"), 0),
        "taxa_frete": _para_float(payload.get("taxa_frete"), 0),
        "comissao": _para_float(payload.get("comissao"), 0),
        "ativo": bool(payload.get("ativo", True)),
    }
    ok = atualizar_regra(idx, nova)
    if not ok:
        raise HTTPException(status_code=404, detail="Regra não encontrada.")
    return {"ok": True, "total": len(carregar_regras())}

@app.delete("/regras/excluir/{idx}")
def regras_excluir(idx: int):
    ok = excluir_regra(idx)
    if not ok:
        raise HTTPException(status_code=404, detail="Regra não encontrada.")
    return {"ok": True, "total": len(carregar_regras())}

@app.post("/regras/importar-excel")
async def regras_importar_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Apenas arquivos .xlsx são aceitos.")
    try:
        import io
        import openpyxl
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        nome_aba = next((n for n in wb.sheetnames if n.lower() in ("regras", "aba2")), wb.sheetnames[0])
        ws = wb[nome_aba]
        regras = []
        for row in ws.iter_rows(min_row=2, max_col=8, values_only=True):
            canal, peso_min, peso_max, preco_min, preco_max, taxa_fixa, taxa_frete, comissao = row
            canal = _normalizar_canal_excel(canal)
            if not canal: continue
            regras.append({
                "canal": canal,
                "peso_min": _para_float(peso_min, 0),
                "peso_max": _para_float(peso_max, 999999),
                "preco_min": _para_float(preco_min, 0),
                "preco_max": _para_float(preco_max, 999999999),
                "taxa_fixa": _para_float(taxa_fixa, 0),
                "taxa_frete": _para_float(taxa_frete, 0),
                "comissao": _para_float(comissao, 0),
                "ativo": True,
            })
        if not regras:
            raise HTTPException(status_code=400, detail="Nenhuma regra encontrada na planilha. Verifique se a aba se chama 'Regras' ou 'Aba2'.")
        substituir_todas_regras(regras)
        return {"ok": True, "total": len(regras), "aba": nome_aba}
    except HTTPException: raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar Excel: {e}")

@app.get("/regras/modelo/download")
def regras_modelo_download():
    modelo = BASE_DIR / "Simulador_modelo.xlsx"
    if not modelo.exists():
        raise HTTPException(status_code=404, detail="Arquivo modelo não encontrado.")
    return FileResponse(
        path=str(modelo),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="Shinsei_Regras_Modelo.xlsx",
    )

def _severidade_negativo(estoque: int) -> str:
    """Classifica a severidade do estoque negativo."""
    if estoque >= -5:   return "leve"
    if estoque >= -20:  return "moderado"
    if estoque >= -100: return "grave"
    return "critico"


@app.get("/auditoria/estoque-negativo")
def auditoria_estoque_negativo_lista(status: str = "", severidade: str = "", busca: str = "", limit: int = 200, offset: int = 0):
    import json as _j
    fila_path = DATA_DIR / "fila_estoque_negativo.json"
    itens = _j.loads(fila_path.read_text(encoding="utf-8")) if fila_path.exists() else []
    # Enriquece com severidade
    for i in itens:
        i["severidade"] = _severidade_negativo(int(i.get("estoque", 0)))
    # Filtra apenas negativos
    itens = [i for i in itens if int(i.get("estoque", 0)) < 0]
    if status:
        itens = [i for i in itens if i.get("status") == status]
    if severidade:
        itens = [i for i in itens if i.get("severidade") == severidade]
    if busca:
        b = busca.lower()
        itens = [i for i in itens if b in str(i.get("sku","")).lower() or b in str(i.get("nome","")).lower()]
    # Ordena do mais grave ao mais leve
    itens.sort(key=lambda x: int(x.get("estoque", 0)))
    total = len(itens)
    pendentes = sum(1 for i in itens if i.get("status") == "pendente")
    # Stats de severidade
    sv = {"leve":0,"moderado":0,"grave":0,"critico":0}
    for i in itens:
        sv[i.get("severidade","leve")] = sv.get(i.get("severidade","leve"), 0) + 1
    return {
        "itens": itens[offset:offset+limit],
        "total": total,
        "stats": {
            "pendente": pendentes,
            "total": total,
            "severidade": sv,
        }
    }


@app.get("/auditoria/estoque-negativo/resumo")
def auditoria_estoque_negativo_resumo():
    """Resumo rápido por severidade — sem retornar todos os itens."""
    import json as _j
    fila_path = DATA_DIR / "fila_estoque_negativo.json"
    itens = _j.loads(fila_path.read_text(encoding="utf-8")) if fila_path.exists() else []
    negativos = [i for i in itens if int(i.get("estoque", 0)) < 0]
    pendentes = [i for i in negativos if i.get("status") == "pendente"]
    sv = {"leve":0,"moderado":0,"grave":0,"critico":0}
    piores = []
    for i in pendentes:
        s = _severidade_negativo(int(i.get("estoque", 0)))
        sv[s] = sv.get(s, 0) + 1
        if s in ("grave","critico"):
            piores.append({"sku": i.get("sku"), "nome": str(i.get("nome",""))[:50], "estoque": i.get("estoque"), "severidade": s})
    piores.sort(key=lambda x: int(x.get("estoque", 0)))
    return {
        "total_negativos": len(negativos),
        "pendentes": len(pendentes),
        "severidade": sv,
        "piores": piores[:20],
    }


@app.post("/auditoria/estoque-negativo/ignorar-leves")
def auditoria_negativo_ignorar_leves():
    """Ignora em batch todos os itens com estoque entre -1 e -5 (provavelmente transientes)."""
    import json as _j
    fila_path = DATA_DIR / "fila_estoque_negativo.json"
    itens = _j.loads(fila_path.read_text(encoding="utf-8")) if fila_path.exists() else []
    count = 0
    for i in itens:
        est = int(i.get("estoque", 0))
        if est < 0 and est >= -5 and i.get("status") == "pendente":
            i["status"] = "ignorado"
            count += 1
    DATA_DIR.mkdir(exist_ok=True)
    fila_path.write_text(_j.dumps(itens, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "ignorados": count}


@app.post("/auditoria/estoque-negativo/ignorar-severidade/{sev}")
def auditoria_negativo_ignorar_por_severidade(sev: str):
    """Ignora em batch todos os itens de uma severidade específica."""
    if sev not in ("leve","moderado","grave","critico"):
        raise HTTPException(status_code=400, detail="Severidade inválida")
    import json as _j
    fila_path = DATA_DIR / "fila_estoque_negativo.json"
    itens = _j.loads(fila_path.read_text(encoding="utf-8")) if fila_path.exists() else []
    count = 0
    for i in itens:
        est = int(i.get("estoque", 0))
        if _severidade_negativo(est) == sev and i.get("status") == "pendente":
            i["status"] = "ignorado"
            count += 1
    DATA_DIR.mkdir(exist_ok=True)
    fila_path.write_text(_j.dumps(itens, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "ignorados": count, "severidade": sev}

@app.post("/auditoria/estoque-negativo/ignorar/{item_id}")
def auditoria_estoque_negativo_ignorar(item_id: str):
    import json as _j
    fila_path = DATA_DIR / "fila_estoque_negativo.json"
    itens = _j.loads(fila_path.read_text(encoding="utf-8")) if fila_path.exists() else []
    for i in itens:
        if i.get("id") == item_id:
            i["status"] = "ignorado"
    DATA_DIR.mkdir(exist_ok=True)
    fila_path.write_text(_j.dumps(itens, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True}

@app.post("/auditoria/estoque-negativo/limpar")
def auditoria_estoque_negativo_limpar():
    import json as _j
    fila_path = DATA_DIR / "fila_estoque_negativo.json"
    itens = _j.loads(fila_path.read_text(encoding="utf-8")) if fila_path.exists() else []
    itens = [i for i in itens if i.get("status") == "pendente"]
    DATA_DIR.mkdir(exist_ok=True)
    fila_path.write_text(_j.dumps(itens, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "pendentes": len(itens)}

@app.post("/auditoria/shopify/limpar-tudo")
def auditoria_shopify_limpar_tudo():
    from shopify_conferencia import salvar_fila_shopify
    salvar_fila_shopify([])
    return {"ok": True}

@app.post("/auditoria/shopify/bling-sem-shopify")
def auditoria_bling_sem_shopify_conferir():
    """Varre o Bling e encontra produtos que NÃO existem na Shopify."""
    from shopify_conferencia import conferir_bling_sem_shopify
    if not BlingClient:
        raise HTTPException(status_code=500, detail="Bling não disponível.")
    try:
        bling = BlingClient()
        resultado = conferir_bling_sem_shopify(bling)
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/auditoria/shopify/bling-sem-shopify")
def auditoria_bling_sem_shopify_lista():
    """Retorna o último resultado da conferência Bling→Shopify."""
    from shopify_conferencia import carregar_bling_sem_shopify
    dados = carregar_bling_sem_shopify()
    if not dados:
        return {"ok": False, "erro": "Nenhuma conferência executada ainda. Use POST para iniciar."}
    return dados

@app.post("/auditoria/estoque-negativo/limpar-tudo")
def auditoria_negativo_limpar_tudo():
    DATA_DIR.mkdir(exist_ok=True)
    (DATA_DIR / "fila_estoque_negativo.json").write_text("[]", encoding="utf-8")
    return {"ok": True}

@app.get("/auditoria/amazon")
def auditoria_amazon_lista(status: str = "", tipo: str = ""):
    from amazon_conferencia import carregar_fila, stats_fila
    itens = [i for i in carregar_fila() if i.get("status") in ("pendente","incompleto")]
    if status:
        itens = [i for i in itens if i.get("status") == status]
    if tipo:
        itens = [i for i in itens if i.get("tipo") == tipo]
    return {"itens": itens, "stats": stats_fila()}

@app.post("/auditoria/amazon/conferir")
def auditoria_amazon_conferir(tipo: str = ""):
    from amazon_conferencia import conferir_amazon
    from amazon_client import AmazonClient
    if not BlingClient:
        raise HTTPException(status_code=500, detail="Bling não disponível.")
    try:
        bling = BlingClient()
        amazon = AmazonClient()
        resultado = conferir_amazon(bling_client=bling, tipo=tipo)
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/auditoria/amazon/corrigir/{item_id}")
def auditoria_amazon_corrigir(item_id: str):
    from amazon_conferencia import carregar_fila, salvar_fila
    from datetime import datetime, timezone
    fila = carregar_fila()
    item = next((i for i in fila if i.get("id") == item_id), None)
    if not item:
        raise HTTPException(status_code=404, detail="Item não encontrado na fila.")
    if item.get("status") != "pendente":
        raise HTTPException(status_code=400, detail=f"Item já está com status '{item['status']}'.")
    try:
        from services.amazon import AmazonService
        service = AmazonService()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Erro ao conectar à Amazon: {exc}")
    tipo = item.get("tipo")
    sku = item.get("sku")
    if tipo == "estoque":
        res = service.atualizar_estoque(sku, item["estoque_bling"])
    elif tipo == "preco":
        res = service.atualizar_com_retry(sku, item["preco_bling"])
    else:
        raise HTTPException(status_code=400, detail=f"Tipo desconhecido: {tipo}")
    if not res.get("success"):
        raise HTTPException(status_code=400, detail=f"Erro ao corrigir na Amazon: {res.get('error')}")
    item["status"] = "corrigido"
    item["corrigido_em"] = datetime.now(timezone.utc).isoformat()
    salvar_fila(fila)
    return {"ok": True, "item_id": item_id, "tipo": tipo, "sku": sku}

@app.post("/auditoria/amazon/ignorar/{item_id}")
def auditoria_amazon_ignorar(item_id: str):
    from amazon_conferencia import carregar_fila, salvar_fila
    itens = [i for i in carregar_fila() if i.get("status") in ("pendente","incompleto")]
    for i in itens:
        if i.get("id") == item_id:
            i["status"] = "ignorado"
    salvar_fila(itens)
    return {"ok": True}

@app.post("/auditoria/amazon/limpar-resolvidos")
def auditoria_amazon_limpar_resolvidos():
    from amazon_conferencia import carregar_fila, salvar_fila, stats_fila
    itens = [i for i in carregar_fila() if i.get("status") == "pendente"]
    salvar_fila(itens)
    return {"ok": True, "stats": stats_fila()}

@app.post("/auditoria/amazon/limpar-tudo")
def auditoria_amazon_limpar_tudo():
    DATA_DIR.mkdir(exist_ok=True)
    (DATA_DIR / "fila_amazon.json").write_text("[]", encoding="utf-8")
    return {"ok": True}

@app.get("/amazon/listings")
def amazon_listings_estoque():
    """
    Retorna todos os listings MFN da Amazon com SKU e quantidade disponível.
    Usa Listings Items API v2021-08-01 com fulfillmentAvailability.
    Público (sem API key) para uso em scripts de conferência.
    """
    try:
        from amazon_client import AmazonClient
        c = AmazonClient()
    except Exception as e:
        return {"ok": False, "erro": str(e), "itens": []}

    itens = []
    page_token = None
    paginas = 0
    MAX_PAGINAS = 200  # segurança

    while paginas < MAX_PAGINAS:
        try:
            resp = c.get_listings(page_token=page_token)
        except Exception as e:
            return {"ok": False, "erro": str(e), "itens": itens}

        erros_api = resp.get("errors") or []
        if erros_api:
            return {"ok": False, "erro": str(erros_api[:2]), "itens": itens}

        for item in (resp.get("items") or []):
            sku = item.get("sku", "").strip()
            if not sku:
                continue

            # Quantidade: fulfillmentAvailability lista canais (AMAZON_NA = MFN)
            qty = 0
            status_listing = ""
            for fa in (item.get("fulfillmentAvailability") or []):
                qty += int(fa.get("quantity") or 0)

            # summaries → status e ASIN
            asin = ""
            nome = ""
            for s in (item.get("summaries") or []):
                if not asin:
                    asin = s.get("asin", "")
                if not nome:
                    nome = s.get("itemName", "")[:60]
                status_listing = s.get("status", "")

            itens.append({
                "sku": sku,
                "asin": asin,
                "nome": nome,
                "qty": qty,
                "status": status_listing,
            })

        paginas += 1
        page_token = (resp.get("pagination") or {}).get("nextToken")
        if not page_token:
            break

    return {"ok": True, "total": len(itens), "itens": itens}


@app.get("/amazon/status")
def amazon_status():
    try:
        from amazon_client import AmazonClient
        c = AmazonClient()
        token = c._get_access_token()

        # Testa a Listings API para verificar se retorna anúncios
        listings_ok = False
        listings_count = 0
        listings_erro = None
        try:
            resp = c.get_listings()
            api_errors = resp.get("errors") or []
            items = resp.get("items") or []
            listings_count = len(items)
            if api_errors:
                listings_erro = str(api_errors[:2])
                listings_ok = False
            else:
                listings_ok = True  # API respondeu, mesmo que vazia
        except Exception as le:
            listings_erro = str(le)

        return {
            "ok": True,
            "configurado": True,
            "conectado": bool(token),
            "seller_id": c.config.get("seller_id", ""),
            "marketplace_id": c.config.get("marketplace_id", ""),
            "listings_api": {
                "ok": listings_ok,
                "itens_primeira_pagina": listings_count,
                "erro": listings_erro,
                "aviso": (
                    "A API retornou 0 itens na 1ª página — a conferência tratará Amazon como desconectada. "
                    "Verifique as permissões do app SP-API (precisa de 'Manage inventory' ou 'Listings')."
                ) if (listings_ok and listings_count == 0) else None,
            },
        }
    except Exception as e:
        return {"ok": False, "configurado": False, "conectado": False, "erro": str(e)}


# ── Amazon SP-API OAuth (self-authorization) ──────────────────────────────────

@app.get("/amazon/auth")
def amazon_auth(request: Request):
    """
    Gera URL de autorização para o SP-API (self-authorization Draft).
    Defina AMAZON_APP_ID com o Application ID do SP-API Developer Portal.
    """
    import json as _json, secrets as _sec
    app_id = os.getenv("AMAZON_APP_ID", "")
    if not app_id:
        # Verifica se já está conectado via LWA (refresh_token configurado)
        amazon_conectado = bool(os.getenv("AMAZON_REFRESH_TOKEN", ""))
        if amazon_conectado:
            return {
                "ok": True,
                "status": "ja_conectado",
                "mensagem": "Amazon ja esta conectada via LWA (refresh_token configurado). "
                            "Este endpoint so e necessario para gerar um novo refresh_token. "
                            "Acesse /amazon/status para confirmar a conexao.",
            }
        return {"ok": False, "erro": "Defina AMAZON_APP_ID no Cloud Run com o ID do app SP-API"}
    state = _sec.token_hex(16)
    (DATA_DIR / "amazon_oauth_state.json").write_text(
        _json.dumps({"state": state}), encoding="utf-8"
    )
    redirect_uri = os.getenv("AMAZON_CALLBACK_URL",
                             f"{request.headers.get('x-forwarded-proto', request.url.scheme)}"
                             f"://{request.headers.get('x-forwarded-host', request.url.netloc)}/amazon/callback")
    url = (
        f"https://sellercentral.amazon.com.br/apps/authorize/consent"
        f"?application_id={app_id}"
        f"&state={state}"
        f"&version=beta"
        f"&redirect_uri={redirect_uri}"
    )
    return {"ok": True, "url": url, "instrucao": "Abra esta URL no navegador logado no Seller Central"}


@app.get("/amazon/callback")
def amazon_callback(
    spapi_oauth_code: str = "",
    selling_partner_id: str = "",
    state: str = "",
    mws_auth_token: str = "",
):
    """
    Callback do SP-API OAuth. Troca o código por refresh_token e salva em
    data/amazon_tokens.json. Copie o refresh_token para AMAZON_REFRESH_TOKEN no Railway.
    """
    import json as _json, requests as _req
    from datetime import datetime, timezone
    try:
        saved_state_path = DATA_DIR / "amazon_oauth_state.json"
        if saved_state_path.exists():
            saved = _json.loads(saved_state_path.read_text(encoding="utf-8"))
            if saved.get("state") and saved.get("state") != state:
                return {"ok": False, "erro": "State inválido"}

        client_id     = os.getenv("AMAZON_CLIENT_ID", "")
        client_secret = os.getenv("AMAZON_CLIENT_SECRET", "")
        redirect_uri  = os.getenv("AMAZON_CALLBACK_URL",
                                   f"{request.headers.get('x-forwarded-proto', request.url.scheme)}"
                                   f"://{request.headers.get('x-forwarded-host', request.url.netloc)}/amazon/callback")

        resp = _req.post(
            "https://api.amazon.com/auth/o2/token",
            data={
                "grant_type":    "authorization_code",
                "code":          spapi_oauth_code,
                "redirect_uri":  redirect_uri,
                "client_id":     client_id,
                "client_secret": client_secret,
            },
            timeout=15,
        )
        if resp.status_code != 200:
            return {"ok": False, "status": resp.status_code, "erro": resp.text[:300]}

        data = resp.json()
        tokens = {
            "access_token":       data.get("access_token"),
            "refresh_token":      data.get("refresh_token"),
            "selling_partner_id": selling_partner_id,
            "salvo_em":           datetime.now(timezone.utc).isoformat(),
        }
        (DATA_DIR / "amazon_tokens.json").write_text(
            _json.dumps(tokens, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        refresh_token = data.get("refresh_token", "")
        return {
            "ok": True,
            "refresh_token": refresh_token,
            "instrucao": f"Copie este refresh_token e defina AMAZON_REFRESH_TOKEN={refresh_token} no Railway",
        }
    except Exception as e:
        return {"ok": False, "erro": str(e)}


# ── Fila unificada de preços (Amazon + Shopify) ───────────────────────────────

def _normalizar_item_preco(item: dict, canal: str) -> dict:
    """Normaliza item de fila de preço para o formato esperado pelo frontend."""
    preco_mkt = item.get("preco_amazon") or item.get("preco_shopify") or 0
    return {
        "id": item.get("id"),
        "sku": item.get("sku"),
        "nome": item.get("nome") or item.get("titulo") or "",
        "canal": canal,
        "preco_shinsei": float(item.get("preco_bling") or 0),
        "preco_marketplace_promocional": float(preco_mkt),
        "preco_virtual_shinsei": float(item.get("preco_bling") or 0),
        "preco_marketplace": float(preco_mkt),
        "diferenca": float(item.get("diferenca") or 0),
        "status": item.get("status", "pendente"),
        "detectado_em": item.get("detectado_em") or item.get("criado_em") or "",
    }


@app.get("/auditoria/precos")
def auditoria_precos_lista(status: str = ""):
    import json as _json
    from pathlib import Path as _P
    itens = []
    for path, canal in [("data/fila_amazon.json", "Amazon"), ("data/fila_shopify.json", "Shopify")]:
        try:
            raw = _json.loads(_P(path).read_text(encoding="utf-8")) if _P(path).exists() else []
            for i in (raw if isinstance(raw, list) else []):
                if i.get("tipo") == "preco":
                    if not status or i.get("status") == status:
                        itens.append(_normalizar_item_preco(i, canal))
        except Exception:
            pass
    return {"itens": itens, "total": len(itens), "pendentes": sum(1 for i in itens if i["status"] == "pendente")}


@app.post("/auditoria/conferir-precos")
def auditoria_conferir_precos():
    if not BlingClient:
        raise HTTPException(status_code=500, detail="bling_client.py nao encontrado.")
    bling = BlingClient()
    total_verificados = 0
    total_divergencias = 0
    erros = []
    try:
        from amazon_conferencia import conferir_amazon
        res = conferir_amazon(bling_client=bling, tipo="preco")
        if res.get("ok"):
            total_verificados += res.get("verificados", 0)
            total_divergencias += res.get("divergencias_preco", 0)
        else:
            erros.append("Amazon: " + str(res.get("erro", "falha")))
    except Exception as e:
        erros.append(f"Amazon: {e}")
    try:
        from shopify_conferencia import conferir_shopify
        res = conferir_shopify(bling_client=bling, tipo="preco")
        if res.get("ok"):
            total_verificados += res.get("verificados", 0)
            total_divergencias += res.get("divergencias_preco", 0)
        else:
            erros.append("Shopify: " + str(res.get("erro", "falha")))
    except Exception as e:
        erros.append(f"Shopify: {e}")
    return {
        "ok": True,
        "verificados": total_verificados,
        "novas_divergencias": total_divergencias,
        "erros": erros,
    }


@app.post("/auditoria/corrigir-preco/{item_id}")
def auditoria_corrigir_preco(item_id: str):
    import json as _json
    from pathlib import Path as _P
    from datetime import datetime, timezone

    # Tenta Amazon
    if item_id.startswith("amz_"):
        from amazon_conferencia import carregar_fila, salvar_fila
        fila = carregar_fila()
        item = next((i for i in fila if i.get("id") == item_id), None)
        if not item:
            raise HTTPException(status_code=404, detail="Item nao encontrado.")
        if item.get("status") != "pendente":
            raise HTTPException(status_code=400, detail=f"Item ja esta '{item['status']}'.")
        try:
            from services.amazon import AmazonService
            res = AmazonService().atualizar_com_retry(item["sku"], item["preco_bling"])
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
        if not res.get("success"):
            raise HTTPException(status_code=400, detail=res.get("error", "Falha ao corrigir preco na Amazon."))
        item["status"] = "corrigido"
        item["corrigido_em"] = datetime.now(timezone.utc).isoformat()
        salvar_fila(fila)
        return {"ok": True}

    # Tenta Shopify
    if item_id.startswith("shp_"):
        from shopify_conferencia import corrigir_shopify
        res = corrigir_shopify(item_id)
        if not res.get("ok"):
            raise HTTPException(status_code=400, detail=res.get("erro", "Falha ao corrigir preco na Shopify."))
        return {"ok": True}

    raise HTTPException(status_code=404, detail="Item nao encontrado em nenhuma fila.")


@app.post("/auditoria/ignorar-preco/{item_id}")
def auditoria_ignorar_preco(item_id: str):
    from datetime import datetime, timezone

    if item_id.startswith("amz_"):
        from amazon_conferencia import carregar_fila, salvar_fila
        fila = carregar_fila()
        item = next((i for i in fila if i.get("id") == item_id), None)
        if not item:
            raise HTTPException(status_code=404, detail="Item nao encontrado.")
        item["status"] = "ignorado"
        item["ignorado_em"] = datetime.now(timezone.utc).isoformat()
        salvar_fila(fila)
        return {"ok": True}

    if item_id.startswith("shp_"):
        from shopify_conferencia import ignorar_shopify
        res = ignorar_shopify(item_id)
        return {"ok": res.get("ok", False)}

    raise HTTPException(status_code=404, detail="Item nao encontrado em nenhuma fila.")


@app.post("/auditoria/mp-token")
async def auditoria_salvar_mp_token(request: Request):
    import json as _json
    from datetime import datetime
    body = await request.json()
    token = (body.get("access_token") or "").strip()
    if not token:
        raise HTTPException(status_code=400, detail="Token vazio.")
    path = DATA_DIR / "mp_token.json"
    DATA_DIR.mkdir(exist_ok=True)
    path.write_text(_json.dumps({"access_token": token, "salvo_em": datetime.utcnow().isoformat()}, indent=2), encoding="utf-8")
    return {"ok": True}





# ─── MARKETING ML ──────────────────────────────────────────────────────────────

@app.get("/marketing", response_class=HTMLResponse)
def marketing_page():
    html_file = PAGES_DIR / "marketing.html"
    if html_file.exists():
        return HTMLResponse(html_file.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="pages/marketing.html nao encontrado.")


@app.post("/marketing/ml/analisar")
async def marketing_ml_analisar(request: Request):
    body = await request.json()
    desconto_pct = float(body.get("desconto_pct", 10))
    margem_alvo = float(body.get("margem_alvo", 20))
    max_itens = int(body.get("max_itens", 500))
    imposto_padrao = float(body.get("imposto_padrao", 12.0))

    import json as _json
    tokens_path = DATA_DIR / "ml_tokens.json"
    if not tokens_path.exists():
        raise HTTPException(status_code=400, detail="Token ML nao configurado. Faca login em /ml/login.")
    tokens = _json.loads(tokens_path.read_text(encoding="utf-8"))
    seller_id = str(tokens.get("user_id", ""))
    if not seller_id:
        raise HTTPException(status_code=400, detail="seller_id nao encontrado no token ML.")

    if not BlingClient:
        raise HTTPException(status_code=500, detail="Bling nao disponivel.")

    from services.marketing_ml import analisar_campanhas_ml
    try:
        client = BlingClient()
        resultado = analisar_campanhas_ml(
            seller_id=seller_id,
            bling_client=client,
            desconto_pct=desconto_pct,
            margem_alvo=margem_alvo,
            imposto_padrao=imposto_padrao,
            max_itens=max_itens,
        )
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/marketing/ml/participar")
async def marketing_ml_participar(request: Request):
    """Atualiza o preco dos itens selecionados no ML para o preco da campanha."""
    body = await request.json()
    itens = body.get("itens", [])  # [{item_id, preco_campanha}, ...]
    if not itens:
        raise HTTPException(status_code=400, detail="Nenhum item selecionado.")

    from services.marketing_ml import atualizar_preco_ml
    resultados = []
    for item in itens:
        res = atualizar_preco_ml(item["item_id"], float(item["preco_campanha"]))
        resultados.append(res)
        import time as _time
        _time.sleep(0.3)

    ok = sum(1 for r in resultados if r.get("ok"))
    erros = len(resultados) - ok
    return {"ok": True, "atualizados": ok, "erros": erros, "detalhes": resultados}


# ─── MARKETING AMAZON ──────────────────────────────────────────────────────────

@app.get("/marketing/amazon/cache")
def marketing_amazon_cache():
    """Retorna resultado cacheado da última análise Amazon (sem refazer a análise)."""
    from services.amazon_marketing import carregar_cache_amazon
    cache = carregar_cache_amazon()
    if not cache:
        return {"ok": False, "erro": "Sem análise em cache"}
    return cache


@app.post("/marketing/amazon/analisar")
async def marketing_amazon_analisar(request: Request):
    body = await request.json()
    margem_alvo = float(body.get("margem_alvo", 20))
    imposto_padrao = float(body.get("imposto_padrao", 12.0))
    max_itens = int(body.get("max_itens", 500))

    if not BlingClient:
        raise HTTPException(status_code=500, detail="Bling nao disponivel.")

    from services.amazon_marketing import analisar_buy_box_amazon
    try:
        client = BlingClient()
        regras = carregar_regras(apenas_ativas=True)
        resultado = analisar_buy_box_amazon(
            bling_client=client,
            regras=regras,
            margem_alvo=margem_alvo,
            imposto_padrao=imposto_padrao,
            max_itens=max_itens,
        )
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/marketing/amazon/participar")
async def marketing_amazon_participar(request: Request):
    """Atualiza preço dos itens selecionados na Amazon para cobrir a Buy Box."""
    body = await request.json()
    itens = body.get("itens", [])
    if not itens:
        raise HTTPException(status_code=400, detail="Nenhum item selecionado.")

    from services.amazon_marketing import atualizar_preco_amazon
    import time as _time
    resultados = []
    for item in itens:
        res = atualizar_preco_amazon(item["sku"], float(item["preco_buy_box"]))
        resultados.append(res)
        _time.sleep(0.3)

    ok = sum(1 for r in resultados if r.get("ok"))
    erros = len(resultados) - ok
    return {"ok": True, "atualizados": ok, "erros": erros, "detalhes": resultados}


# ── SEO Health ────────────────────────────────────────────────────────────────

SEO_CACHE_PATH = DATA_DIR / "seo_health_cache.json"


@app.get("/seo-health", response_class=HTMLResponse)
def seo_health_page():
    path = PAGES_DIR / "seo_health.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="seo_health.html não encontrado.")
    return HTMLResponse(path.read_text(encoding="utf-8"))


@app.get("/seo-health/dados")
def seo_health_dados():
    cache = _load_json(SEO_CACHE_PATH, None)
    if not cache:
        return {"ok": False, "erro": "Nenhuma análise em cache. Clique em Analisar agora."}
    return {"ok": True, **cache}


@app.post("/seo-health/analisar")
async def seo_health_analisar():
    """Audita coleções, produtos e blog do Shopify e salva cache."""
    import asyncio, importlib.util
    seo_path = BASE_DIR / "shinsei_seo.py"
    if not seo_path.exists():
        raise HTTPException(status_code=500, detail="shinsei_seo.py não encontrado.")

    def _run():
        spec = importlib.util.spec_from_file_location("shinsei_seo", seo_path)
        seo = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(seo)
        return seo.health_score()

    try:
        loop = asyncio.get_event_loop()
        resultado = await loop.run_in_executor(None, _run)
        _save_json(SEO_CACHE_PATH, resultado)
        return {"ok": True, **resultado}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/seo-health/pagespeed")
async def seo_health_pagespeed():
    """Busca scores do PageSpeed Insights (mobile + desktop) e atualiza cache."""
    try:
        import importlib.util, os as _os
        seo_path = BASE_DIR / "shinsei_seo.py"
        if not seo_path.exists():
            raise HTTPException(status_code=500, detail="shinsei_seo.py não encontrado.")
        spec = importlib.util.spec_from_file_location("shinsei_seo", seo_path)
        seo = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(seo)

        import requests as _req
        store = getattr(seo, "STORE", "pknw4n-eg")
        api_key = _os.getenv("PAGESPEED_API_KEY", "")
        url_loja = f"https://www.shinseimarket.com.br"
        ps_url = "https://www.googleapis.com/pagespeedonline/v5/runPagespeed"
        scores = {}
        for strategy in ("mobile", "desktop"):
            params = [
                ("url", url_loja),
                ("strategy", strategy),
                ("category", "performance"),
                ("category", "seo"),
                ("category", "accessibility"),
                ("category", "best-practices"),
            ]
            if api_key:
                params.append(("key", api_key))
            r = _req.get(ps_url, params=params, timeout=90)
            if r.status_code == 200:
                d = r.json()
                cats = d.get("lighthouseResult", {}).get("categories", {})
                audits = d.get("lighthouseResult", {}).get("audits", {})
                oportunidades = [
                    v.get("title", "") for v in audits.values()
                    if isinstance(v, dict) and v.get("score") is not None
                    and float(v.get("score", 1)) < 0.9 and v.get("title")
                    and v.get("details", {}).get("type") not in ("table", "list", "criticalrequestchain")
                ]
                scores[strategy] = {
                    "performance": round((cats.get("performance", {}).get("score") or 0) * 100),
                    "seo": round((cats.get("seo", {}).get("score") or 0) * 100),
                    "accessibility": round((cats.get("accessibility", {}).get("score") or 0) * 100),
                    "best_practices": round((cats.get("best-practices", {}).get("score") or 0) * 100),
                    "lcp": audits.get("largest-contentful-paint", {}).get("displayValue", "—"),
                    "cls": audits.get("cumulative-layout-shift", {}).get("displayValue", "—"),
                    "tbt": audits.get("total-blocking-time", {}).get("displayValue", "—"),
                    "fcp": audits.get("first-contentful-paint", {}).get("displayValue", "—"),
                    "ttfb": audits.get("server-response-time", {}).get("displayValue", "—"),
                    "oportunidades": oportunidades[:10],
                }
            else:
                scores[strategy] = None

        cache = _load_json(SEO_CACHE_PATH, {})
        cache["pagespeed"] = scores
        _save_json(SEO_CACHE_PATH, cache)
        return {"ok": True, "pagespeed": scores}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/seo-health/merchant")
async def seo_health_merchant():
    """Audita cobertura do feed Google Merchant Center via Shopify API."""
    import requests as _req, asyncio, random as _rand

    cfg_path = DATA_DIR / "shopify_config.json"
    if not cfg_path.exists():
        raise HTTPException(status_code=500, detail="shopify_config.json não encontrado.")

    token = _load_json(cfg_path, {}).get("access_token", "")
    if not token:
        raise HTTPException(status_code=500, detail="Token Shopify não configurado.")

    hdrs = {"X-Shopify-Access-Token": token}
    store = "pknw4n-eg"
    base = f"https://{store}.myshopify.com/admin/api/2024-01"

    def _fetch_shopify():
        # Buscar 1ª página (250 produtos) para estimativa rápida de cobertura
        all_prods = []
        url = f"{base}/products.json?limit=250&fields=id,vendor,product_type"
        for _ in range(6):  # até 6 páginas = 1500 produtos
            try:
                r = _req.get(url, headers=hdrs, timeout=20)
                prods = r.json().get("products", [])
                if not prods:
                    break
                all_prods.extend(prods)
                links = r.headers.get("Link", "")
                url = None
                for part in links.split(","):
                    if 'rel="next"' in part:
                        url = part.strip().split(";")[0].strip("<> ")
                        break
                if not url:
                    break
            except Exception:
                break

        total = len(all_prods)
        if total == 0:
            return None

        sem_type = sum(1 for p in all_prods if not (p.get("product_type") or "").strip())
        vendor_errado = sum(1 for p in all_prods if p.get("vendor", "") in ("Shinsei Market", ""))

        # Checar google_product_category numa amostra de 40 produtos
        sample = _rand.sample(all_prods, min(40, total))
        com_gcat = 0
        for p in sample:
            try:
                rm = _req.get(f"{base}/products/{p['id']}/metafields.json?namespace=mm-google-shopping&limit=1",
                              headers=hdrs, timeout=10)
                if rm.json().get("metafields"):
                    com_gcat += 1
            except Exception:
                pass
        pct_gcat = round(com_gcat / len(sample) * 100) if sample else 0
        estimado_gcat = round(total * pct_gcat / 100)

        return {
            "analisado_em": datetime.utcnow().isoformat(),
            "total_produtos": total,
            "com_product_type": total - sem_type,
            "sem_product_type": sem_type,
            "pct_product_type": round((total - sem_type) / total * 100),
            "vendor_incorreto": vendor_errado,
            "pct_vendor_ok": round((total - vendor_errado) / total * 100),
            "estimado_google_cat": estimado_gcat,
            "pct_google_cat": pct_gcat,
            "pendencias_merchant": [
                {
                    "prioridade": "alto",
                    "titulo": "Forçar re-sincronização no Merchant Center",
                    "descricao": "product_type, vendor e google_product_category foram atualizados. O Merchant Center precisa re-processar o feed para refletir as mudanças.",
                    "impacto": "Produtos aprovados começam a aparecer no Google Shopping em até 24h",
                    "acao_url": "https://merchants.google.com/",
                    "acao_label": "Abrir Merchant Center",
                },
                {
                    "prioridade": "alto",
                    "titulo": "Verificar produtos reprovados no Merchant Center",
                    "descricao": "Após o re-sync, verifique em Produtos → Problemas se há itens suspensos por dado ausente, preço divergente ou imagem inválida.",
                    "impacto": "Cada produto reprovado é tráfego perdido de Shopping",
                    "acao_url": "https://merchants.google.com/",
                    "acao_label": "Ver Problemas",
                },
                {
                    "prioridade": "medio",
                    "titulo": "Ativar campanhas Performance Max",
                    "descricao": "Com product_type e google_product_category definidos, o Google Ads consegue segmentar campanhas PMax por categoria. Criar grupos de ativos por linha (Coloração, Tratamento, Maquiagem).",
                    "impacto": "Performance Max com dados estruturados pode aumentar impressões em 30-50%",
                    "acao_url": "https://ads.google.com/",
                    "acao_label": "Abrir Google Ads",
                },
                {
                    "prioridade": "medio",
                    "titulo": "Configurar e-mail de review pós-compra (Judge.me)",
                    "descricao": "Reviews com estrelas aparecem nos anúncios Shopping como Seller Ratings. Configurar envio automático 7 dias após entrega no painel Judge.me.",
                    "impacto": "Seller Ratings aumentam CTR do Shopping em média 17%",
                    "acao_url": "https://judge.me/",
                    "acao_label": "Abrir Judge.me",
                },
                {
                    "prioridade": "baixo",
                    "titulo": "Expandir para Google Merchant Center — Promoções",
                    "descricao": "Cadastrar promoções no Merchant Center (ex: frete grátis acima de R$99) para exibir badge 'Promoção' nos anúncios Shopping.",
                    "impacto": "Badge de promoção aumenta CTR em média 12%",
                    "acao_url": "https://merchants.google.com/",
                    "acao_label": "Ver Promoções",
                },
            ],
        }

    try:
        loop = asyncio.get_event_loop()
        resultado = await loop.run_in_executor(None, _fetch_shopify)
        if not resultado:
            raise HTTPException(status_code=500, detail="Não foi possível buscar dados do Shopify.")
        cache = _load_json(SEO_CACHE_PATH, {})
        cache["merchant"] = resultado
        _save_json(SEO_CACHE_PATH, cache)
        return {"ok": True, "merchant": resultado}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── SCBOT — Robô de Indexação Google ──────────────────────────────────────

@app.get("/scbot/status")
def scbot_status_endpoint():
    """Retorna status e histórico do SCBOT."""
    return {"ok": True, "scbot": scbot_status()}


@app.post("/scbot/executar")
async def scbot_executar_endpoint(urls_extras: list[str] = None):
    """Dispara um ciclo manual do SCBOT (ignora agendamento diário)."""
    import asyncio
    loop = asyncio.get_event_loop()
    resultado = await loop.run_in_executor(None, lambda: scbot_executar(urls_extras or [], force=True))
    # Salva no cache de SEO Health também
    cache = _load_json(SEO_CACHE_PATH, {})
    cache["scbot"] = scbot_status()
    _save_json(SEO_CACHE_PATH, cache)
    return {"ok": True, "resultado": resultado}


# ── Módulos Avançados — Endpoints de API ─────────────────────────────────────

_MOD_DIR = DATA_DIR

@app.get("/modulos/cost-engine")
def get_cost_engine():
    return _load_json(_MOD_DIR / "cost_engine.json", {})

@app.post("/modulos/cost-engine")
async def post_cost_engine(request: Request):
    body = await request.json()
    _save_json(_MOD_DIR / "cost_engine.json", body)
    return {"ok": True}

@app.get("/modulos/cost-allocation")
def get_cost_allocation():
    return _load_json(_MOD_DIR / "cost_allocation_module.json", {})

@app.post("/modulos/cost-allocation")
async def post_cost_allocation(request: Request):
    body = await request.json()
    _save_json(_MOD_DIR / "cost_allocation_module.json", body)
    return {"ok": True}

@app.get("/rateio/visoes")
def get_rateio_visoes():
    data = _load_json(_MOD_DIR / "cost_allocation_module.json", {})
    return {"visoes": data.get("visoes", [])}

@app.get("/config/automacao")
def get_automacao():
    oee = _load_json(_MOD_DIR / "oee_module.json", {})
    return {"oee": oee}

@app.post("/config/automacao")
async def post_automacao(request: Request):
    body = await request.json()
    oee_data = body.get("oee", body)
    _save_json(_MOD_DIR / "oee_module.json", oee_data)
    return {"ok": True}

@app.get("/modulos/sie")
def get_sie():
    return _load_json(_MOD_DIR / "sie_module.json", {})

@app.post("/modulos/sie")
async def post_sie(request: Request):
    body = await request.json()
    _save_json(_MOD_DIR / "sie_module.json", body)
    return {"ok": True}

@app.get("/modulos/regras-calculo")
def get_regras_calculo():
    return _load_json(_MOD_DIR / "regras_calculo.json", {})

@app.post("/modulos/regras-calculo")
async def post_regras_calculo(request: Request):
    body = await request.json()
    _save_json(_MOD_DIR / "regras_calculo.json", body)
    return {"ok": True}

@app.get("/config/regras-precificacao")
def get_regras_precificacao():
    return _load_json(_MOD_DIR / "regras_precificacao.json", {})

@app.post("/config/regras-precificacao")
async def post_regras_precificacao(request: Request):
    body = await request.json()
    _save_json(_MOD_DIR / "regras_precificacao.json", body)
    return {"ok": True}

@app.post("/config/regras-precificacao/ativar/{perfil_id}")
def ativar_perfil_precificacao(perfil_id: str):
    data = _load_json(_MOD_DIR / "regras_precificacao.json", {"perfis": []})
    for p in data.get("perfis", []):
        p["ativo"] = (str(p.get("id")) == perfil_id)
    _save_json(_MOD_DIR / "regras_precificacao.json", data)
    return {"ok": True, "ativo": perfil_id}


# ── SIE: Simulação + Motor Anti-Colapso + Health Dashboard ───────────────────

@app.post("/sie/simular")
async def sie_simular(request: Request):
    """
    Calcula o SIE completo + Motor Anti-Colapso para um produto.
    Body: { velocidade_venda, share_faturamento, margem_real,
            prazo_entrega_fornecedor, prazo_pagamento_fornecedor,
            tempo_para_vender, devolucao_defeito, devolucao_transporte,
            preco_final?, custo_base?, estoque?, nome_sku? }
    """
    body = await request.json()

    if _calculate_sie_fn is None:
        return {"erro": "product_intelligence.py não disponível"}

    sie_result = _calculate_sie_fn(body)

    anti_colapso = None
    if _motor_anti_colapso_fn:
        preco_final = float(body.get("preco_final", 0) or 0)
        custo_base  = float(body.get("custo_base", 0) or 0)
        if preco_final > 0 and custo_base > 0:
            anti_colapso = _motor_anti_colapso_fn(
                preco_final=preco_final,
                custo_base=custo_base,
                estoque=int(body.get("estoque", 0) or 0),
                sie_score=sie_result["sie"],
                icg=sie_result["icg"],
                velocidade_venda=float(body.get("velocidade_venda", 50) or 50),
            )

    # Persiste dados no sie_module.json para histórico
    sie_store = _load_json(_MOD_DIR / "sie_module.json", {})
    sku = str(body.get("nome_sku", "ultimo_simulado")).strip() or "ultimo_simulado"
    sie_store[sku] = {**body, "sie_result": sie_result, "anti_colapso": anti_colapso}
    _save_json(_MOD_DIR / "sie_module.json", sie_store)

    return {
        "sie": sie_result,
        "anti_colapso": anti_colapso,
        "sku": sku,
    }


@app.post("/sie/anti-colapso")
async def sie_anti_colapso(request: Request):
    """
    Roda apenas o Motor Anti-Colapso isolado.
    Body: { preco_final, custo_base, estoque, sie_score, icg,
            velocidade_venda, regra_colapso? }
    """
    if _motor_anti_colapso_fn is None:
        return {"erro": "motor_anti_colapso não disponível"}
    body = await request.json()
    result = _motor_anti_colapso_fn(
        preco_final=float(body.get("preco_final", 0)),
        custo_base=float(body.get("custo_base", 0)),
        estoque=int(body.get("estoque", 0)),
        sie_score=float(body.get("sie_score", 1.0)),
        icg=float(body.get("icg", 1.0)),
        velocidade_venda=float(body.get("velocidade_venda", 50)),
        regra_colapso=body.get("regra_colapso"),
    )
    return result


@app.get("/sie/health-dashboard")
def sie_health_dashboard():
    """
    Retorna snapshot de saúde de todos os SKUs salvos no sie_module.json.
    Combina: SIE score + classificação + ICG + sinais anti-colapso.
    """
    sie_store = _load_json(_MOD_DIR / "sie_module.json", {})
    skus = []

    for sku, dados in sie_store.items():
        if not isinstance(dados, dict):
            continue
        sie_result = dados.get("sie_result") or {}
        anti_colapso = dados.get("anti_colapso") or {}

        # Recalcula SIE se não tiver resultado salvo
        if not sie_result and _calculate_sie_fn:
            try:
                sie_result = _calculate_sie_fn(dados)
            except Exception:
                sie_result = {}

        skus.append({
            "sku": sku,
            "sie": sie_result.get("sie", 0),
            "classificacao": sie_result.get("classificacao", "?"),
            "icg": sie_result.get("icg", 0),
            "scores": sie_result.get("scores", {}),
            "nivel_risco": anti_colapso.get("nivel_risco", "ok"),
            "label_risco": anti_colapso.get("label", "Saudável"),
            "cor_risco": anti_colapso.get("cor", "#22c55e"),
            "sinais": anti_colapso.get("sinais", []),
            "margem_atual": anti_colapso.get("margem_atual"),
            "preco_original": anti_colapso.get("preco_original"),
            "preco_protegido": anti_colapso.get("preco_protegido"),
            "ajuste_pp": anti_colapso.get("ajuste_pp", 0),
            "entradas": sie_result.get("entradas", {}),
        })

    # Ordena: piores primeiro (maior risco = menor SIE)
    skus.sort(key=lambda x: x.get("sie", 1))

    resumo = {
        "total": len(skus),
        "estrela":  sum(1 for s in skus if s["classificacao"] == "estrela"),
        "saudavel": sum(1 for s in skus if s["classificacao"] == "saudavel"),
        "atencao":  sum(1 for s in skus if s["classificacao"] == "atencao"),
        "problema": sum(1 for s in skus if s["classificacao"] == "problema"),
        "em_colapso": sum(1 for s in skus if s["nivel_risco"] == "colapso_iminente"),
        "em_protecao": sum(1 for s in skus if s["nivel_risco"] == "protecao"),
    }

    return {"skus": skus, "resumo": resumo}


@app.post("/sie/score-config")
async def sie_score_config(request: Request):
    """
    Salva a configuração de ajuste de margem por SIE (usada pelo motor de precificação).
    Body: { ajuste_ativo, ajuste_estrela, ajuste_saudavel, ajuste_atencao, ajuste_problema }
    """
    body = await request.json()
    cfg = _load_json(_MOD_DIR / "sie_score_config.json", {})
    cfg.update(body)
    _save_json(_MOD_DIR / "sie_score_config.json", cfg)
    return {"ok": True, "config": cfg}


@app.get("/sie/score-config")
def get_sie_score_config():
    """Retorna a configuração atual de ajuste SIE → margem."""
    return _load_json(_MOD_DIR / "sie_score_config.json", {
        "ajuste_ativo": False,
        "ajuste_estrela": 0,
        "ajuste_saudavel": 0,
        "ajuste_atencao": 8,
        "ajuste_problema": 18,
    })
