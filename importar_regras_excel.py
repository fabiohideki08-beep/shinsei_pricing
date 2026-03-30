import json
from pathlib import Path

import openpyxl

ARQUIVO_EXCEL = "Simulador_2026_V8_beta.xlsx"
NOME_ABA = "Aba2"
ARQUIVO_SAIDA = Path("data") / "regras.json"

MAPA_CANAIS = {
    "Classico": "Mercado Livre Classico",
    "Premium": "Mercado Livre Premium",
    "Shopfy": "Shopify",
}


def para_float(valor, padrao=0.0):
    if valor is None or valor == "":
        return padrao
    try:
        return float(valor)
    except Exception:
        texto = str(valor).strip().replace(".", "").replace(",", ".")
        try:
            return float(texto)
        except Exception:
            return padrao


def normalizar_canal(canal):
    if canal is None:
        return ""
    texto = str(canal).strip()
    return MAPA_CANAIS.get(texto, texto)


def main():
    if not Path(ARQUIVO_EXCEL).exists():
        raise FileNotFoundError(
            f"Não encontrei o arquivo '{ARQUIVO_EXCEL}' dentro da pasta do projeto."
        )

    wb = openpyxl.load_workbook(ARQUIVO_EXCEL, data_only=True)

    if NOME_ABA not in wb.sheetnames:
        raise ValueError(f"A aba '{NOME_ABA}' não existe no arquivo Excel.")

    ws = wb[NOME_ABA]

    regras = []

    # Mapeamento esperado da Aba2:
    # A: canal
    # B: peso_min
    # C: peso_max
    # D: preco_min
    # E: preco_max
    # F: taxa_fixa
    # G: taxa_frete
    # H: comissao
    for row in ws.iter_rows(min_row=2, max_col=8, values_only=True):
        canal, peso_min, peso_max, preco_min, preco_max, taxa_fixa, taxa_frete, comissao = row

        canal = normalizar_canal(canal)
        if not canal:
            continue

        regras.append(
            {
                "canal": canal,
                "peso_min": para_float(peso_min),
                "peso_max": para_float(peso_max),
                "preco_min": para_float(preco_min),
                "preco_max": para_float(preco_max),
                "frete": para_float(taxa_frete),
                "comissao": para_float(comissao),
                "taxa_fixa": para_float(taxa_fixa),
            }
        )

    ARQUIVO_SAIDA.parent.mkdir(parents=True, exist_ok=True)
    with open(ARQUIVO_SAIDA, "w", encoding="utf-8") as f:
        json.dump(regras, f, ensure_ascii=False, indent=2)

    print(f"Importação concluída com sucesso.")
    print(f"Total de regras importadas: {len(regras)}")
    print(f"Arquivo gerado: {ARQUIVO_SAIDA}")


if __name__ == "__main__":
    main()